"""
ProQuote - Streamlit interface.

Run:  streamlit run app.py   (from the _ProQuote folder)

Workflow 2 (New Offer): type-ahead catalogue search auto-fills Brand/costs/
prices; live-editable grid recalculates Total Cost / T. Price / T. Price SAR
instantly; discount row + bottom totals block with the markup factor.
Workflow 3 (PDF): one click generates the client-facing Quotation PDF.
"""
from __future__ import annotations
import os
import io
import html
import importlib
import inspect
import zipfile
import datetime as dt

import pandas as pd
import streamlit as st

# Streamlit Community Cloud stores production credentials in its Secrets panel.
# Promote the database URL before importing the persistence modules so CLI/local
# runs can use DATABASE_URL while cloud runs use the same backend transparently.
if not os.environ.get("DATABASE_URL"):
    try:
        _secret_database_url = str(st.secrets.get("DATABASE_URL", "")).strip()
        if _secret_database_url:
            os.environ["DATABASE_URL"] = _secret_database_url
    except Exception:
        pass

import calc
import repo
import pdf_export
import auth
import db
import db_backup
import ingest
import reports
import audit as audit_log
import runtime_env
import updater
from version import APP_VERSION

# Streamlit reruns this file, but imported helper modules can stay cached in a
# long-lived app process (especially during Streamlit Cloud hot deployments).
# Reload calculation + persistence modules together so their function signatures
# and DB fields always match the currently deployed app.py.
calc = importlib.reload(calc)
db = importlib.reload(db)
repo = importlib.reload(repo)
audit_log = importlib.reload(audit_log)

_LOGO = db.banner_path()                       # per-company banner (follows BOQ_DATA_DIR)
_COMPANY = repo.get_setting("company_name") or "Company Name"
st.set_page_config(page_title=f"ProQuote - {_COMPANY}", layout="wide",
                   initial_sidebar_state="expanded")


def _request_scroll_top():
    st.session_state["_scroll_to_top"] = True


def _scroll_to_top_if_requested():
    if not st.session_state.pop("_scroll_to_top", False):
        return
    st.iframe(
        """
        <script>
        const scrollToTop = () => {
            const doc = window.parent.document;
            const targets = [
                doc.querySelector('[data-testid="stMain"]'),
                doc.querySelector('[data-testid="stAppViewContainer"]'),
                doc.scrollingElement
            ];
            targets.forEach((target) => {
                if (!target) return;
                target.scrollTop = 0;
                if (typeof target.scrollTo === 'function') target.scrollTo(0, 0);
            });
            window.parent.scrollTo(0, 0);
        };
        scrollToTop();
        requestAnimationFrame(scrollToTop);
        setTimeout(scrollToTop, 100);
        </script>
        """,
        height=1,
        width=1,
    )


def _choose_local_folder(initial_dir: str = "") -> tuple[str, str]:
    """Open a Windows folder picker on the machine running Streamlit."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askdirectory(initialdir=initial_dir or os.getcwd())
        root.destroy()
        return selected or "", ""
    except Exception as exc:
        return "", str(exc)


def _run_excel_import(import_root: str):
    """Safety-backup, ingest every workbook under import_root, and show the stats."""
    safety_backup = db_backup.create_profile_backup("before-import")
    st.info(f"Safety backup created before import: {os.path.basename(safety_backup)}")
    progress = st.progress(0, text="Scanning Excel workbooks...")

    def _p(done, total, path):
        if total:
            progress.progress(min(done / total, 1.0),
                              text=f"Importing {done}/{total}: {os.path.basename(path)}")
    try:
        with st.spinner("Importing Excel workbooks..."):
            stats = ingest.ingest_folder(import_root, progress=_p)
        progress.empty()
        st.success("Import completed.")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Workbooks found", stats.get("workbooks_found", 0))
        m2.metric("Files ingested", stats.get("files", 0))
        m3.metric("BOQ sheets", stats.get("sheets", 0))
        m4.metric("Catalogue items", stats.get("catalogue_items", 0))
        d1, d2, d3 = st.columns(3)
        d1.metric("Offer lines", stats.get("lines", 0))
        d2.metric("Spare items", stats.get("spares", 0))
        d3.metric("Skipped", stats.get("skipped_no_boq", 0))
        if stats.get("errors"):
            with st.expander(f"Import warnings / errors ({len(stats['errors'])})"):
                for name, err in stats["errors"][:30]:
                    st.write(f"- {name}: {err}")
    except Exception as exc:
        progress.empty()
        st.error(f"Import failed: {exc}")

# Larger button text + icons; tracking status cells are compact and centered.
st.markdown("""<style>
[data-testid="stMainBlockContainer"],
.main .block-container {
  padding-top: 2.5rem !important;
}
.stButton button,
.stDownloadButton button {
  font-size: 1.05rem;
  font-weight: 600;
  min-height: 2.9rem;
}
.stDownloadButton,
.stDownloadButton > div { width: 100%; }
.stDownloadButton button {
  width: 100%;
  height: 2.9rem;
  max-height: 2.9rem;
  box-sizing: border-box;
  display: flex;
  align-items: center;
  justify-content: center;
}
.stButton button p,
.stDownloadButton button p {
  font-size: 1.05rem;
  line-height: 1.2;
  margin: 0;
}
div[data-testid="stPopover"] button {
  font-size: 1.05rem;
  font-weight: 600;
  min-height: 2.9rem;
}
div[data-testid="stPopover"] button p { font-size: 1.05rem; }
[class*="st-key-ed_cancel"] button,
[class*="st-key-ed_cancel_modal_discard"] button {
  background: #dc2626 !important;
  border-color: #dc2626 !important;
  color: #fff !important;
}
[class*="st-key-ed_cancel"] button p,
[class*="st-key-ed_cancel_modal_discard"] button p {
  color: #fff !important;
}
[class*="st-key-offer_tab_"] button {
  min-height: 4.15rem !important;
  border-radius: 8px 8px 0 0 !important;
  border: 1px solid #cbd5df !important;
  border-bottom-width: 3px !important;
  font-size: 1.35rem !important;
  font-weight: 800 !important;
}
[class*="st-key-offer_tab_"] button p {
  font-size: 1.35rem !important;
  font-weight: 800 !important;
}
[class*="st-key-offer_tab_active_"] button {
  background: #17324d !important;
  border-color: #17324d !important;
  color: #fff !important;
}
[class*="st-key-offer_tab_active_"] button p { color: #fff !important; }
[class*="st-key-offer_tab_locked_"] button {
  background: #f1f4f7 !important;
  color: #8a96a3 !important;
  border-color: #d5dde5 !important;
  opacity: 0.72 !important;
}
[class*="st-key-offer_tab_locked_"] button p { color: #8a96a3 !important; }
[class*="st-key-trkbtn_"] {
  display: flex !important;
  justify-content: center !important;
}
[class*="st-key-trkbtn_"] button {
  min-height: 1.9rem !important;
  height: 1.9rem !important;
  max-height: 1.9rem !important;
  min-width: 2.15rem !important;
  max-width: 2.15rem !important;
  width: 2.15rem !important;
  padding: 0 !important;
  margin: 0 auto !important;
  border-radius: 7px !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
}
[class*="st-key-trkbtn_on_"] button {
  background: #2ea44f !important;
  border-color: #2ea44f !important;
  color: #fff !important;
}
[class*="st-key-trkbtn_partial_"] button {
  background: #f2b84b !important;
  border-color: #f2b84b !important;
  color: #fff !important;
}
[class*="st-key-trkbtn_off_"] button {
  background: #fff !important;
  border-color: #cfd6df !important;
  color: transparent !important;
}
[class*="st-key-trkbtn_"] button p {
  font-size: 1.3rem !important;
  line-height: 1 !important;
  margin: 0 !important;
}
[class*="st-key-trkbtn_on_"] button p { color: #fff !important; }
[class*="st-key-trkbtn_partial_"] button p { color: #fff !important; }
[class*="st-key-trkbtn_off_"] button p { color: transparent !important; }
[class*="st-key-po_"] input,
[class*="st-key-dn_"] input,
[class*="st-key-trkqty_"] input {
  min-height: 2.15rem !important;
  padding-top: 0.25rem !important;
  padding-bottom: 0.25rem !important;
}
[class*="st-key-trkqty_"] input {
  text-align: center !important;
  padding-left: 0.2rem !important;
  padding-right: 0.2rem !important;
}
.tracking-header {
  width: 100%;
  text-align: center;
  font-size: 0.98rem;
  font-weight: 600;
  color: #333;
}
.tracking-stamp {
  width: 100%;
  text-align: center;
  white-space: nowrap;
  font-size: 0.82rem;
  color: #7b8490;
  line-height: 1.2;
  min-height: 0.95rem;
  margin-top: 0.05rem;
}
.tracking-center-cell {
  width: 100%;
  text-align: center;
  line-height: 2.15rem;
}
.tracking-row-separator {
  width: 100%;
  border-top: 1px solid #e3e8ef;
  margin: 0.28rem 0 0.42rem 0;
}
[data-testid="stHorizontalBlock"]:has(.tracking-header),
[data-testid="stHorizontalBlock"]:has(.tracking-stamp) {
  gap: 0.4rem !important;
}
[class*="price_updated_"] input:disabled {
  color: #172033 !important;
  -webkit-text-fill-color: #172033 !important;
  opacity: 1 !important;
  font-weight: 500 !important;
}
[class*="price_updated_"] label p {
  color: #172033 !important;
}
</style>""", unsafe_allow_html=True)

# Full internal grid (builder always sees costs; the client PDF never shows costs).
BUILDER_COLS = ["Area", "System", "Description", "Brand", "Model", "Qty",
                "Cur", "List Price $", "Ex Unit Cost $", "Shipping %", "Unit Cost $", "Total Cost $",
                "Markup x", "U. Price $", "T. Price $", "U. Price SAR", "T. Price SAR"]
# Pure outputs - locked in the editor (everything else is an input/driver).
COMPUTED = ["Total Cost $", "T. Price $", "U. Price SAR", "T. Price SAR"]
MONEY_COLS = ["List Price $", "Ex Unit Cost $", "Unit Cost $", "Total Cost $",
              "U. Price $", "T. Price $", "U. Price SAR", "T. Price SAR"]
# Numeric inputs that affect computed prices - a change triggers one auto-rerun
# so the recomputed columns refresh immediately (no st.data_editor 1-step lag).
NUM_DRIVERS = ["Qty", "Ex Unit Cost $", "Shipping %", "Unit Cost $", "Markup x", "U. Price $"]
# Reviewing a loaded offer shows selling prices only - all cost columns hidden.
PRICE_VIEW_COLS = ["Area", "System", "Description", "Brand", "Model", "Qty",
                   "U. Price $", "T. Price $", "U. Price SAR", "T. Price SAR"]

# Offer terms/notes - keys match repo.TERMS_KEYS; defaults from the historical Quotation sheets.
TERMS_KEYS = repo.TERMS_KEYS
PROJECT_SHEET_KEYS = repo.PROJECT_SHEET_KEYS
DEFAULT_TERMS = {
    "subject": "",
    "greeting": ("Dear Sir,\n\nThank you for the opportunity to quote for the above-mentioned "
                 "project. Kindly find hereinafter our offer for your kind review."),
    "system_note": "",
    "scope": "Supply, Installation, Testing & Commissioning.",
    "exclusions": ("Cables, electrical wiring & conduits, back boxes, cabling and/or civil work. "
                   "Pulling cables."),
    "prerequisites": "Power must be available and fully operational prior to the start of our work.",
    "delivery": "8-10 weeks from date of receiving the down payment.",
    "payment": "70% Down payment, 20% upon delivery, 10% upon installation, testing & commissioning.",
    "validity": "30 Days from its date of issuance.",
    "notes": "",
}

PROJECT_LEAD_SOURCE_OPTIONS = [
    "Self Generated", "International Specs", "Hilights", "Lumiere Studio",
    "Follow-Up", "Selection/Alternative",
]
PROJECT_SHIPMENT_OPTIONS = ["Air", "Sea"]
DEFAULT_PROJECT_SHEET_INFO = {
    "job_reference": "",
    "sheet_date": "",
    "lead_source": "Self Generated",
    "commission": "",
    "shipment_by": "Air",
    "downpayment_date": "",
    "invoice_to": "",
    "delivery_instructions": "",
    "gm_signature": "",
}

# Which user-role each offer people-field is picked from.
PEOPLE_ROLES = {"sales": "sales", "presales": "Pre-Sales", "pm": "Project Manager"}
# The Sales Person picker spans several roles - anyone who may own/sign an offer.
SALES_PERSON_ROLES = ["sales", "Pre-Sales", "Project Manager", "Top Management"]


def _role_label(role) -> str:
    """Display a role name starting with a capital letter (cosmetic only - the stored
    role key is unchanged, so permissions / the protected 'owner' logic still work)."""
    r = str(role or "")
    return r[:1].upper() + r[1:]


def _ps_enabled() -> bool:
    """Project Sheet (info section + Excel export) is on unless disabled in Settings."""
    if "cached_ps_enabled" not in st.session_state:
        st.session_state.cached_ps_enabled = repo.get_setting("project_sheet_enabled") != "0"
    return st.session_state.cached_ps_enabled


def _inclusion_enabled() -> bool:
    """Installation Included pricing mode is off unless explicitly enabled in Settings."""
    if "cached_inclusion_enabled" not in st.session_state:
        st.session_state.cached_inclusion_enabled = (
            repo.get_setting("installation_inclusion_enabled") == "1"
        )
    return st.session_state.cached_inclusion_enabled


@st.fragment
def _catalog_dedupe_tool():
    """Find catalogue duplicates (same Model + Description) and delete chosen ones.
    A fragment, so ticking checkboxes doesn't re-run the whole catalogue page."""
    if st.button("🔎 Find duplicates", key="cat_find_dups"):
        st.session_state["cat_dups"] = repo.catalog_duplicates()
        st.rerun(scope="fragment")
    dups = st.session_state.get("cat_dups")
    if dups is None:
        st.caption("Click **Find duplicates** to scan the catalogue.")
        return
    if not dups:
        st.success("No duplicates found (by Model + Description).")
        return
    n_ident = sum(1 for g in dups if g["identical"])
    st.caption(f"{len(dups)} duplicate group(s) — {n_ident} identical. Identical groups "
               "pre-select all but the most-quoted copy; review/adjust, then delete.")
    widths = [0.5, 1.5, 1.3, 1.0, 1.0, 0.6, 1.1, 0.8]
    to_delete = []
    for g in dups:
        badge = "🟢 Identical" if g["identical"] else "🟠 Differs"
        st.markdown(f"**{badge}** · {g['model'] or '—'} · {g['description'] or '—'}  "
                    f"({len(g['items'])} copies)")
        items = sorted(g["items"], key=lambda it: -int(it.get("TimesQuoted") or 0))
        keep_id = items[0]["ItemID"]                      # default keep = most-quoted
        hc = st.columns(widths)
        for col, t in zip(hc, ["Del", "Brand", "Model", "List", "Ex", "Cur", "U.SAR", "Quoted"]):
            col.caption(t)
        for it in items:
            rc = st.columns(widths, vertical_alignment="center")
            default_del = g["identical"] and it["ItemID"] != keep_id
            if rc[0].checkbox("del", value=default_del, key=f"dupdel_{it['ItemID']}",
                              label_visibility="collapsed"):
                to_delete.append(int(it["ItemID"]))
            rc[1].write(str(it.get("Brand") or ""))
            rc[2].write(str(it.get("Model") or ""))
            rc[3].write(f"{calc._num(it.get('ListPriceUSD')):,.2f}")
            rc[4].write(f"{calc._num(it.get('ExUnitCostUSD')):,.2f}")
            rc[5].write(str(it.get("Currency") or ""))
            rc[6].write(f"{calc._num(it.get('DefaultUPriceSAR')):,.2f}")
            rc[7].write(str(int(it.get("TimesQuoted") or 0)))
        st.divider()
    if st.button(f"🗑️ Delete selected ({len(to_delete)})", type="primary",
                 key="cat_dup_del", disabled=not to_delete):
        repo.delete_catalog_items(to_delete)
        st.session_state.pop("cat_dups", None)
        st.toast(f"Deleted {len(to_delete)} duplicate item(s).", icon="✅")
        st.rerun()                                        # app-level: refresh the main grid


def _person_select(col, label, role, current, key, **widget_kwargs):
    """Dropdown of active users holding `role` (a single role name or a list of roles);
    keeps any legacy stored value selectable. Always exactly one '-' (unassign) entry."""
    raw = (auth.users_in_roles(role) if isinstance(role, (list, tuple, set))
           else auth.users_in_role(role))
    names = []
    for n in raw:                            # drop blank/placeholder entries and dedupe
        n = str(n or "").strip()
        if n and n != "-" and n not in names:
            names.append(n)
    cur = (current or "").strip()
    opts = ["-"] + names
    if cur and cur != "-" and cur not in names:
        opts = ["-", cur] + names            # preserve a real stored name not in the list
    initial = {} if key in st.session_state else {
        "index": opts.index(cur) if cur in opts else 0
    }
    pick = col.selectbox(label, opts, key=key, **initial, **widget_kwargs)
    return "" if pick == "-" else pick


def _region_select(col, current, key, **widget_kwargs):
    """Managed Region dropdown that keeps legacy stored values selectable."""
    regions = repo.regions()
    cur = _text(current).strip()
    widget_cur = _text(st.session_state.get(key)).strip()
    options = [""] + regions
    for legacy in (cur, widget_cur):
        if legacy and legacy not in options:
            options.append(legacy)
    selected = widget_cur if widget_cur in options else (cur if cur in options else "")
    initial = {} if key in st.session_state else {"index": options.index(selected)}
    return col.selectbox(
        "Region",
        options,
        key=key,
        format_func=lambda value: value or "-",
        help="Manage this dropdown in Settings → Company Details → Project regions.",
        **initial,
        **widget_kwargs,
    )


def _system_select(col, current, key, **widget_kwargs):
    """Full-name System dropdown; the configured abbreviation is shown alongside it."""
    names = repo.system_names()
    cur = repo.system_name(current)
    widget_cur = repo.system_name(st.session_state.get(key))
    options = [""] + names
    for legacy in (cur, widget_cur):
        if legacy and legacy not in options:
            options.append(legacy)
    selected = widget_cur if widget_cur in options else (cur if cur in options else "")

    def _label(name):
        if not name:
            return "-"
        abbreviation = repo.system_abbreviation(name)
        return f"{name} - {abbreviation}" if abbreviation and abbreviation != name else name

    initial = {} if key in st.session_state else {"index": options.index(selected)}
    return col.selectbox(
        "System",
        options,
        key=key,
        format_func=_label,
        help=("The full system name is stored on the project; its abbreviation replaces "
              "*TYPE* in the Offer #. Manage systems in Settings → Company Details."),
        **initial,
        **widget_kwargs,
    )


def _report_filter_options(df, column):
    """Return data values plus everyone currently assignable as Sales."""
    values = {
        str(value).strip()
        for value in df[column].dropna().unique()
        if str(value).strip() and str(value).strip() != "-"
    }
    if column == "Sales Person":
        values.update(
            str(name).strip()
            for name in auth.users_in_roles(SALES_PERSON_ROLES)
            if str(name).strip() and str(name).strip() != "-"
        )
    return sorted(values, key=str.casefold)


def _report_filter_label(column):
    return "Assigned as Sales" if column == "Sales Person" else column


def _project_person_filter_options(projects, column, active_names=()):
    """Stored project assignees plus active users valid for that people field."""
    values = {
        str(value).strip()
        for value in projects[column].dropna().unique()
        if str(value).strip() and str(value).strip() != "-"
    }
    values.update(
        str(name).strip()
        for name in active_names
        if str(name).strip() and str(name).strip() != "-"
    )
    return sorted(values, key=str.casefold)


def _empty_grid() -> pd.DataFrame:
    df = pd.DataFrame([calc.blank_row()])
    df["LineType"] = "item"
    df["_ItemID"] = None
    return df.iloc[0:0]


def _default_system():
    names = repo.system_names()
    legacy_default = repo.system_name("LCS")
    return legacy_default if legacy_default in names else (names[0] if names else "")


def _ensure_state():
    if "grid" not in st.session_state:
        st.session_state.grid = _empty_grid()
    if "header" not in st.session_state:
        st.session_state.header = {
            **DEFAULT_TERMS,
            "client": "", "project": "", "contact": "", "phone": "",
            "contractor": "", "region": "",
            "sales": "", "presales": "", "pm": "",
            "option": "",
            "offer": _next_offer_no(), "system": _default_system(),
            "date": dt.date.today().isoformat(), "margin": 1.60,
            "project_sheet": dict(DEFAULT_PROJECT_SHEET_INFO),
        }
    if "discount" not in st.session_state:
        st.session_state.discount = 0.0
    if "commission" not in st.session_state:
        st.session_state.commission = 0.0


def _new_offer_header(overrides: dict | None = None) -> dict:
    header = {
        **DEFAULT_TERMS,
        "client": "", "project": "", "contact": "", "phone": "",
        "contractor": "", "region": "",
        "sales": "", "presales": "", "pm": "",
        "offer": _next_offer_no(), "system": _default_system(),
        "date": dt.date.today().isoformat(), "margin": 1.60,
        "project_sheet": dict(DEFAULT_PROJECT_SHEET_INFO),
        "inclusion_mode": "excluded",
    }
    if overrides:
        header.update(overrides)
    return header


def _prime_new_offer_form(header: dict | None = None, grid: pd.DataFrame | None = None,
                          discount: float = 0.0, commission: float = 0.0,
                          commission_percent: float = 0.0,
                          commission_mode: str = "Protect profit"):
    """Load data into the New Offer form before its widgets are rendered."""
    h = _new_offer_header(header or {})
    h["project_sheet"] = {**DEFAULT_PROJECT_SHEET_INFO, **(h.get("project_sheet") or {})}
    st.session_state.header = h
    st.session_state.grid = grid.copy() if grid is not None else _empty_grid()
    st.session_state.discount = abs(float(discount or 0.0))
    st.session_state.commission = abs(float(commission or 0.0))
    st.session_state.no_commission_percent = max(float(commission_percent or 0.0), 0.0)
    st.session_state.no_commission_mode = (
        commission_mode if commission_mode in ("Protect profit", "Deduct from profit")
        else "Deduct from profit"
    )
    st.session_state.no_commission_applied_percent = st.session_state.no_commission_percent
    if st.session_state.no_commission_mode != "Protect profit":
        st.session_state.no_commission_applied_percent = 0.0
    st.session_state.no_offer_lock = None
    st.session_state.no_saved_options = []
    for key in ("editor", "pdf_bytes", "project_sheet_bytes", "saved_rev"):
        st.session_state.pop(key, None)
    for key in (
        "no_discount_percent", "no_discount_driver", "no_discount_subtotal",
        "no_commission_driver", "no_commission_subtotal",
        "no_commission_base_subtotal",
    ):
        st.session_state.pop(key, None)

    st.session_state["no_client"] = h.get("client", "")
    st.session_state["no_project"] = h.get("project", "")
    st.session_state["no_contact"] = h.get("contact", "")
    st.session_state["no_phone"] = h.get("phone", "")
    st.session_state["no_contractor"] = h.get("contractor", "")
    st.session_state["no_region"] = h.get("region", "")
    st.session_state["no_sales"] = h.get("sales", "")
    st.session_state["no_presales"] = h.get("presales", "")
    st.session_state["no_pm"] = h.get("pm", "")
    st.session_state["no_offer_ov"] = ""
    st.session_state["no_option"] = h.get("option", "")

    system = repo.system_name(h.get("system"))
    st.session_state["no_offer_type"] = system

    term_keys = {
        "subject": "no_subject", "greeting": "no_greet",
        "scope": "no_scope", "exclusions": "no_excl", "prerequisites": "no_prereq",
        "delivery": "no_deliv", "validity": "no_valid", "payment": "no_pay",
        "notes": "no_notes",
    }
    for src, key in term_keys.items():
        st.session_state[key] = h.get(src, DEFAULT_TERMS.get(src, ""))

    project_sheet_keys = {
        "job_reference": "no_ps_job_reference",
        "sheet_date": "no_ps_sheet_date",
        "lead_source": "no_ps_lead_source",
        "commission": "no_ps_commission",
        "shipment_by": "no_ps_shipment_by",
        "downpayment_date": "no_ps_downpayment_date",
        "invoice_to": "no_ps_invoice_to",
        "delivery_instructions": "no_ps_delivery_instructions",
        "gm_signature": "no_ps_gm_signature",
    }
    ps_info = h.get("project_sheet") or {}
    for src, key in project_sheet_keys.items():
        st.session_state[key] = ps_info.get(src, DEFAULT_PROJECT_SHEET_INFO.get(src, ""))


def _next_offer_no() -> str:
    yr = dt.date.today().strftime("%y")
    return f"OFR-SWS-RUH-{yr}-NEW"


def _add_row_to(state_key: str, row: dict):
    g = st.session_state[state_key]
    st.session_state[state_key] = pd.concat([g, pd.DataFrame([row])], ignore_index=True)


def _add_row(row: dict):
    _add_row_to("grid", row)


def _text(value, default: str = "") -> str:
    """Display-safe text for DB/pandas values, treating None/NaN as empty."""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return text if text else default


def _fmt_date(value, default: str = "-") -> str:
    """Show a stored ISO date (YYYY-MM-DD...) as DD-MM-YYYY."""
    s = _text(value)[:10]
    parts = s.split("-")
    if len(parts) == 3 and len(parts[0]) == 4:
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return s or default


def _fmt_month_year(value, default: str = "01-2025") -> str:
    """Show a stored date as MM-YYYY."""
    text = _text(value)
    if not text:
        return default
    try:
        if isinstance(value, (dt.datetime, dt.date)):
            return value.strftime("%m-%Y")
        return dt.date.fromisoformat(text[:10]).strftime("%m-%Y")
    except (TypeError, ValueError):
        parts = text.replace("/", "-").split("-")
        if len(parts) >= 2 and len(parts[0]) == 2 and len(parts[1]) == 4:
            return f"{parts[0]}-{parts[1]}"
        if len(parts) >= 2 and len(parts[0]) == 4:
            return f"{parts[1].zfill(2)}-{parts[0]}"
        return default


def _ctr(col, text, header: bool = False):
    """Write center-aligned text into a column (header = small grey caption style)."""
    style = ("text-align:center;font-size:0.8rem;color:#808495" if header
             else "text-align:center")
    col.markdown(f"<div style='{style}'>{text}</div>", unsafe_allow_html=True)


def _builder_column_order(editor_key: str, host=st, width="content", show_include: bool = False):
    # The app owns this preference so hidden columns stay hidden after editor reruns.
    with host.popover("⚙ Columns", help="Show / hide columns - your choice sticks across edits.",
                      width=width):
        st.markdown("<div style='min-width:320px'></div>", unsafe_allow_html=True)
        cc = st.columns(2)
        all_cols = (["_IncludedInItems"] if show_include else []) + list(BUILDER_COLS)
        labels = {"_IncludedInItems": "Include"}
        visible = [col for i, col in enumerate(all_cols)
                   if cc[i % 2].checkbox(labels.get(col, col), value=True,
                                         key=f"{editor_key}_show_{col}")]
    base = tuple(c for c in BUILDER_COLS if c in visible) or tuple(BUILDER_COLS)
    if show_include and "_IncludedInItems" in visible:
        return ("_IncludedInItems",) + base
    return base


def _editor_full_height(row_count: int) -> int:
    return 76 + max(row_count, 1) * 35


EDIT_WIDGET_KEYS = (
    "edit_grid", "edit_key", "edit_pid", "edit_system", "edit_terms", "edit_header",
    "edit_project_sheet", "edit_discount", "edit_dirty_snapshot", "edit_show_cancel_dialog",
    "edit_close_after_save", "pending_close_edit", "pending_save", "pending_option_label",
    "edit_commission", "edit_editor", "ed_option", "ed_discount_percent",
    "ed_discount_driver", "ed_discount_subtotal", "ed_commission_percent",
    "ed_commission_driver", "ed_commission_subtotal", "ed_commission_mode",
    "ed_commission_applied_percent",
    "ed_commission_base_subtotal", "eh_client", "eh_project", "eh_contact",
    "eh_phone", "eh_contractor", "eh_region", "eh_system", "eh_sales", "eh_presales", "eh_pm",
    "ed_subject", "ed_greet", "ed_scope", "ed_excl", "ed_prereq",
    "ed_deliv", "ed_valid", "ed_pay", "ed_notes",
    "ed_ps_job_reference", "ed_ps_sheet_date", "ed_ps_lead_source", "ed_ps_commission",
    "ed_ps_shipment_by", "ed_ps_downpayment_date", "ed_ps_invoice_to",
    "ed_ps_delivery_instructions", "ed_ps_gm_signature",
    "ed_inclusion_mode",
    "_edit_just_saved",
)


def _snapshot_value(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    return str(value).strip()


def _grid_snapshot(grid: pd.DataFrame) -> list[dict]:
    if grid is None:
        return []
    df = calc.recompute(grid).reset_index(drop=True).copy()
    cols = [c for c in BUILDER_COLS + ["LineType", "_ItemID", "_IncludedInItems"] if c in df.columns]
    return [
        {col: _snapshot_value(row.get(col)) for col in cols}
        for row in df[cols].to_dict("records")
    ]


def _dict_snapshot(values: dict | None) -> dict:
    values = values or {}
    return {key: _snapshot_value(values.get(key)) for key in sorted(values)}


def _edit_snapshot() -> dict:
    return {
        "grid": _grid_snapshot(st.session_state.get("edit_grid")),
        "system": _snapshot_value(st.session_state.get("edit_system")),
        "terms": _dict_snapshot(st.session_state.get("edit_terms")),
        "header": _dict_snapshot(st.session_state.get("edit_header")),
        "project_sheet": _dict_snapshot(st.session_state.get("edit_project_sheet")),
        "discount": _snapshot_value(st.session_state.get("edit_discount")),
        "commission": _snapshot_value(st.session_state.get("edit_commission")),
        "commission_percent": _snapshot_value(st.session_state.get("ed_commission_percent")),
        "commission_mode": _snapshot_value(st.session_state.get("ed_commission_mode")),
        "option": _snapshot_value(st.session_state.get("ed_option")),
        "inclusion_mode": _snapshot_value(st.session_state.get("ed_inclusion_mode")),
    }


def _sync_edit_state_from_widgets():
    if "edit_header" in st.session_state:
        header_keys = {
            "client": "eh_client", "project": "eh_project", "contact": "eh_contact",
            "phone": "eh_phone", "contractor": "eh_contractor", "region": "eh_region",
            "sales": "eh_sales", "presales": "eh_presales", "pm": "eh_pm",
        }
        for field, key in header_keys.items():
            if key in st.session_state:
                value = st.session_state.get(key)
                if field in ("sales", "presales", "pm") and value == "-":
                    value = ""
                st.session_state.edit_header[field] = value

    if "eh_system" in st.session_state:
        st.session_state.edit_system = repo.system_name(st.session_state.get("eh_system"))

    if "edit_terms" in st.session_state:
        term_keys = {
            "subject": "ed_subject", "greeting": "ed_greet",
            "scope": "ed_scope", "exclusions": "ed_excl", "prerequisites": "ed_prereq",
            "delivery": "ed_deliv", "validity": "ed_valid", "payment": "ed_pay",
            "notes": "ed_notes",
        }
        for field, key in term_keys.items():
            if key in st.session_state:
                st.session_state.edit_terms[field] = st.session_state.get(key)
        st.session_state.edit_terms["system_note"] = st.session_state.get("edit_system", "")

    if "edit_project_sheet" in st.session_state:
        ps_keys = {
            "job_reference": "ed_ps_job_reference", "sheet_date": "ed_ps_sheet_date",
            "lead_source": "ed_ps_lead_source", "commission": "ed_ps_commission",
            "shipment_by": "ed_ps_shipment_by", "downpayment_date": "ed_ps_downpayment_date",
            "invoice_to": "ed_ps_invoice_to",
            "delivery_instructions": "ed_ps_delivery_instructions",
            "gm_signature": "ed_ps_gm_signature",
        }
        for field, key in ps_keys.items():
            if key in st.session_state:
                st.session_state.edit_project_sheet[field] = st.session_state.get(key)


def _queue_edit_save(action: str):
    _sync_edit_state_from_widgets()
    st.session_state.pending_save = action
    st.session_state.pending_option_label = (
        st.session_state.get("ed_option") or ""
    ).strip()


def _sync_new_header_from_widgets():
    if "header" not in st.session_state:
        return

    h = st.session_state.header
    header_keys = {
        "client": "no_client", "project": "no_project", "contact": "no_contact",
        "phone": "no_phone", "contractor": "no_contractor", "region": "no_region",
        "sales": "no_sales", "presales": "no_presales", "pm": "no_pm",
        "option": "no_option",
    }
    for field, key in header_keys.items():
        if key in st.session_state:
            h[field] = st.session_state.get(key)

    if "no_offer_type" in st.session_state:
        h["system"] = st.session_state.get("no_offer_type") or ""
    if "no_offer_ov" in st.session_state:
        h["offer_override"] = (st.session_state.get("no_offer_ov") or "").strip()
    if "no_offer_lock" in st.session_state or "no_offer_type" in st.session_state:
        h["offer"] = (
            st.session_state.get("no_offer_lock")
            or h.get("offer_override")
            or repo.make_offer_no(h.get("system", ""))
        )

    term_keys = {
        "subject": "no_subject", "greeting": "no_greet",
        "scope": "no_scope", "exclusions": "no_excl", "prerequisites": "no_prereq",
        "delivery": "no_deliv", "validity": "no_valid", "payment": "no_pay",
        "notes": "no_notes",
    }
    for field, key in term_keys.items():
        if key in st.session_state:
            h[field] = st.session_state.get(key)
    h["system_note"] = h.get("system", "")

    ps = h.setdefault("project_sheet", dict(DEFAULT_PROJECT_SHEET_INFO))
    ps_keys = {
        "job_reference": "no_ps_job_reference", "sheet_date": "no_ps_sheet_date",
        "lead_source": "no_ps_lead_source", "commission": "no_ps_commission",
        "shipment_by": "no_ps_shipment_by", "downpayment_date": "no_ps_downpayment_date",
        "invoice_to": "no_ps_invoice_to",
        "delivery_instructions": "no_ps_delivery_instructions",
        "gm_signature": "no_ps_gm_signature",
    }
    for field, key in ps_keys.items():
        if key in st.session_state:
            ps[field] = st.session_state.get(key)
    h["project_sheet"] = ps
    st.session_state.header = h


def _copy_widget_to_state_dict(state_key: str, field: str, widget_key: str):
    target = st.session_state.get(state_key)
    if isinstance(target, dict) and widget_key in st.session_state:
        target[field] = st.session_state.get(widget_key)
        st.session_state[state_key] = target


def _copy_edit_header_widget(field: str, widget_key: str, blank_marker=None):
    """Immediately copy an Edit Project Details widget into the save payload."""
    if "edit_header" not in st.session_state or widget_key not in st.session_state:
        return
    value = st.session_state.get(widget_key)
    if blank_marker is not None and value == blank_marker:
        value = ""
    st.session_state.edit_header[field] = value


def _copy_edit_system_widget():
    """Immediately synchronize the Edit System selector and its dependent note."""
    system = repo.system_name(st.session_state.get("eh_system"))
    st.session_state.edit_system = system
    if "edit_terms" in st.session_state:
        st.session_state.edit_terms["system_note"] = system


def _clear_offer_data_caches():
    """Ensure a just-saved offer is re-read when the editor closes or the page reruns."""
    for name in ("_cached_project_index", "_cached_project_grid", "_cached_offers_df",
                 "_cached_finance_df", "_cached_report_dataset", "_cached_project_bundle"):
        cached = globals().get(name)
        if cached is not None and hasattr(cached, "clear"):
            cached.clear()


def _edit_has_unsaved_changes() -> bool:
    _sync_edit_state_from_widgets()
    original = st.session_state.get("edit_dirty_snapshot")
    return bool(original) and _edit_snapshot() != original


def _mark_edit_clean():
    _sync_edit_state_from_widgets()
    st.session_state.edit_dirty_snapshot = _edit_snapshot()
    st.session_state.edit_show_cancel_dialog = False
    st.session_state.edit_close_after_save = False
    st.session_state["_edit_just_saved"] = True


def _clear_edit_widget_state():
    """Remove every per-offer edit value so it cannot leak into another edit."""
    for key in EDIT_WIDGET_KEYS:
        st.session_state.pop(key, None)


def _init_edit_state(pid: int, meta: dict, grid, sheet):
    """Populate all session-state keys needed to open the edit panel for a project."""
    _clear_edit_widget_state()
    st.session_state.edit_grid = calc.recompute(grid.copy())
    st.session_state.edit_key = f"{pid}::{sheet}"
    st.session_state.edit_pid = pid
    _loaded_system = (repo.base_name(sheet or "").replace("BOQ", "").strip() or _default_system())
    st.session_state.edit_system = repo.system_name(_loaded_system)
    st.session_state.edit_discount = abs(float(meta.get("DiscountAmount") or 0))
    st.session_state.edit_commission = abs(float(meta.get("CommissionAmount") or 0))
    st.session_state.ed_commission_percent = abs(float(meta.get("CommissionPercent") or 0))
    st.session_state.ed_commission_mode = (
        meta.get("CommissionMode") if meta.get("CommissionMode") in COMMISSION_MODES
        else "Deduct from profit"
    )
    st.session_state.ed_commission_applied_percent = (
        st.session_state.ed_commission_percent
        if st.session_state.ed_commission_mode == "Protect profit" else 0.0
    )
    st.session_state.edit_terms = {**DEFAULT_TERMS, **repo.load_terms(meta)}
    st.session_state.edit_terms["system_note"] = st.session_state.edit_system
    st.session_state.edit_header = {
        "client": _text(meta.get("ClientName")), "project": _text(meta.get("ProjectName")),
        "contact": _text(meta.get("ContactName")), "phone": _text(meta.get("ContactPhone")),
        "contractor": _text(meta.get("Contractor")), "region": _text(meta.get("Region")),
        "sales": _text(meta.get("SalesPerson")),
        "presales": _text(meta.get("PresalesEngineer")),
        "pm": _text(meta.get("ProjectManager")),
    }
    for field, widget_key in {
        "client": "eh_client", "project": "eh_project", "contact": "eh_contact",
        "phone": "eh_phone", "contractor": "eh_contractor", "region": "eh_region",
        "sales": "eh_sales", "presales": "eh_presales", "pm": "eh_pm",
    }.items():
        value = st.session_state.edit_header.get(field, "")
        st.session_state[widget_key] = value if value else (
            "-" if field in ("sales", "presales", "pm") else ""
        )
    st.session_state["eh_system"] = st.session_state.edit_system
    st.session_state.edit_project_sheet = {
        **DEFAULT_PROJECT_SHEET_INFO, **repo.load_project_sheet_info(meta)
    }
    for src, key in {
        "job_reference": "ed_ps_job_reference", "sheet_date": "ed_ps_sheet_date",
        "lead_source": "ed_ps_lead_source", "commission": "ed_ps_commission",
        "shipment_by": "ed_ps_shipment_by", "downpayment_date": "ed_ps_downpayment_date",
        "invoice_to": "ed_ps_invoice_to",
        "delivery_instructions": "ed_ps_delivery_instructions",
        "gm_signature": "ed_ps_gm_signature",
    }.items():
        st.session_state[key] = st.session_state.edit_project_sheet.get(
            src, DEFAULT_PROJECT_SHEET_INFO.get(src, ""))
    st.session_state["ed_option"] = meta.get("OptionLabel") or ""
    # Read inclusion fields directly from DB to bypass any stale cache in meta
    _incl_meta = repo.project_meta(pid)
    st.session_state["ed_inclusion_mode"] = _incl_meta.get("InclusionMode") or "excluded"
    for k in ("ed_discount_percent", "ed_discount_driver", "ed_discount_subtotal"):
        st.session_state.pop(k, None)
    st.session_state.edit_dirty_snapshot = _edit_snapshot()
    st.session_state.edit_show_cancel_dialog = False
    st.session_state.edit_close_after_save = False
    st.session_state.edit_mode = True
    for k in ("pdf_bytes", "project_sheet_bytes", "saved_rev",
              "saved_export_header", "saved_export_grid", "saved_export_summary"):
        st.session_state.pop(k, None)


def _close_edit_mode():
    st.session_state.edit_mode = False
    _request_scroll_top()
    _clear_edit_widget_state()
    for key in (
        "pdf_bytes", "project_sheet_bytes", "saved_rev",
        "saved_export_header", "saved_export_grid", "saved_export_summary",
    ):
        st.session_state.pop(key, None)
    # If we were editing a just-saved offer, navigate to the load view so the
    # user lands on the offer page (not the empty new-offer form).
    _fam = st.session_state.pop("_just_saved_fam", None)
    _pid = st.session_state.pop("_just_saved_pid", None)
    st.session_state.pop("_just_saved_meta", None)
    if _fam:
        for _k in ("load_search_snapshot", "load_search_name", "load_search_offer",
                   "load_filter_sales", "load_filter_presales", "load_filter_pm"):
            st.session_state.pop(_k, None)
        st.session_state["load_fam"] = _fam
        if _pid:
            st.session_state["view_pid"] = _pid
        st.session_state["project_workspace_view"] = "load"


@st.dialog("Unsaved changes", width="small", dismissible=False)
def _cancel_edit_dialog():
    _, xcol = st.columns([8, 1])
    if xcol.button("X", key="ed_cancel_modal_back", width="stretch",
                   help="Back to editor"):
        st.session_state.edit_show_cancel_dialog = False
        st.rerun()

    st.warning("You have unsaved changes in this edit.")
    st.write("Save this revision before closing, or close the editor without saving.")
    save_col, close_col = st.columns(2)
    if save_col.button("💾 Save", type="primary", key="ed_cancel_modal_save",
                       width="stretch"):
        _queue_edit_save("this")
        st.session_state.edit_close_after_save = True
        st.session_state.edit_show_cancel_dialog = False
        st.rerun()
    if close_col.button("X Close without saving", key="ed_cancel_modal_discard",
                        width="stretch"):
        _close_edit_mode()
        st.rerun()


@st.dialog("Confirm approval", width="small", dismissible=False)
def _approve_offer_dialog(project_id: int, offer_label: str):
    _, xcol = st.columns([8, 1])
    if xcol.button("X", key="approve_modal_close", width="stretch",
                   help="Close without approving"):
        st.rerun()

    st.write("Are you sure you want to approve this offer?")
    st.info(offer_label or f"Offer #{project_id}")
    st.caption("Approving it will archive any other revision or option in the same offer family.")

    yes_col, no_col = st.columns(2)
    if yes_col.button("Yes", type="primary", key="approve_modal_yes", width="stretch"):
        archived = repo.approve_offer(project_id)
        st.cache_data.clear()
        st.toast(
            f"Approved. {archived} other entr{'y' if archived == 1 else 'ies'} archived."
            if archived else "Approved.",
            icon="✅",
        )
        st.rerun()
    if no_col.button("No", key="approve_modal_no", width="stretch"):
        st.rerun()


def render_editable_grid(state_key: str, editor_key: str, in_fragment: bool = False,
                         column_order=None, show_included_col: bool = False):
    """Full editable grid (all columns) with live recompute.

    Pre-applies the pending delta before rendering so computed columns are correct in
    the single auto-rerun that data_editor already fires — no second rerun needed."""
    original_base = st.session_state[state_key].reset_index(drop=True)
    if "_RowOrder" not in original_base.columns:
        original_base["_RowOrder"] = range(1, len(original_base) + 1)
    n_base = len(original_base)

    colcfg = {c: st.column_config.NumberColumn(c, format="accounting") for c in MONEY_COLS}
    colcfg["Qty"] = st.column_config.NumberColumn("Qty", format="%d", min_value=0)
    colcfg["Cur"] = st.column_config.SelectboxColumn(
        "Cur", options=calc.CURRENCIES, required=False, width="small",
        help="Currency of List Price & Ex Unit Cost. The Unit Cost is converted to "
             "USD automatically (EUR rate from Settings; SAR pegged at 3.75).")
    # List Price & Ex Unit Cost are in the row's chosen currency (drop the misleading $).
    colcfg["List Price $"] = st.column_config.NumberColumn("List Price", format="accounting")
    colcfg["Ex Unit Cost $"] = st.column_config.NumberColumn("Ex Unit Cost", format="accounting")
    colcfg["Unit Cost $"] = st.column_config.NumberColumn(
        "Unit Cost (USD)", format="accounting",
        help="⚡ Conditional: editable only when Ex Unit Cost is blank. "
             "If Ex Unit Cost is filled, Unit Cost is computed automatically.")
    colcfg["Shipping %"] = st.column_config.NumberColumn(
        "Shipping %", format="%.2f", min_value=0.0, step=5.0,
        help="Added to Ex Unit Cost. Unit Cost = Ex Unit Cost × (1 + Shipping% / 100), in USD.")
    colcfg["Markup x"] = st.column_config.NumberColumn(
        "Markup x", format="plain", min_value=0.0, step=None,
        help="Multiplier on Unit Cost → U. Price $. Set to 0 to enter U. Price $ manually.")
    colcfg["U. Price $"] = st.column_config.NumberColumn(
        "U. Price $", format="accounting",
        help="⚡ Conditional: editable only when Markup x = 0. "
             "If Markup x > 0, this is computed as ⌈Unit Cost × Markup x⌉.")
    colcfg["Total Cost $"] = st.column_config.NumberColumn(
        "Total Cost $", format="accounting",
        help="🔒 Computed: Qty × Unit Cost $. Cannot be edited.")
    colcfg["T. Price $"] = st.column_config.NumberColumn(
        "T. Price $", format="accounting",
        help="🔒 Computed: Qty × U. Price $. Cannot be edited.")
    colcfg["U. Price SAR"] = st.column_config.NumberColumn(
        "U. Price SAR", format="accounting",
        help="🔒 Computed: U. Price $ × 3.75, rounded up to next multiple of 10. Cannot be edited.")
    colcfg["T. Price SAR"] = st.column_config.NumberColumn(
        "T. Price SAR", format="accounting",
        help="🔒 Computed: Qty × U. Price SAR. Cannot be edited.")
    if show_included_col:
        colcfg["_IncludedInItems"] = st.column_config.CheckboxColumn(
            "Include", default=False,
            help="When checked, this row's cost is distributed proportionally to other item prices.")
    colcfg["_RowOrder"] = st.column_config.NumberColumn(
        "#", min_value=1, step=1, format="%d",
        help="Row position. Change the number to reorder; lower = higher up.")
    _base_order = column_order if column_order is not None else _builder_column_order(editor_key)
    order = ("_RowOrder",) + tuple(c for c in _base_order if c != "_RowOrder")

    # Read the delta BEFORE data_editor processes it.
    raw_delta = st.session_state.get(editor_key)
    delta = raw_delta if isinstance(raw_delta, dict) else {}
    _structural = bool(delta.get("added_rows") or delta.get("deleted_rows"))

    if _structural:
        # Structural change (add/delete row): reconstruct from base + delta, clear delta, rerun.
        rec = original_base.copy()
        for row_idx, changes in delta.get("edited_rows", {}).items():
            idx = int(row_idx)
            for col, val in changes.items():
                if col in rec.columns and idx < len(rec):
                    rec.at[idx, col] = val
        for idx in sorted([int(i) for i in delta.get("deleted_rows", [])], reverse=True):
            rec = rec.drop(index=idx, errors="ignore")
        rec = rec.reset_index(drop=True)
        for new_row in delta.get("added_rows", []):
            blank = calc.blank_row()
            blank.update({k: v for k, v in new_row.items() if k in blank})
            blank["_RowOrder"] = len(rec) + 1
            rec = pd.concat([rec, pd.DataFrame([blank])], ignore_index=True)
        rec["_RowOrder"] = range(1, len(rec) + 1)
        new_grid = calc.recompute(rec)
        st.session_state[state_key] = new_grid
        st.session_state.pop(editor_key, None)
        st.rerun(scope="fragment") if in_fragment else st.rerun()

    # Pre-apply any pending cell edits to the base so the recomputed columns
    # (Total SAR, U.Price etc.) are already correct when data_editor renders.
    # Streamlit re-applies the same delta on top, which is idempotent for cell edits
    # (absolute values, not deltas), so no second rerun is needed.
    working = original_base.copy()
    for row_idx, changes in delta.get("edited_rows", {}).items():
        idx = int(row_idx)
        for col, val in changes.items():
            if col in working.columns and idx < len(working):
                working.at[idx, col] = val
    n = len(working)
    working["LineType"] = ["discount" if str(d).strip().lower() == "discount" else "item"
                           for d in working.get("Description", pd.Series([""] * n))]
    working["_ItemID"] = [original_base["_ItemID"].iloc[i] if (i < n_base and "_ItemID" in original_base.columns)
                          else None for i in range(n)]
    if "_IncludedInItems" not in working.columns:
        working["_IncludedInItems"] = False
    else:
        working["_IncludedInItems"] = working["_IncludedInItems"].fillna(False).astype(bool)
    if show_included_col:
        incl_mask = working["_IncludedInItems"].astype(bool)
        working.loc[incl_mask, "Markup x"] = 0.0
    if "_RowOrder" not in working.columns:
        working["_RowOrder"] = range(1, len(working) + 1)
    else:
        working = working.sort_values("_RowOrder", kind="stable").reset_index(drop=True)
        working["_RowOrder"] = range(1, len(working) + 1)
    working = calc.recompute(working)
    # Store now so the Totals section rendered after this call sees the updated values.
    st.session_state[state_key] = working

    display_cols = BUILDER_COLS + (["_IncludedInItems"] if show_included_col else []) + ["_RowOrder"]
    display_grid = working[display_cols] if not working.empty else working
    edited = st.data_editor(
        display_grid,
        column_config=colcfg, disabled=[c for c in COMPUTED if c in display_cols],
        column_order=order,
        num_rows="dynamic", width="stretch", height=_editor_full_height(len(display_grid)),
        row_height=35, key=editor_key, hide_index=True,
    ).reset_index(drop=True)

    # Incorporate what data_editor actually returned (handles any edge-case divergence
    # between our pre-apply and Streamlit's delta re-application; idempotent for normal edits).
    n = len(edited)
    edited["LineType"] = ["discount" if str(d).strip().lower() == "discount" else "item"
                          for d in edited.get("Description", pd.Series([""] * n))]
    edited["_ItemID"] = [original_base["_ItemID"].iloc[i] if (i < n_base and "_ItemID" in original_base.columns)
                         else None for i in range(n)]
    # Restore _IncludedInItems when it wasn't part of the displayed columns.
    if "_IncludedInItems" not in edited.columns:
        m = len(working)
        edited["_IncludedInItems"] = [
            bool(working["_IncludedInItems"].iloc[i]) if i < m else False
            for i in range(n)
        ]
    # _RowOrder from edited (user may have changed it); fall back to working.
    if "_RowOrder" in edited.columns:
        edited["_RowOrder"] = pd.to_numeric(edited["_RowOrder"], errors="coerce").fillna(0).astype(int)
    else:
        m = len(working)
        edited["_RowOrder"] = [int(working["_RowOrder"].iloc[i]) if i < m else n + 1 for i in range(n)]
    final_grid = calc.recompute(edited)
    st.session_state[state_key] = final_grid
    return final_grid


@st.fragment
def _edit_panel(meta):
    # -------------------- EDIT: all columns -> new revision OR new option --------------------
    base = meta.get("BaseName") or repo.base_name(meta.get("ProjectName") or "Offer")
    src_rev = int(meta.get("RevisionNo") or 0)
    _nrk = f"_nextrev_{base}"
    if _nrk not in st.session_state:
        st.session_state[_nrk] = repo.next_revision(base)
    nextrev = st.session_state[_nrk]
    if "edit_terms" not in st.session_state:
        st.session_state.edit_terms = {**DEFAULT_TERMS, **repo.load_terms(meta)}
    if "edit_header" not in st.session_state:
        st.session_state.edit_header = {
            "client": _text(meta.get("ClientName")), "project": _text(meta.get("ProjectName")),
            "contact": _text(meta.get("ContactName")), "phone": _text(meta.get("ContactPhone")),
            "contractor": _text(meta.get("Contractor")), "region": _text(meta.get("Region")),
            "sales": _text(meta.get("SalesPerson")),
            "presales": _text(meta.get("PresalesEngineer")),
            "pm": _text(meta.get("ProjectManager")),
        }
    if "edit_project_sheet" not in st.session_state:
        st.session_state.edit_project_sheet = {
            **DEFAULT_PROJECT_SHEET_INFO, **repo.load_project_sheet_info(meta)
        }
    if "edit_dirty_snapshot" not in st.session_state:
        st.session_state.edit_dirty_snapshot = _edit_snapshot()

    _sync_edit_state_from_widgets()
    ea1, ea2, ea3, ea4 = st.columns([1.55, 1.55, 1.45, 0.36], vertical_alignment="center")
    ea1.button(
        "💾 Save on this revision", type="primary", width="stretch",
        key="ed_save_this_top", on_click=_queue_edit_save, args=("this",),
        help="Overwrite the current revision/option in place - keeps the same "
             "offer #, revision and approval.",
    )
    ea2.button(
        "💾 Save as new revision", width="stretch", key="ed_save_revision_top",
        on_click=_queue_edit_save, args=("revision",),
        help=f"Create {repo.revision_token(nextrev)} as a changed version.",
    )
    ea3.button(
        "💾 Save as new option", width="stretch", key="ed_save_option_top",
        on_click=_queue_edit_save, args=("option",),
        help="Needs an Option label.",
    )
    if ea4.button("X", key="ed_cancel", width="stretch", help="Close editor"):
        if st.session_state.pop("_edit_just_saved", False):
            _close_edit_mode()
            st.rerun(scope="app")
        else:
            st.session_state.pending_close_edit = True
            st.rerun()
    if st.session_state.get("edit_show_cancel_dialog"):
        _cancel_edit_dialog()

    # ---- Editable offer header (client / project / system / people) ----
    eh = st.session_state.edit_header
    with st.expander("✏️ Project Details (client · project · system · people)", expanded=False):
        hc1, hc2, hc3 = st.columns(3)
        eh["client"] = hc1.text_input(
            "Client", key="eh_client",
            on_change=_copy_edit_header_widget, args=("client", "eh_client"))
        eh["project"] = hc1.text_input(
            "Project", key="eh_project",
            on_change=_copy_edit_header_widget, args=("project", "eh_project"))
        eh["contact"] = hc2.text_input(
            "Contact", key="eh_contact",
            on_change=_copy_edit_header_widget, args=("contact", "eh_contact"))
        eh["phone"] = hc2.text_input(
            "Phone", key="eh_phone",
            on_change=_copy_widget_to_state_dict,
            args=("edit_header", "phone", "eh_phone"))
        eh["contractor"] = hc3.text_input(
            "Contractor", key="eh_contractor",
            on_change=_copy_widget_to_state_dict,
            args=("edit_header", "contractor", "eh_contractor"))
        eh["region"] = _region_select(
            hc3, eh.get("region", ""), "eh_region",
            on_change=_copy_edit_header_widget, args=("region", "eh_region"))
        ph1, ph2, ph3 = st.columns(3)
        eh["sales"] = _person_select(ph1, "Sales Person", SALES_PERSON_ROLES,
                                     eh.get("sales", ""), "eh_sales",
                                     on_change=_copy_edit_header_widget,
                                     args=("sales", "eh_sales", "-"))
        eh["presales"] = _person_select(ph2, "Pre-sales Engineer", PEOPLE_ROLES["presales"],
                                        eh.get("presales", ""), "eh_presales",
                                        on_change=_copy_edit_header_widget,
                                        args=("presales", "eh_presales", "-"))
        eh["pm"] = _person_select(ph3, "Project Manager", PEOPLE_ROLES["pm"],
                                  eh.get("pm", ""), "eh_pm",
                                  on_change=_copy_edit_header_widget,
                                  args=("pm", "eh_pm", "-"))
        sy1, sy2 = st.columns([1, 2], vertical_alignment="bottom")
        st.session_state.edit_system = _system_select(
            sy1, st.session_state.get("edit_system", ""), "eh_system",
            on_change=_copy_edit_system_widget)
        st.session_state.edit_terms["system_note"] = st.session_state.edit_system
        sy2.caption("Changing the System updates the BOQ system; the existing Offer # stays unchanged.")
    st.session_state.edit_header = eh

    terms_form(st.session_state.edit_terms, "ed")
    edit_header_for_ps = {
        **st.session_state.edit_terms,
        "client": eh.get("client"), "project": eh.get("project"),
        "contact": eh.get("contact"), "phone": eh.get("phone"),
        "contractor": eh.get("contractor"), "region": eh.get("region"),
        "sales": eh.get("sales"), "date": meta.get("UpdatedDate") or meta.get("CreationDate"),
        "project_sheet": st.session_state.edit_project_sheet,
    }
    if _ps_enabled():
        project_sheet_info_form(edit_header_for_ps, "ed_ps")
        st.session_state.edit_project_sheet = edit_header_for_ps["project_sheet"]
    if "cached_default_margin" not in st.session_state:
        st.session_state.cached_default_margin = float(repo.get_setting("default_margin") or 1.6)
    catalogue_add("edit_grid", st.session_state.cached_default_margin, "ed",
                  st.session_state.get("edit_system", ""))
    col_pick, option_col, inc_c1 = st.columns([0.9, 1.8, 1.8], vertical_alignment="bottom")
    _show_incl = _inclusion_enabled() and st.session_state.get("ed_inclusion_mode") == "included"
    edit_column_order = _builder_column_order("edit_editor", host=col_pick, width="stretch",
                                              show_include=_show_incl)
    opt_label = option_col.text_input(
        "Option label",
        key="ed_option",
        placeholder="e.g. Dynalite, KNX",
        help="Names this alternative. Required for 'Save as new option'.",
    )

    if "ed_inclusion_mode" not in st.session_state:
        st.session_state["ed_inclusion_mode"] = "excluded"
    if _inclusion_enabled():
        inc_mode = inc_c1.selectbox(
            "Installation pricing",
            options=["excluded", "included"],
            format_func=lambda x: "Installation Excluded (standard)" if x == "excluded"
                                  else "Installation Included in item prices",
            key="ed_inclusion_mode",
        )
    else:
        inc_mode = "excluded"



    grid = render_editable_grid("edit_grid", "edit_editor", in_fragment=True,
                                column_order=edit_column_order,
                                show_included_col=(inc_mode == "included"))

    st.markdown("##### Totals")
    edit_calc_grid = calc.recompute(st.session_state.edit_grid)
    if inc_mode == "included":
        edit_totals_grid = calc.apply_inclusion(edit_calc_grid)
    else:
        edit_totals_grid = edit_calc_grid
    edit_base_summary = calc.summarize(edit_totals_grid, 0)
    tcol, pcol, cmcol, ccol, cpcol = st.columns(5)
    edit_discount = _discount_inputs(
        "ed", "edit_discount", edit_base_summary["subtotal_sar"], tcol, pcol)
    commission_base = max(edit_base_summary["subtotal_sar"] - edit_discount, 0.0)
    edit_commission, edit_commission_percent, edit_commission_mode = _commission_inputs(
        "ed", "edit_commission", commission_base,
        "edit_grid", "edit_editor", "edit_discount", cmcol, ccol, cpcol)
    s = calc.summarize(edit_totals_grid, edit_discount, edit_commission)
    m1, m2, m3 = st.columns(3)
    _subtotal_metric(m1, s)
    m2.metric(f"VAT {calc.VAT_RATE * 100:g}% (SAR)", f"{s['vat_amount_sar']:,.2f}")
    m3.metric("Grand Total (SAR)", f"{s['grand_total_sar']:,.2f}")
    _profit_banner(s)

    edit_terms = st.session_state.get("edit_terms", dict(DEFAULT_TERMS))
    edit_project_sheet = st.session_state.get("edit_project_sheet", dict(DEFAULT_PROJECT_SHEET_INFO))
    edit_header = st.session_state.get("edit_header", {})

    def _post_save(npid, nname, nrev):
        _saved_meta = repo.project_meta(npid)
        offer_rev = _saved_meta.get("OfferNo") or nname   # actual saved offer #
        h = {**edit_terms,
             "client": edit_header.get("client"), "project": edit_header.get("project") or nname,
             "contact": edit_header.get("contact"), "phone": edit_header.get("phone", ""),
             "contractor": edit_header.get("contractor"), "region": edit_header.get("region"),
             "sales": edit_header.get("sales"), "presales": edit_header.get("presales"),
             "pm": edit_header.get("pm"), "system": st.session_state.get("edit_system", ""),
             "offer": offer_rev, "date": dt.date.today().isoformat(),
             "option_label": _saved_meta.get("OptionLabel") or "",
             "project_sheet": edit_project_sheet}
        st.session_state.saved_export_header = h
        raw_grid = calc.recompute(st.session_state.edit_grid)
        if inc_mode == "included":
            st.session_state.saved_export_grid = calc.apply_inclusion(raw_grid)
        else:
            st.session_state.saved_export_grid = raw_grid
        st.session_state.saved_export_summary = dict(s)
        st.session_state.saved_rev = (npid, nname, nrev)

    # Execute the save requested from the action row above the grid (the buttons
    # there only set the flag; the work needs the totals/option computed above).
    _cur_rev = int(meta.get("RevisionNo") or 0)
    pending = st.session_state.pop("pending_save", None)
    pending_option_label = (
        st.session_state.pop("pending_option_label", opt_label) or ""
    ).strip() if pending else opt_label.strip()
    close_after_save = bool(st.session_state.pop("edit_close_after_save", False)) if pending else False
    if pending:
        _sync_edit_state_from_widgets()
        edit_terms = st.session_state.get("edit_terms", dict(DEFAULT_TERMS))
        edit_project_sheet = st.session_state.get("edit_project_sheet", dict(DEFAULT_PROJECT_SHEET_INFO))
        edit_header = st.session_state.get("edit_header", {})
    if pending == "this":
        repo.update_offer(
            st.session_state.edit_pid, calc.recompute(st.session_state.edit_grid),
            discount_sar=edit_discount,
            commission_sar=edit_commission,
            commission_percent=edit_commission_percent,
            commission_mode=edit_commission_mode,
            factors=(s["markup_factor"], None, None),
            system_suffix=st.session_state.get("edit_system") or _default_system(), terms=edit_terms,
            project_sheet_info=edit_project_sheet, header=edit_header,
            option_label=pending_option_label,
            inclusion_mode=inc_mode)
        _clear_offer_data_caches()
        saved_option_label = _text(
            repo.project_meta(st.session_state.edit_pid).get("OptionLabel")
        )
        _post_save(st.session_state.edit_pid,
                   edit_header.get("project") or meta.get("ProjectName"), _cur_rev)
        _name = edit_header.get("project") or meta.get("ProjectName")
        st.toast(f"Updated {_name}", icon="✅")
        st.success(
            f"Updated **{_name}** in place. Option label saved as "
            f"**{saved_option_label or 'Main'}**."
        )
        if close_after_save:
            _close_edit_mode()
            st.rerun(scope="app")
        _mark_edit_clean()
    elif pending == "revision":
        npid, nname, nrev = repo.save_revision(
            st.session_state.edit_pid, calc.recompute(st.session_state.edit_grid),
            discount_sar=edit_discount,
            commission_sar=edit_commission,
            commission_percent=edit_commission_percent,
            commission_mode=edit_commission_mode,
            factors=(s["markup_factor"], None, None),
            system_suffix=st.session_state.get("edit_system") or _default_system(),
            terms=edit_terms, option_label=pending_option_label,
            project_sheet_info=edit_project_sheet, header=edit_header,
            inclusion_mode=inc_mode)
        _clear_offer_data_caches()
        saved_option_label = _text(repo.project_meta(npid).get("OptionLabel"))
        _post_save(npid, nname, nrev)
        st.toast(f"Saved {nname} as ProjectID {npid}.", icon="✅")
        st.success(
            f"Saved **{nname}** as ProjectID {npid}. Option label: "
            f"**{saved_option_label or 'Main'}**."
        )
        _mark_edit_clean()
    elif pending == "option":
        if not pending_option_label:
            st.toast("Enter an Option label first.", icon="⚠️")
            st.warning("Enter an Option label first (e.g. Dynalite / KNX).")
        else:
            npid, nname, nrev = repo.save_option(
                st.session_state.edit_pid, calc.recompute(st.session_state.edit_grid),
                option_label=pending_option_label,
                discount_sar=edit_discount,
                commission_sar=edit_commission,
                commission_percent=edit_commission_percent,
                commission_mode=edit_commission_mode,
                factors=(s["markup_factor"], None, None),
                system_suffix=st.session_state.get("edit_system") or _default_system(), terms=edit_terms,
                project_sheet_info=edit_project_sheet, header=edit_header,
                inclusion_mode=inc_mode)
            _clear_offer_data_caches()
            saved_option_label = _text(repo.project_meta(npid).get("OptionLabel"))
            _post_save(npid, nname, nrev)
            st.toast(f"Saved option {nname} as ProjectID {npid}.", icon="✅")
            st.success(
                f"Saved option **{nname}** as ProjectID {npid}. Option label: "
                f"**{saved_option_label or 'Main'}**."
            )
            _mark_edit_clean()

    if st.session_state.pop("pending_close_edit", False):
        if _edit_has_unsaved_changes():
            st.session_state.edit_show_cancel_dialog = True
            st.rerun()
        else:
            _close_edit_mode()
            st.rerun(scope="app")

    if st.session_state.get("saved_rev"):
        pdf_col, download_col = st.columns(2)
        if pdf_col.button("📄 Generate saved offer PDF", key="ed_generate_saved_pdf",
                          width="stretch"):
            _make_pdf_download(
                st.session_state.saved_export_header,
                st.session_state.saved_export_grid,
                st.session_state.saved_export_summary,
            )
        fn = f"Quotation_{st.session_state.saved_rev[1]}.pdf".replace(" ", "")
        if "pdf_bytes" in st.session_state:
            download_col.download_button(
                "⬇️ Download PDF",
                st.session_state.pdf_bytes,
                file_name=fn,
                mime="application/pdf",
                width="stretch",
            )
        else:
            download_col.button("⬇️ Download PDF", disabled=True, width="stretch")


def _new_offer_actions():
    """Render and handle New Offer actions directly below the offer number."""
    _sync_new_header_from_widgets()
    h = st.session_state.header
    calc_grid = calc.recompute(st.session_state.grid)
    s = calc.summarize(
        calc_grid,
        st.session_state.get("discount") or 0.0,
        st.session_state.get("commission") or 0.0,
    )
    discount_sar = s["discount_sar"]
    commission_percent = max(
        _safe_float(st.session_state.get("no_commission_percent")), 0.0
    )
    commission_mode = st.session_state.get("no_commission_mode", "Protect profit")
    _optname = (st.session_state.get("no_option") or h.get("option") or "").strip()

    if st.session_state.get("no_offer_lock"):
        st.caption(
            f"Adding options to **{h['offer']}** - build this option, name it, then "
            "**Save option**. Use **➕ Add another option** to start the next one, or "
            "**🆕 New offer** to begin a fresh offer."
        )

    ac1, ac2, ac3, ac4, ac5, ac6, ac7 = st.columns(
        [1.0, 1.05, 0.85, 1.0, 0.85, 1.0, 0.95]
    )
    if ac1.button(
        "💾 Save option" if st.session_state.get("no_offer_lock") else "💾 Save offer",
        type="primary",
        width="stretch",
    ):
        _sync_new_header_from_widgets()
        h = st.session_state.header
        _optname = (st.session_state.get("no_option") or h.get("option") or "").strip()
        h["option"] = _optname
        _locked_now = st.session_state.get("no_offer_lock")
        _done = st.session_state.get("no_saved_options", [])
        if st.session_state.grid.empty:
            st.warning("Grid is empty.")
        elif not h.get("project", "").strip():
            st.warning("Project name is required.")
        elif not h.get("client", "").strip():
            st.warning("Client name is required.")
        elif _locked_now and not _optname:
            st.warning("Enter an Option label for this alternative (e.g. KNX).")
        elif _optname and _optname in _done:
            st.warning(f"Option '{_optname}' is already saved for this offer.")
        else:
            name = repo.project_name_with_option(h["project"] or "Untitled", _optname)
            pid = repo.save_offer(
                name=name,
                client=h["client"],
                contact=h["contact"],
                offer_no=h["offer"],
                system_suffix=h["system"],
                grid=calc_grid,
                discount_sar=discount_sar,
                commission_sar=s["commission_sar"],
                commission_percent=commission_percent,
                commission_mode=commission_mode,
                factors=(s["markup_factor"], None, None),
                sales_person=h.get("sales"),
                presales_engineer=h.get("presales"),
                project_manager=h.get("pm"),
                terms={k: h.get(k) for k in TERMS_KEYS},
                option_label=_optname,
                project_sheet_info=h.get("project_sheet"),
                phone=h.get("phone"),
                contractor=h.get("contractor"),
                region=h.get("region"),
                inclusion_mode=h.get("inclusion_mode", "excluded"),
            )
            # Load the saved offer and open it directly in edit mode.
            _sys, _meta, _grid, _ = _cached_project_bundle(pid, (pid,), _db_cache_stamp())
            _sheet = _sys[0] if _sys else None
            _init_edit_state(pid, _meta, _grid, _sheet)
            _prime_new_offer_form()
            _clear_offer_data_caches()
            st.session_state["_just_saved_meta"] = _meta
            st.session_state["_just_saved_fam"] = repo.family_key(
                _meta.get("OfferNo") or "", _meta.get("ProjectName") or "")
            st.session_state["_just_saved_pid"] = pid
            _request_scroll_top()
            st.rerun()

    if ac2.button(
        "➕ Add another option",
        width="stretch",
        disabled=not st.session_state.get("no_offer_lock"),
    ):
        st.session_state.grid = _empty_grid()
        st.session_state["_no_reset_option"] = True
        st.rerun()

    if ac3.button("🆕 New offer", width="stretch"):
        st.session_state.grid = _empty_grid()
        st.session_state.no_offer_lock = None
        st.session_state.no_saved_options = []
        st.session_state["_no_reset_all"] = True
        st.session_state.pop("pdf_bytes", None)
        st.session_state.pop("project_sheet_bytes", None)
        st.rerun()

    if ac4.button("📄 Generate Offer PDF", width="stretch"):
        _make_pdf_download({**h, "option_label": _optname}, st.session_state.grid, s)

    _pdf_name = f"Quotation_{h['offer']}{(' ' + _optname) if _optname else ''}.pdf"
    if "pdf_bytes" in st.session_state:
        ac5.download_button(
            "⬇️ Download PDF",
            st.session_state.pdf_bytes,
            file_name=_pdf_name,
            mime="application/pdf",
            width="stretch",
        )
    else:
        ac5.button("⬇️ Download PDF", disabled=True, width="stretch")

    if _ps_enabled() and ac6.button(
        "📊 Generate Sheet",
        help="Generate Project Sheet",
        width="stretch",
    ):
        _make_project_sheet_download(h, s)

    if _ps_enabled():
        if "project_sheet_bytes" in st.session_state:
            ac7.download_button(
                "⬇️ Download Sheet",
                st.session_state.project_sheet_bytes,
                file_name=(
                    f"Project_Sheet_{_safe_filename(h.get('offer') or h.get('project'))}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                help="Download Project Sheet",
                width="stretch",
            )
        else:
            ac7.button(
                "⬇️ Download Sheet",
                disabled=True,
                help="Generate the Project Sheet first",
                width="stretch",
            )


@st.fragment
def _new_project_editor():
    _sync_new_header_from_widgets()
    h = st.session_state.header

    # ---- Add items from catalogue ----
    if "cached_default_margin" not in st.session_state:
        st.session_state.cached_default_margin = float(repo.get_setting("default_margin") or 1.6)
    _dm = st.session_state.cached_default_margin
    catalogue_add("grid", _dm, "no", st.session_state.header["system"])

    # ---- Columns / Option label / Installation pricing row (same layout as Edit Offer) ----
    _show_incl = _inclusion_enabled() and h.get("inclusion_mode") == "included"
    col_pick, option_col, inc_c1 = st.columns([0.9, 1.8, 1.8], vertical_alignment="bottom")
    no_column_order = _builder_column_order("editor", host=col_pick, width="stretch",
                                            show_include=_show_incl)
    h["option"] = option_col.text_input(
        "Option label",
        key="no_option",
        placeholder="e.g. Dynalite, KNX",
        help="Name this alternative. Leave blank for a single-option offer.",
    )
    if _inclusion_enabled():
        h["inclusion_mode"] = inc_c1.selectbox(
            "Installation pricing",
            options=["excluded", "included"],
            index=0 if h.get("inclusion_mode", "excluded") == "excluded" else 1,
            format_func=lambda x: "Installation Excluded (standard)" if x == "excluded"
                                  else "Installation Included in item prices",
            key="no_inclusion_mode",
        )
    else:
        h["inclusion_mode"] = "excluded"
    st.session_state.header = h

    # ---- Editable grid (builder always shows costs) ----
    grid = render_editable_grid("grid", "editor", in_fragment=True,
                                column_order=no_column_order,
                                show_included_col=_show_incl)

    # ---- Discount + totals ----
    st.markdown("##### Totals")
    calc_grid = calc.recompute(st.session_state.grid)
    base_summary = calc.summarize(calc_grid, 0)
    dcol, pcol, cmcol, ccol, cpcol = st.columns(5)
    discount_sar = _discount_inputs("no", "discount", base_summary["subtotal_sar"], dcol, pcol)
    commission_base = max(base_summary["subtotal_sar"] - discount_sar, 0.0)
    commission_sar, commission_percent, commission_mode = _commission_inputs(
        "no", "commission", commission_base,
        "grid", "editor", "discount", cmcol, ccol, cpcol)
    s = calc.summarize(calc_grid, discount_sar, commission_sar)

    m1, m2, m3 = st.columns(3)
    _subtotal_metric(m1, s)
    m2.metric(f"VAT {calc.VAT_RATE * 100:g}% (SAR)", f"{s['vat_amount_sar']:,.2f}")
    m3.metric("Grand Total (SAR)", f"{s['grand_total_sar']:,.2f}")
    _profit_banner(s)
    if admin:
        a1, a2, a3 = st.columns(3)
        a1.metric("Total Cost (USD)", f"{s['total_cost_usd']:,.2f}")
        a2.metric("Cost in SAR", f"{s['cost_sar']:,.2f}")
        a3.metric("Total Selling (USD)", f"{s['total_sell_usd']:,.2f}")

def catalogue_add(state_key: str, default_margin: float, kp: str, default_system: str = "",
                  show_clear: bool = False):
    """Type-ahead catalogue search + add controls writing into st.session_state[state_key]."""
    st.markdown("##### Add item from catalogue")
    term = st.text_input("Search Model / Description / Brand", key=f"{kp}_term",
                         placeholder="e.g. PDEG, keypad, Dynalite…").strip()
    results = _cached_search_catalog(term, 20, _db_cache_stamp()) if term else pd.DataFrame()
    if not results.empty:
        results = results.assign(_label=results.apply(
            lambda r: f"{r['Model']} - {str(r['Description'])[:48]} ({r['Brand']})  ·x{r['TimesQuoted']}", axis=1))
        a1, a2, a3, a4, a5, a6 = st.columns(
            [4.0, 1.0, 0.85, 1.3, 1.3, 0.9], vertical_alignment="bottom")
        pick = a1.selectbox("Match", results["_label"].tolist(), key=f"{kp}_pick")
        chosen = results[results["_label"] == pick].iloc[0].to_dict()
        a2.text_input("Updated On", _fmt_month_year(chosen.get("PriceUpdatedAt")),
                      disabled=True, key=f"{kp}_price_updated_{chosen.get('ItemID')}")
        qty = a3.number_input("Qty", min_value=1, value=1, step=1, key=f"{kp}_qty")
        area = a4.text_input("Area", value="All Areas", key=f"{kp}_area")
        system = a5.text_input("System", value=default_system, key=f"{kp}_system")
        if a6.button("➕ Add", width="stretch", key=f"{kp}_add"):
            _add_row_to(state_key, repo.item_to_grid_row(
                chosen, area=area, system=system, qty=int(qty), default_margin=default_margin))
            st.rerun()
    elif term:
        st.info("No catalogue match - use the + row in the table to type freely.")
    if show_clear:
        bc1, _ = st.columns([1, 5])
        if bc1.button("🧹 Clear grid", key=f"{kp}_clear", width="stretch"):
            st.session_state[state_key] = _empty_grid()
            st.rerun()


def terms_form(store: dict, kp: str):
    """Editable Quotation terms/notes (subject, greeting, scope, payment, ...)."""
    # Pre-populate widget keys from store only when absent — avoids the Streamlit
    # warning that fires when both an explicit value and a session-state key are
    # provided to the same widget after the key already exists in session state.
    _wkeys = {
        f"{kp}_subject": "subject", f"{kp}_greet": "greeting",
        f"{kp}_scope": "scope", f"{kp}_excl": "exclusions",
        f"{kp}_prereq": "prerequisites", f"{kp}_deliv": "delivery",
        f"{kp}_valid": "validity", f"{kp}_pay": "payment",
        f"{kp}_notes": "notes",
    }
    for wkey, field in _wkeys.items():
        if wkey not in st.session_state:
            st.session_state[wkey] = store.get(field, "")
    with st.expander("📋 Terms, scope & notes (appear on the quotation)", expanded=False):
        store["subject"] = st.text_input("Subject (offer title)",
            key=f"{kp}_subject", placeholder="e.g. Low Current Systems Offer")
        store["greeting"] = st.text_area("Greeting", key=f"{kp}_greet", height=80)
        store["scope"] = st.text_input("Scope", key=f"{kp}_scope")
        store["exclusions"] = st.text_area("Exclusions", key=f"{kp}_excl", height=70)
        store["prerequisites"] = st.text_area("Pre-requirements", key=f"{kp}_prereq", height=70)
        c3, c4 = st.columns(2)
        store["delivery"] = c3.text_input("Delivery", key=f"{kp}_deliv")
        store["validity"] = c4.text_input("Validity", key=f"{kp}_valid")
        store["payment"] = st.text_area("Payment Terms", key=f"{kp}_pay", height=70)
        store["notes"] = st.text_area("Special notes & instructions", key=f"{kp}_notes", height=70)


def _project_sheet_info(info: dict | None = None, h: dict | None = None) -> dict:
    h = h or {}
    data = {**DEFAULT_PROJECT_SHEET_INFO, **(info or {})}
    if not data.get("sheet_date"):
        data["sheet_date"] = h.get("date") or dt.date.today().isoformat()
    if not data.get("invoice_to"):
        data["invoice_to"] = h.get("client") or ""
    if not data.get("delivery_instructions"):
        contact = " ".join(x for x in [h.get("contact"), h.get("phone")] if x)
        data["delivery_instructions"] = contact
    if data.get("lead_source") not in PROJECT_LEAD_SOURCE_OPTIONS:
        data["lead_source"] = data.get("lead_source") or PROJECT_LEAD_SOURCE_OPTIONS[0]
    if data.get("shipment_by") not in PROJECT_SHIPMENT_OPTIONS:
        data["shipment_by"] = data.get("shipment_by") or PROJECT_SHIPMENT_OPTIONS[0]
    return {k: data.get(k, "") for k in PROJECT_SHEET_KEYS}


def _option_index(options: list[str], value: str) -> tuple[list[str], int]:
    opts = list(options)
    value = _text(value)
    if value and value not in opts:
        opts = [value] + opts
    return opts, opts.index(value) if value in opts else 0


def project_sheet_info_form(store: dict, kp: str):
    """Editable Project Sheet export details."""
    ps = _project_sheet_info(store.get("project_sheet"), store)

    def _txt(container, field, label, area=False, **kw):
        # If the key is already in Session State (set by the edit/duplicate prefill,
        # or owned by the widget after the first render), let Session State drive it.
        # Passing a default value as well triggers a Streamlit warning.
        key = f"{kp}_{field}"
        if key not in st.session_state:
            kw["value"] = ps.get(field, "")
        widget = container.text_area if area else container.text_input
        ps[field] = widget(label, key=key, **kw)

    def _sel(container, field, label, options):
        key = f"{kp}_{field}"
        opts, idx = _option_index(options, st.session_state.get(key, ps.get(field, "")))
        kw = {} if key in st.session_state else {"index": idx}   # omit index when SS-driven
        ps[field] = container.selectbox(label, opts, key=key, **kw)

    with st.expander("Project Sheet Information", expanded=False):
        c1, c2, c3 = st.columns(3)
        _txt(c1, "job_reference", "Project Job Reference")
        _txt(c2, "sheet_date", "Project Sheet Date")
        _sel(c3, "lead_source", "Project Lead Source", PROJECT_LEAD_SOURCE_OPTIONS)

        c4, c5, c6 = st.columns(3)
        _txt(c4, "commission", "Architect/Contractor Commissions (if any)")
        _sel(c5, "shipment_by", "Based on Shipments by", PROJECT_SHIPMENT_OPTIONS)
        _txt(c6, "downpayment_date", "Downpayment Date")

        _txt(st, "invoice_to", "Invoice to")
        _txt(st, "delivery_instructions",
             "Delivery Instructions / Contact person & details", area=True, height=70)
        c9, _ = st.columns(2)
        _txt(c9, "gm_signature", "GM Signature Name")

    store["project_sheet"] = ps
    return ps


def _make_pdf_download(h, grid, summary, options=None):
    global db, pdf_export
    if "header_left_path" not in dir(db):
        db = importlib.reload(db)
    if "template" not in inspect.signature(pdf_export.generate_quotation_pdf).parameters:
        pdf_export = importlib.reload(pdf_export)
    notes = {
        "System": h.get("system") or h.get("system_note"), "Scope": h.get("scope"),
        "Exclusions": h.get("exclusions"), "Pre-requirements": h.get("prerequisites"),
        "Delivery": h.get("delivery"), "Payment Terms": h.get("payment"),
        "Validity": h.get("validity"), "Notes": h.get("notes"),
    }
    header = {"title": h.get("subject") or "Quotation",
              "client": h.get("client"), "project": h.get("project"),
              "contact": h.get("contact"), "phone": h.get("phone"),
              "sales": h.get("sales"), "presales": h.get("presales"), "pm": h.get("pm"),
              "offer": h.get("offer"), "date": h.get("date"),
              "greeting": h.get("greeting") or DEFAULT_TERMS["greeting"]}
    company = {
        "name": repo.get_setting("company_name") or "Company Name",
        "tagline": repo.get_setting("company_tagline") or "",
        "contact": repo.get_setting("company_contact") or "",
        "vat_number": repo.get_setting("company_vat_number") or "",
        "cr_number": repo.get_setting("company_cr_number") or "",
        "color": repo.get_setting("company_brand_color") or "#002060",
        "header_left": repo.get_setting("header_left_text") or "",
        "header_middle": repo.get_setting("header_middle_text") or "",
        "header_right": repo.get_setting("header_right_text") or "",
        "footer_left": repo.get_setting("footer_left_text") or "",
        "footer_middle": repo.get_setting("footer_middle_text") or "",
        "footer_right": repo.get_setting("footer_right_text") or "",
    }
    pdf_body_template = repo.get_setting("pdf_body_template") or "template1"
    tmp = os.path.join(db.DATA_DIR, "_last_quotation.pdf")
    if options:                       # one document, a section per option
        pdf_export.generate_options_pdf(tmp, header, options, notes=notes,
                                        company=company, show_costs=False,
                                        template=pdf_body_template)
    else:
        pdf_export.generate_quotation_pdf(tmp, header, grid, summary, notes=notes,
                                          company=company, show_costs=False,
                                          template=pdf_body_template,
                                          option_label=h.get("option_label") or "")
    with open(tmp, "rb") as f:
        st.session_state.pdf_bytes = f.read()
    n = len(options) if options else 1
    st.toast(
        f"PDF ready ({n} option{'s' if n > 1 else ''}) - click Download PDF.",
        icon="📄",
    )


def _project_sheet_bytes(h: dict, s: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "PROJECT SHEET"
    ws.sheet_view.showGridLines = False

    ps = _project_sheet_info(h.get("project_sheet"), h)
    net_sales = s.get("discounted_subtotal_sar") or s.get("subtotal_sar") or 0.0
    landed_cost = s.get("cost_sar") or 0.0

    def _display_date(value):
        raw = _text(value)
        if not raw:
            return dt.date.today().strftime("%B %d,%Y")
        try:
            return dt.date.fromisoformat(raw[:10]).strftime("%B %d,%Y")
        except ValueError:
            return raw

    def _shipment_text(value):
        ship = (_text(value) or "Air").lower()
        air = "X" if ship == "air" else " "
        sea = "X" if ship == "sea" else " "
        return f"(  {air}  ) Air                   (  {sea}  ) Sea"

    header_fill = PatternFill("solid", fgColor="FFFFFF")
    fallback_header_fill = PatternFill("solid", fgColor="002060")
    sidebar_fill = PatternFill("solid", fgColor="002060")
    side = Side(style="thin", color="1F1F1F")
    border = Border(left=side, right=side, top=side, bottom=side)
    label_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=11)

    ws.merge_cells("A1:E4")
    has_banner = os.path.exists(_LOGO)
    for row in range(1, 5):
        ws.row_dimensions[row].height = 20
        for col in range(1, 6):
            ws.cell(row, col).fill = header_fill if has_banner else fallback_header_fill
    if has_banner:
        banner = XLImage(_LOGO)
        banner.width = 695
        banner.height = 77
        ws.add_image(banner, "A1")
    else:
        ws["A1"] = repo.get_setting("company_name") or "Company Name"
        ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[5].height = 10

    ws.merge_cells("A6:A24")
    ws["A6"] = "Project Sheet"
    ws["A6"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws["A6"].fill = sidebar_fill
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center",
                                   text_rotation=90, wrap_text=True)
    ws["A6"].border = border

    rows = [
        (6, "Project Job Reference:", ps.get("job_reference"), False),
        (7, "Date:", _display_date(ps.get("sheet_date")), False),
        (8, "Project Name:", h.get("project") or h.get("subject") or "", False),
        (9, "Client/Contractor Name:", h.get("client") or "", False),
        (10, "Confirmed Offer Reference:", h.get("offer") or "", False),
        (11, "Salesman:", h.get("sales") or "", False),
        (12, "Project Lead Source:", ps.get("lead_source"), False),
        (13, "Net Projects Sales Amount without VAT:", net_sales, True),
        (14, "Architect/Contractor Commissions (If any):", ps.get("commission"), False),
        (15, "Projected Total Project Landed Cost:", landed_cost, True),
        (16, "Projected Margin:", "=IF(D13>0,(D13-D15)/D13,0)", False, "percent"),
        (17, "Projected Profit:", "=D13-D15", True),
        (18, "Based on Shipments by:", _shipment_text(ps.get("shipment_by")), False),
        (19, "Payment Terrms:", h.get("payment") or "", False),
        (20, "Downpayment Date:", ps.get("downpayment_date"), False),
        (21, "Invoice to:", ps.get("invoice_to"), False),
        (22, "Contractual Project Delivery Date:", h.get("delivery") or "", False),
        (23, "Delivery Instructions / Contact person & details:", ps.get("delivery_instructions"), False),
        (24, "Notes: ", h.get("notes") or "", False),
    ]
    for item in rows:
        row, label, value, money_row = item[:4]
        value_kind = item[4] if len(item) > 4 else ""
        ws.cell(row, 2, label)
        if money_row:
            ws.cell(row, 3, "SAR")
            ws.cell(row, 4, value)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.cell(row, 3, value)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row, 2).font = label_font
        ws.cell(row, 3).font = value_font
        ws.cell(row, 4).font = value_font
        ws.cell(row, 2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row, 3).alignment = Alignment(horizontal="center",
                                              vertical="center", wrap_text=True)
        ws.cell(row, 4).alignment = Alignment(horizontal="right" if money_row else "center",
                                              vertical="center", wrap_text=True)
        if value_kind == "percent":
            ws.cell(row, 3).number_format = "0%"
        for col in range(2, 6):
            ws.cell(row, col).border = border

    ws["D13"].number_format = "#,##0.00"
    ws["D15"].number_format = "#,##0.00"
    ws["C16"].number_format = "0%"
    ws["D17"].number_format = "#,##0.00"

    ws.merge_cells("A25:B27")
    ws["A25"] = "Salesman Signature\n\n" + _text(h.get("sales"))
    ws.merge_cells("C25:E27")
    ws["C25"] = "GM Signature" + (("\n\n" + ps.get("gm_signature")) if ps.get("gm_signature") else "")
    for cell in (ws["A25"], ws["C25"]):
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row in range(25, 28):
        for col in range(1, 6):
            ws.cell(row, col).border = border

    for row in range(6, 25):
        ws.row_dimensions[row].height = 27
    ws.row_dimensions[13].height = 31
    ws.row_dimensions[14].height = 31
    ws.row_dimensions[15].height = 31
    ws.row_dimensions[17].height = 31
    ws.row_dimensions[19].height = 35
    ws.row_dimensions[23].height = 33
    ws.row_dimensions[24].height = 33
    ws.row_dimensions[25].height = 28
    ws.row_dimensions[26].height = 28
    ws.row_dimensions[27].height = 28
    for col, width in {
        "A": 9, "B": 39, "C": 9, "D": 22, "E": 20,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"
    ws.print_area = "A1:E27"
    ws.print_title_rows = "1:5"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _make_project_sheet_download(h: dict, summary: dict):
    st.session_state.project_sheet_bytes = _project_sheet_bytes(h, summary)
    st.toast("Project Sheet ready - use the download button.", icon="📊")


def _safe_filename(value, fallback="Project"):
    name = _text(value, fallback)
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in name).strip("_") or fallback


def revision_options(base_pid):
    """All active options of the same revision as `base_pid`, as PDF sections."""
    meta = repo.project_meta(base_pid)
    fam = repo.family_key(meta.get("OfferNo"), meta.get("ProjectName"))
    rev = int(meta.get("RevisionNo") or 0)
    projs = repo.list_projects()
    same = projs[projs.apply(
        lambda r: repo.family_key(r.get("OfferNo"), r.get("ProjectName")) == fam
        and int(r.get("RevisionNo") or 0) == rev, axis=1)]
    active = same[same["Archived"].fillna(0) == 0]
    rows = active if not active.empty else same[same["ProjectID"] == base_pid]
    rows = rows.sort_values("OptionLabel", na_position="first")
    out = []
    for _, r in rows.iterrows():
        pid = int(r["ProjectID"])
        m = repo.project_meta(pid)
        sh = (repo.list_systems(pid) or [None])[0]
        g = repo.load_project_grid(pid, sh).copy()
        for col in MONEY_COLS:
            if col in g.columns:
                g[col] = g[col].map(lambda v: calc.roundup(v, 0))
        if m.get("InclusionMode") == "included":
            g = calc.apply_inclusion(g)
        out.append({"label": m.get("OptionLabel") or "",
                    "grid": g, "summary": calc.summarize(
                        g, m.get("DiscountAmount") or 0, m.get("CommissionAmount") or 0)})
    return out


def _profit_banner(s: dict):
    """Profit bubble: profit big (left) with Margin/markup beneath, cost big (right)."""
    profit = s.get("gross_margin_sar") or 0.0
    profit_usd = s.get("gross_margin_usd") or 0.0
    sub = s.get("discounted_subtotal_sar") or 0.0
    margin_pct = (profit / sub * 100) if sub else 0.0
    factor = s.get("markup_factor")
    commission = s.get("commission_sar", 0) or 0.0
    cost_sar = s.get("cost_sar", 0) or 0.0
    cost_usd = s.get("total_cost_usd", 0) or 0.0
    markup_txt = f"Markup x{factor:.2f}" if factor else "Markup -"
    if profit >= 0:
        bg, fg, sub_fg = "rgba(33,195,84,0.12)", "#0b6b34", "#3f7d59"
    else:
        bg, fg, sub_fg = "rgba(255,43,43,0.10)", "#b02a37", "#a05560"
    lbl = f"font-size:0.8rem;font-weight:700;color:{sub_fg};text-transform:uppercase;letter-spacing:0.04em"
    big = f"font-size:1.5rem;font-weight:800;color:{fg};line-height:1.2"
    small = f"font-size:0.9rem;font-weight:700;color:{sub_fg}"
    mid = f"font-size:1.0rem;font-weight:800;color:{fg}"

    def _block(label, sar, usd, align):
        return (f"<div style='flex:1;min-width:130px;text-align:{align}'>"
                f"<div style='{lbl}'>{label}</div>"
                f"<div style='{big}'>SAR {sar:,.2f}</div>"
                f"<div style='{small}'>$ {usd:,.2f}</div></div>")

    commission_html = (
        f"<div style='{small}'>Commission cost SAR {commission:,.2f}</div>"
        if commission else ""
    )
    html = (
        f"<div style='background:{bg};border-radius:8px;padding:14px 24px;margin:2px 0 10px;"
        f"display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap'>"
        + _block("🧾 Cost", cost_sar, cost_usd, "left")
        + f"<div style='flex:1;min-width:120px;text-align:center'>"
          f"<div style='{mid}'>{markup_txt}</div>"
          + commission_html
          + f"<div style='{mid}'>Margin {margin_pct:.1f}%</div></div>"
        + _block("💰 Gross Profit", profit, profit_usd, "right")
        + "</div>"
    )
    st.markdown(html, unsafe_allow_html=True)


def _fin_bubble(left_label, left_val, middle_lines, right_label, right_val, positive=True):
    """Green (or red) summary bubble used under each Finance table."""
    bg, fg, sub_fg = (("rgba(33,195,84,0.12)", "#0b6b34", "#3f7d59") if positive
                      else ("rgba(255,43,43,0.10)", "#b02a37", "#a05560"))
    lbl = f"font-size:0.78rem;font-weight:700;color:{sub_fg};text-transform:uppercase;letter-spacing:0.04em"
    big = f"font-size:1.3rem;font-weight:800;color:{fg};line-height:1.25"
    mid = f"font-size:0.92rem;font-weight:800;color:{fg}"
    mid_html = "".join(f"<div style='{mid}'>{m}</div>" for m in middle_lines)
    st.markdown(
        f"<div style='background:{bg};border-radius:8px;padding:12px 20px;margin:8px 0 2px;"
        f"display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap'>"
        f"<div style='flex:1;min-width:120px;text-align:left'><div style='{lbl}'>{left_label}</div>"
        f"<div style='{big}'>{left_val}</div></div>"
        f"<div style='flex:1;min-width:110px;text-align:center'>{mid_html}</div>"
        f"<div style='flex:1;min-width:120px;text-align:right'><div style='{lbl}'>{right_label}</div>"
        f"<div style='{big}'>{right_val}</div></div></div>", unsafe_allow_html=True)


def _subtotal_metric(col, s: dict):
    discount = s.get("discount_sar") or 0.0
    if discount:
        col.metric("Subtotal after discount (SAR)",
                   f"{s['discounted_subtotal_sar']:,.2f}",
                   delta=f"-{discount:,.2f} discount", delta_color="inverse")
    else:
        col.metric("Subtotal (SAR)", f"{s['subtotal_sar']:,.2f}")


def _summary_metrics(s: dict):
    commission = s.get("commission_sar", 0) or 0.0
    if commission:
        m1, m2, m3, m4 = st.columns(4)
        _subtotal_metric(m1, s)
        m2.metric("Internal Commission (SAR)", f"{commission:,.2f}",
                  help="Internal expense only; excluded from the client quotation and profit.")
        m3.metric(f"VAT {calc.VAT_RATE * 100:g}% (SAR)", f"{s['vat_amount_sar']:,.2f}")
        m4.metric("Grand Total (SAR)", f"{s['grand_total_sar']:,.2f}")
    else:
        m1, m2, m3 = st.columns(3)
        _subtotal_metric(m1, s)
        m2.metric(f"VAT {calc.VAT_RATE * 100:g}% (SAR)", f"{s['vat_amount_sar']:,.2f}")
        m3.metric("Grand Total (SAR)", f"{s['grand_total_sar']:,.2f}")


def _project_details_readonly(meta: dict, system=""):
    def _readonly_field(container, label, value):
        label_html = html.escape(label)
        value_html = html.escape(_text(value, "-"))
        container.markdown(
            "<div style='margin-bottom:0.7rem'>"
            f"<div style='font-size:0.86rem;font-weight:650;color:#516173;margin-bottom:0.25rem'>{label_html}</div>"
            "<div style='min-height:2.45rem;border:1px solid #d9e0e8;background:#f8fafc;"
            "border-radius:8px;padding:0.62rem 0.8rem;color:#182333;font-size:0.98rem;"
            f"font-weight:550;line-height:1.25'>{value_html}</div>"
            "</div>",
            unsafe_allow_html=True,
        )

    with st.expander("Project Details", expanded=False):
        c1, c2, c3 = st.columns(3)
        _readonly_field(c1, "Client", meta.get("ClientName"))
        _readonly_field(c1, "Project", meta.get("ProjectName"))
        _readonly_field(c2, "Contact", meta.get("ContactName"))
        _readonly_field(c2, "Phone", meta.get("ContactPhone"))
        _readonly_field(c3, "Contractor", meta.get("Contractor"))
        _readonly_field(c3, "Region", meta.get("Region"))
        _readonly_field(c3, "System", repo.system_name(system))

        p1, p2, p3 = st.columns(3)
        _readonly_field(p1, "Sales Person", meta.get("SalesPerson"))
        _readonly_field(p2, "Pre-sales Engineer", meta.get("PresalesEngineer"))
        _readonly_field(p3, "Project Manager", meta.get("ProjectManager"))

        d1, d2, d3 = st.columns(3)
        _readonly_field(d1, "Created", _fmt_date(meta.get("CreationDate")))
        _readonly_field(
            d2, "Last Updated",
            _fmt_date(meta.get("UpdatedDate") or meta.get("CreationDate")),
        )
        _readonly_field(d3, "Offer #", meta.get("OfferNo"))


def _safe_float(value, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        if pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _discount_percent(amount: float, subtotal: float) -> float:
    return round((amount / subtotal * 100), 4) if subtotal else 0.0


def _sync_discount_from_amount(amount_key: str, percent_key: str,
                               subtotal_key: str, driver_key: str):
    subtotal = max(_safe_float(st.session_state.get(subtotal_key)), 0.0)
    amount = min(abs(_safe_float(st.session_state.get(amount_key))), subtotal)
    st.session_state[amount_key] = amount
    st.session_state[percent_key] = _discount_percent(amount, subtotal)
    st.session_state[driver_key] = "amount"


def _sync_discount_from_percent(amount_key: str, percent_key: str,
                                subtotal_key: str, driver_key: str):
    subtotal = max(_safe_float(st.session_state.get(subtotal_key)), 0.0)
    percent = min(abs(_safe_float(st.session_state.get(percent_key))), 100.0)
    st.session_state[percent_key] = percent
    st.session_state[amount_key] = round(subtotal * percent / 100, 2)
    st.session_state[driver_key] = "percent"


def _discount_inputs(prefix: str, amount_key: str, subtotal: float,
                     amount_col=None, percent_col=None) -> float:
    subtotal = max(_safe_float(subtotal), 0.0)
    percent_key = f"{prefix}_discount_percent"
    subtotal_key = f"{prefix}_discount_subtotal"
    driver_key = f"{prefix}_discount_driver"
    st.session_state[subtotal_key] = subtotal

    amount = min(abs(_safe_float(st.session_state.get(amount_key))), subtotal)
    st.session_state[amount_key] = amount
    if st.session_state.get(driver_key) not in ("amount", "percent"):
        st.session_state[driver_key] = "amount"

    if st.session_state[driver_key] == "percent":
        percent = min(abs(_safe_float(st.session_state.get(percent_key))), 100.0)
        st.session_state[percent_key] = percent
        st.session_state[amount_key] = round(subtotal * percent / 100, 2)
    else:
        st.session_state[percent_key] = _discount_percent(st.session_state[amount_key], subtotal)

    if amount_col is None or percent_col is None:
        amount_col, percent_col = st.columns(2)
    if st.session_state[amount_key] > subtotal:
        st.session_state[amount_key] = subtotal
        st.session_state[percent_key] = _discount_percent(subtotal, subtotal)

    amount_col.number_input(
        "Discount (SAR)", min_value=0.0, max_value=subtotal, step=100.0, format="%.2f", key=amount_key,
        help="Fixed discount amount to subtract from the subtotal.",
        on_change=_sync_discount_from_amount,
        args=(amount_key, percent_key, subtotal_key, driver_key))
    percent_col.number_input(
        "Discount %", min_value=0.0, max_value=100.0, step=1.0, format="%.2f", key=percent_key,
        help="Percentage of the current subtotal. Editing this updates Discount (SAR).",
        on_change=_sync_discount_from_percent,
        args=(amount_key, percent_key, subtotal_key, driver_key))
    return abs(_safe_float(st.session_state.get(amount_key)))


COMMISSION_MODES = ("Protect profit", "Deduct from profit")


def _reprice_commission(prefix: str, amount_key: str, grid_key: str,
                        editor_key: str, discount_key: str, driver: str):
    """Synchronize commission fields and apply/remove a margin gross-up."""
    percent_key = f"{prefix}_commission_percent"
    mode_key = f"{prefix}_commission_mode"
    applied_key = f"{prefix}_commission_applied_percent"
    base_key = f"{prefix}_commission_base_subtotal"
    mode = st.session_state.get(mode_key, "Protect profit")
    if mode not in COMMISSION_MODES:
        mode = "Deduct from profit"
        st.session_state[mode_key] = mode

    discount = abs(_safe_float(st.session_state.get(discount_key)))
    grid = calc.recompute(st.session_state.get(grid_key, _empty_grid()))
    old_applied = max(_safe_float(st.session_state.get(applied_key)), 0.0)

    # First recover the original, non-grossed-up margins. This makes changing
    # 15 -> 10, switching modes, or removing commission non-compounding.
    if old_applied > 0:
        reverse = (1 / (1 + old_applied / 100) - 1) * 100
        grid, _ = calc.increase_margins(grid, reverse)
    base_subtotal = calc.summarize(grid, discount)["discounted_subtotal_sar"]
    st.session_state[base_key] = base_subtotal

    entered_amount = max(_safe_float(st.session_state.get(amount_key)), 0.0)
    entered_percent = min(
        max(_safe_float(st.session_state.get(percent_key)), 0.0), 1000.0
    )
    if driver == "amount":
        target_percent = (entered_amount / base_subtotal * 100) if base_subtotal else 0.0
    else:
        target_percent = entered_percent
    target_percent = min(max(target_percent, 0.0), 1000.0)

    if mode == "Protect profit" and target_percent > 0:
        grid, changed = calc.increase_margins(grid, target_percent)
        increased_subtotal = calc.summarize(grid, discount)["discounted_subtotal_sar"]
        # Use the exact uplift after line rounding so gross profit is preserved.
        commission_amount = round(max(increased_subtotal - base_subtotal, 0.0), 2)
        applied_percent = target_percent if changed else 0.0
    else:
        commission_amount = round(base_subtotal * target_percent / 100, 2)
        applied_percent = 0.0

    st.session_state[amount_key] = commission_amount
    st.session_state[percent_key] = target_percent
    st.session_state[applied_key] = applied_percent
    st.session_state[grid_key] = grid
    st.session_state.pop(editor_key, None)


def _commission_inputs(prefix: str, amount_key: str, subtotal: float,
                       grid_key: str, editor_key: str, discount_key: str,
                       mode_col=None, amount_col=None, percent_col=None
                       ) -> tuple[float, float, str]:
    """Editable commission amount/rate with selectable profit treatment."""
    subtotal = max(_safe_float(subtotal), 0.0)
    percent_key = f"{prefix}_commission_percent"
    mode_key = f"{prefix}_commission_mode"
    applied_key = f"{prefix}_commission_applied_percent"
    base_key = f"{prefix}_commission_base_subtotal"
    st.session_state.setdefault(percent_key, 0.0)
    st.session_state.setdefault(mode_key, "Protect profit")
    st.session_state.setdefault(
        applied_key,
        _safe_float(st.session_state[percent_key])
        if st.session_state[mode_key] == "Protect profit" else 0.0,
    )
    st.session_state.setdefault(
        base_key,
        max(subtotal - abs(_safe_float(st.session_state.get(amount_key))), 0.0)
        if st.session_state[mode_key] == "Protect profit" else subtotal,
    )
    st.session_state[amount_key] = abs(_safe_float(st.session_state.get(amount_key)))

    if mode_col is None or amount_col is None or percent_col is None:
        mode_col, amount_col, percent_col = st.columns(3)
    mode_col.selectbox(
        "Commission treatment", COMMISSION_MODES, key=mode_key,
        help="Protect profit adds the commission to item margins. Deduct from profit keeps client prices unchanged.",
        on_change=_reprice_commission,
        args=(prefix, amount_key, grid_key, editor_key, discount_key, "percent"))
    amount_col.number_input(
        "Commission Amount (SAR)", min_value=0.0, step=100.0,
        format="%.2f", key=amount_key,
        help="Editable cashback amount. Commission % updates automatically.",
        on_change=_reprice_commission,
        args=(prefix, amount_key, grid_key, editor_key, discount_key, "amount"))
    percent_col.number_input(
        "Commission %", min_value=0.0, max_value=1000.0,
        step=1.0, format="%.2f", key=percent_key,
        help="Editable cashback percentage. Commission Amount updates automatically.",
        on_change=_reprice_commission,
        args=(prefix, amount_key, grid_key, editor_key, discount_key, "percent"))
    return (
        abs(_safe_float(st.session_state.get(amount_key))),
        max(_safe_float(st.session_state.get(percent_key)), 0.0),
        st.session_state.get(mode_key, "Protect profit"),
    )


def _set_offer_active_tab(state_key: str, tab: str):
    st.session_state[state_key] = tab


def _offer_tab_selector(project_id: int, approved: bool) -> str:
    tabs = ("BoQ", "Tracking", "Finance")
    state_key = f"offer_active_tab_{project_id}"
    active = st.session_state.get(state_key, "BoQ")
    if active not in tabs or (not approved and active != "BoQ"):
        active = "BoQ"
        st.session_state[state_key] = active

    tab_css = ["<style>"]
    for label in tabs:
        slug = label.lower()
        selector = f'[class*="st-key-offer_tab_{slug}_{project_id}"] button'
        text_selector = f'[class*="st-key-offer_tab_{slug}_{project_id}"] button p'
        locked = label != "BoQ" and not approved
        if label == active:
            tab_css.append(
                f"{selector} {{ background: #17324d !important; border-color: #17324d !important; "
                "color: #fff !important; }}"
            )
            tab_css.append(f"{text_selector} {{ color: #fff !important; }}")
        elif locked:
            tab_css.append(
                f"{selector} {{ background: #f1f4f7 !important; color: #8a96a3 !important; "
                "border-color: #d5dde5 !important; opacity: 0.72 !important; }}"
            )
            tab_css.append(f"{text_selector} {{ color: #8a96a3 !important; }}")
    tab_css.append("</style>")
    st.markdown("\n".join(tab_css), unsafe_allow_html=True)

    cols = st.columns(3, gap="small")
    for label, col in zip(tabs, cols):
        locked = label != "BoQ" and not approved
        key = f"offer_tab_{label.lower()}_{project_id}"
        col.button(label, key=key, disabled=locked, width="stretch",
                   type="primary" if active == label else "secondary",
                   on_click=_set_offer_active_tab, args=(state_key, label))

    if not approved:
        st.caption("Tracking and Finance are enabled after the offer is approved.")
    return active


def _fmt_tracking_stamp(value) -> str:
    raw = _text(value)
    if not raw:
        return ""
    try:
        return dt.datetime.fromisoformat(raw).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return raw.replace("T", " ")[:16]


def _tracking_center_text(col, value):
    text = html.escape(_text(value)) or "&nbsp;"
    col.markdown(f"<div class='tracking-center-cell'>{text}</div>", unsafe_allow_html=True)


def _tracking_keys(lid: int, key_name: str):
    value_key = f"trk_status_{key_name}_{lid}"
    stamp_key = f"trk_{key_name}_stamp_{lid}"
    qty_key = f"trkqty_{key_name}_{lid}"
    return value_key, stamp_key, qty_key


def _bounded_tracking_qty(value, max_qty: float) -> float:
    qty = max(_safe_float(value), 0.0)
    if max_qty > 0:
        qty = min(qty, max_qty)
    return round(qty, 4)


def _set_tracking_qty(value_key: str, stamp_key: str, qty_key: str,
                      qty, max_qty: float):
    qty = _bounded_tracking_qty(qty, max_qty)
    st.session_state[qty_key] = qty
    checked = qty > 0
    st.session_state[value_key] = checked
    if checked and not st.session_state.get(stamp_key):
        st.session_state[stamp_key] = dt.datetime.now().isoformat(timespec="minutes")
    elif not checked:
        st.session_state[stamp_key] = ""


def _open_tracking_qty_prompt(value_key: str, stamp_key: str, qty_key: str,
                              action: str, description: str, full_qty: float,
                              region_key: str = ""):
    st.session_state.tracking_qty_prompt = {
        "value_key": value_key,
        "stamp_key": stamp_key,
        "qty_key": qty_key,
        "action": action,
        "description": description,
        "full_qty": _bounded_tracking_qty(full_qty, full_qty),
        "region_key": region_key,
    }
    st.session_state["tracking_prompt_qty"] = _bounded_tracking_qty(full_qty, full_qty)


def _clear_tracking_qty_prompt():
    st.session_state.pop("tracking_qty_prompt", None)
    st.session_state.pop("tracking_prompt_qty", None)


def _render_tracking_qty_prompt():
    prompt = st.session_state.get("tracking_qty_prompt")
    if not prompt:
        return
    action = _text(prompt.get("action")).title()
    desc = _text(prompt.get("description"), "this item")
    full_qty = _bounded_tracking_qty(prompt.get("full_qty"), prompt.get("full_qty") or 0.0)

    region_key = _text(prompt.get("region_key"))
    _regions = repo.regions() if (action.lower() == "received" and region_key) else []

    def prompt_body():
        st.write(f"**{desc}**")
        st.write(
            f"Is the total quantity ({full_qty:g}) {action.lower()}, "
            "or only part of it?"
        )
        if _regions:
            st.selectbox("Receiving region", ["", *_regions],
                         key=region_key, label_visibility="visible")
        if st.button(f"Full quantity ({full_qty:g})", key="tracking_prompt_full",
                     type="primary", width="stretch"):
            _set_tracking_qty(prompt["value_key"], prompt["stamp_key"], prompt["qty_key"],
                              full_qty, full_qty)
            _clear_tracking_qty_prompt()
            st.rerun()
        partial_qty = st.number_input(
            "Partial quantity",
            min_value=0.0,
            max_value=full_qty,
            step=1.0,
            format="%.2f",
            key="tracking_prompt_qty",
        )
        b1, b2 = st.columns(2)
        if b1.button(f"Use partial", key="tracking_prompt_partial",
                     width="stretch"):
            _set_tracking_qty(prompt["value_key"], prompt["stamp_key"], prompt["qty_key"],
                              partial_qty, full_qty)
            _clear_tracking_qty_prompt()
            st.rerun()
        if b2.button("Cancel", key="tracking_prompt_cancel", width="stretch"):
            _clear_tracking_qty_prompt()
            st.rerun()

    if hasattr(st, "dialog"):
        @st.dialog(f"{action} quantity")
        def tracking_qty_dialog():
            prompt_body()
        tracking_qty_dialog()
    else:
        with st.container(border=True):
            prompt_body()


def _toggle_tracking_status(value_key: str, stamp_key: str,
                            qty_key: str | None = None, full_qty: float = 0.0):
    checked = not bool(st.session_state.get(value_key))
    st.session_state[value_key] = checked
    if checked and not st.session_state.get(stamp_key):
        st.session_state[stamp_key] = dt.datetime.now().isoformat(timespec="minutes")
    elif not checked:
        st.session_state[stamp_key] = ""
    if qty_key:
        st.session_state[qty_key] = _bounded_tracking_qty(full_qty, full_qty) if checked else 0.0


def _handle_tracking_status_click(value_key: str, stamp_key: str,
                                  qty_key: str | None = None,
                                  full_qty: float = 0.0,
                                  action: str = "",
                                  description: str = "",
                                  region_key: str = ""):
    if qty_key and not bool(st.session_state.get(value_key)):
        current_qty = _bounded_tracking_qty(st.session_state.get(qty_key), full_qty)
        if current_qty <= 0:
            _open_tracking_qty_prompt(value_key, stamp_key, qty_key,
                                      action, description, full_qty, region_key)
            return
    _toggle_tracking_status(value_key, stamp_key, qty_key, full_qty)


def _sync_tracking_qty_status(qty_key: str, value_key: str, stamp_key: str, max_qty: float):
    _set_tracking_qty(value_key, stamp_key, qty_key,
                      st.session_state.get(qty_key), max_qty)


def _tracking_status_cell(col, lid: int, key_name: str, current: bool, stamp_value,
                          full_qty: float | None = None, current_qty=0.0,
                          description: str = "", region_key: str = ""):
    value_key, stamp_key, qty_key = _tracking_keys(lid, key_name)
    if value_key not in st.session_state:
        st.session_state[value_key] = bool(current)
    if stamp_key not in st.session_state:
        st.session_state[stamp_key] = _text(stamp_value)
    max_qty = _bounded_tracking_qty(full_qty or 0.0, full_qty or 0.0)
    if full_qty is not None:
        default_qty = _bounded_tracking_qty(current_qty, max_qty)
        if bool(st.session_state[value_key]) and default_qty <= 0 and max_qty > 0:
            default_qty = max_qty
        if qty_key not in st.session_state:
            st.session_state[qty_key] = default_qty
        st.session_state[qty_key] = _bounded_tracking_qty(st.session_state.get(qty_key), max_qty)

    checked = bool(st.session_state[value_key])
    stamp = _text(st.session_state.get(stamp_key))
    if checked and not stamp:
        stamp = dt.datetime.now().isoformat(timespec="minutes")
        st.session_state[stamp_key] = stamp
    elif not checked:
        stamp = ""
        st.session_state[stamp_key] = ""

    btn_state = "on" if checked else "off"
    if full_qty is not None and checked:
        tracked_qty = _bounded_tracking_qty(st.session_state.get(qty_key), max_qty)
        if max_qty > 0 and 0 < tracked_qty < max_qty:
            btn_state = "partial"
    btn_key = f"trkbtn_{btn_state}_{key_name}_{lid}"
    label = "✓" if checked else " "
    col.button(label, key=btn_key, disabled=not can("tracking"), width="stretch",
               on_click=_handle_tracking_status_click,
               args=(value_key, stamp_key,
                     qty_key if full_qty is not None else None,
                     _bounded_tracking_qty(full_qty or 0.0, full_qty or 0.0),
                     key_name, description, region_key))
    stamp_text = html.escape(_fmt_tracking_stamp(stamp)) if stamp else "&nbsp;"
    col.markdown(f"<div class='tracking-stamp'>{stamp_text}</div>", unsafe_allow_html=True)
    return checked, stamp


def _tracking_qty_cell(col, lid: int, key_name: str, current_qty, line_qty: float) -> float:
    value_key, stamp_key, qty_key = _tracking_keys(lid, key_name)
    max_qty = max(float(line_qty or 0.0), 0.0)
    default_qty = _bounded_tracking_qty(current_qty, max_qty)
    if bool(st.session_state.get(value_key)) and default_qty <= 0 and max_qty > 0:
        default_qty = max_qty
    if qty_key not in st.session_state:
        st.session_state[qty_key] = default_qty
    st.session_state[qty_key] = _bounded_tracking_qty(st.session_state.get(qty_key), max_qty)
    col.number_input(
        f"{key_name} quantity",
        min_value=0.0,
        max_value=max_qty,
        step=1.0,
        format="%.2f",
        key=qty_key,
        label_visibility="collapsed",
        disabled=not can("tracking"),
        on_change=_sync_tracking_qty_status,
        args=(qty_key, value_key, stamp_key, max_qty),
    )
    return _bounded_tracking_qty(st.session_state.get(qty_key), max_qty)


def _render_finance_tab(project_id: int, grand_total: float):
    """Two side-by-side tables for an offer: client payments/invoices and purchases/costs."""
    if not can("finance"):
        st.info("🔒 Your role doesn't have Finance access. "
                "An owner can grant it in Settings → Roles & permissions.")
        return

    gt = float(grand_total or 0.0)
    commission = abs(float(repo.project_meta(project_id).get("CommissionAmount") or 0.0))
    # Cache the editor sources once per offer so in-progress edits stay consistent
    # (we never reload from the DB mid-edit; we just persist changes back to it).
    pay_src_key, pur_src_key = f"fin_pay_src_{project_id}", f"fin_pur_src_{project_id}"
    sig_key = f"fin_sig_{project_id}"
    if pay_src_key not in st.session_state or pur_src_key not in st.session_state:
        pays, purs = repo.get_finance(project_id)
        st.session_state[pay_src_key] = pd.DataFrame(
            [{"Description": r["Description"], "Amount (SAR)": r["AmountSAR"],
              "Invoice #": r["InvoiceNo"] or ""} for r in pays]
            or [{"Description": d, "Amount (SAR)": 0.0, "Invoice #": ""}
                for d in ("Downpayment", "Payment 1", "Payment 2")])
        st.session_state[pur_src_key] = pd.DataFrame(
            [{"Description": r["Description"], "Cost (SAR)": r["AmountSAR"],
              "PO #": r["PORef"] or ""} for r in purs]
            or [{"Description": "", "Cost (SAR)": 0.0, "PO #": ""}])

    _pay_wkey = f"fin_pay_{project_id}"
    _pur_wkey = f"fin_pur_{project_id}"

    col_pay, col_pur = st.columns(2)
    with col_pay:
        st.markdown("#### 💵 Payments / Invoices")
        pay_cfg = {
            "Description": st.column_config.TextColumn("Payment Description", width="medium"),
            "Amount (SAR)": st.column_config.NumberColumn("Amount (SAR)", format="%.2f", min_value=0.0),
            "Invoice #": st.column_config.TextColumn("Invoice #", help="Invoice number (free text)"),
        }
        pay_edit = st.data_editor(st.session_state[pay_src_key], column_config=pay_cfg,
                                  num_rows="dynamic", hide_index=True, width="stretch",
                                  key=_pay_wkey)
        collected = pay_edit["Amount (SAR)"].map(calc._num).sum()
        remaining = gt - collected
        pct = (collected / gt * 100) if gt else 0.0
        _fin_bubble("Collected", f"SAR {collected:,.2f}",
                    [f"{pct:.0f}% collected", f"of SAR {gt:,.2f}"],
                    "Remaining / Due", f"SAR {remaining:,.2f}",
                    positive=remaining >= 0)

    with col_pur:
        st.markdown("#### 🧾 Purchases / Costs")
        pur_cfg = {
            "Description": st.column_config.TextColumn("Dispense Description", width="medium"),
            "Cost (SAR)": st.column_config.NumberColumn("Cost (SAR)", format="%.2f", min_value=0.0),
            "PO #": st.column_config.TextColumn("PO #", help="Purchase-order reference (free text)"),
        }
        pur_edit = st.data_editor(st.session_state[pur_src_key], column_config=pur_cfg,
                                  num_rows="dynamic", hide_index=True, width="stretch",
                                  key=_pur_wkey)
        cost_total = pur_edit["Cost (SAR)"].map(calc._num).sum()
        vat = gt * calc.VAT_RATE
        internal_cost = cost_total + commission
        net_profit = gt - internal_cost - vat
        markup = (gt / internal_cost) if internal_cost > 0 else None
        margin_pct = (net_profit / gt * 100) if gt else 0.0
        markup_txt = f"Markup x{markup:.2f}" if markup else "Markup -"
        _fin_bubble("🧾 Cost (POs + Commission)", f"SAR {internal_cost:,.2f}",
                    [markup_txt, f"Margin {margin_pct:.1f}%",
                     f"Commission SAR {commission:,.2f}",
                     f"VAT ({calc.VAT_RATE * 100:g}%) SAR {vat:,.2f}"],
                    "💰 Net Profit", f"SAR {net_profit:,.2f}",
                    positive=net_profit >= 0)

    if st.button("💾 Save Finance", type="primary"):
        repo.save_finance(project_id, pay_edit.to_dict("records"), pur_edit.to_dict("records"))
        # Clear source and widget state so next render reloads cleanly from DB.
        for k in (pay_src_key, pur_src_key, _pay_wkey, _pur_wkey):
            st.session_state.pop(k, None)
        st.toast("Finance saved", icon="💾")
        st.rerun()


def _company_dict():
    return {
        "name": repo.get_setting("company_name") or "Company",
        "vat_number": repo.get_setting("company_vat_number") or "",
        "cr_number": repo.get_setting("company_cr_number") or "",
        "color": repo.get_setting("company_brand_color") or "#002060",
    }


def _excel_bytes(df_map: dict):
    """df_map: {sheet_name: DataFrame} -> xlsx bytes. Money columns (… SAR) use
    accounting number format (1,234.00); counts stay whole; % keeps one decimal."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for sheet, d in df_map.items():
            name = sheet[:31]
            d.to_excel(w, index=False, sheet_name=name)
            ws = w.sheets[name]
            for ci, col in enumerate(d.columns, start=1):
                cs = str(col)
                if cs.endswith("SAR"):
                    fmt = "#,##0.00"
                elif cs.endswith("%"):
                    fmt = "0.0"
                elif cs in ("Qty", "Count"):
                    fmt = "#,##0"
                else:
                    continue
                for row in range(2, ws.max_row + 1):       # row 1 is the header
                    ws.cell(row=row, column=ci).number_format = fmt
    return buf.getvalue()


def _catalogue_zip_bytes(catalogue_df: pd.DataFrame) -> bytes:
    excel = _excel_bytes({"Catalogue": catalogue_df.drop(columns=["ItemID"], errors="ignore")})
    manifest = (
        "ProQuote catalogue backup\n"
        "Contains catalogue.xlsx. Restore from Settings > Backup & Restore.\n"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_STORED) as zf:
        zf.writestr("catalogue.xlsx", excel)
        zf.writestr("catalogue-backup.txt", manifest)
    return buf.getvalue()


def _catalogue_df_from_zip(uploaded_bytes: bytes) -> pd.DataFrame:
    if not uploaded_bytes:
        raise ValueError("Uploaded catalogue backup is empty.")
    try:
        with zipfile.ZipFile(io.BytesIO(uploaded_bytes)) as zf:
            names = [
                info.filename
                for info in zf.infolist()
                if not info.is_dir() and info.filename.lower().endswith(".xlsx")
            ]
            if not names:
                raise ValueError("Catalogue backup ZIP does not contain an .xlsx file.")
            preferred = "catalogue.xlsx" if "catalogue.xlsx" in names else names[0]
            with zf.open(preferred) as src:
                return pd.read_excel(io.BytesIO(src.read()))
    except zipfile.BadZipFile as exc:
        raise ValueError("Not a valid catalogue ZIP backup.") from exc


def _db_cache_stamp():
    """Cache key that changes whenever the active database is written to."""
    if os.environ.get("DATABASE_URL", "").startswith("postgres"):
        import db_postgres
        return ("pg", db_postgres.write_epoch())
    # SQLite: use file mtime + WAL size
    stamp = []
    for path in (db.DB_PATH, f"{db.DB_PATH}-wal"):
        try:
            s = os.stat(path)
            stamp.append((s.st_mtime_ns, s.st_size))
        except OSError:
            stamp.append((0, 0))
    return tuple(stamp)


@st.cache_data(show_spinner=False, max_entries=12)
def _cached_report_dataset(ds_name: str, include_archived: bool, db_stamp):
    return reports.DATASETS[ds_name]["builder"](include_archived=include_archived)


@st.cache_data(show_spinner=False, max_entries=12)
def _cached_offers_df(include_archived: bool, db_stamp):
    return reports.offers_df(include_archived=include_archived)


@st.cache_data(show_spinner=False, max_entries=12)
def _cached_finance_df(include_archived: bool, db_stamp):
    return reports.finance_df(include_archived=include_archived)


@st.cache_data(show_spinner=False, max_entries=48)
def _cached_project_grid(project_id: int, sheet_name: str | None, db_stamp):
    return repo.load_project_grid(project_id, sheet_name)


@st.cache_data(show_spinner=False, max_entries=100)
def _cached_search_catalog(term: str, limit: int, db_stamp):
    return repo.search_catalog(term, limit=limit)


def _build_project_index(projects):
    if projects.empty:
        return projects, []

    projects = projects.copy()
    groups = {}
    fam_keys = []
    for row in projects.itertuples(index=False):
        fam = repo.family_key(row.OfferNo, row.ProjectName)
        fam_keys.append(fam)
        rev = int(row.RevisionNo or 0) if pd.notna(row.RevisionNo) else 0
        option = _text(row.OptionLabel)
        sort_key = (rev, bool(option), option)
        group = groups.setdefault(fam, {
            "fam": fam, "offer_nos": set(), "project_names": set(),
            "sales_people": set(), "presales_people": set(), "project_managers": set(),
            "revision_counts": {}, "approved": False,
            "created_date": "", "updated_date": "",
            "rep_sort": None, "base": "", "client": "", "sales": "",
            "system": "", "region": "",
        })
        offer_no = _text(row.OfferNo)
        project_name = _text(row.ProjectName)
        # The matching-results row represents the whole offer family, which can
        # contain several options. Keep its Project column uncluttered; option
        # labels are shown after opening the offer in Revisions & options.
        project_display_name = repo.project_name_with_option(project_name, "", option)
        if offer_no:
            group["offer_nos"].add(offer_no)
        if project_display_name:
            group["project_names"].add(project_display_name)
        sales_person = _text(row.SalesPerson).strip()
        presales_person = _text(row.PresalesEngineer).strip()
        project_manager = _text(row.ProjectManager).strip()
        if sales_person:
            group["sales_people"].add(sales_person)
        if presales_person:
            group["presales_people"].add(presales_person)
        if project_manager:
            group["project_managers"].add(project_manager)
        group["revision_counts"][rev] = group["revision_counts"].get(rev, 0) + 1
        group["approved"] = group["approved"] or bool(row.Approved or 0)
        created_date = _text(row.CreationDate)
        updated_date = _text(row.UpdatedDate) or created_date
        if created_date and (
            not group["created_date"] or created_date < group["created_date"]
        ):
            group["created_date"] = created_date
        group["updated_date"] = max(group["updated_date"], updated_date)
        if group["rep_sort"] is None or sort_key >= group["rep_sort"]:
            group["rep_sort"] = sort_key
            group["base"] = _text(row.BaseName) or repo.base_name(project_name or "Offer")
            group["client"] = _text(row.ClientName)
            group["sales"] = _text(row.SalesPerson)
            group["system"] = repo.system_name(_text(row.System))
            group["region"] = _text(row.Region)

    projects["_fam"] = fam_keys
    fams = []
    for group in groups.values():
        offer_nos = sorted(group["offer_nos"])
        project_names = sorted(group["project_names"])
        base = group["base"]
        client = group["client"]
        offer_label = repo.base_name(offer_nos[0]) if offer_nos else ""
        project_label = ", ".join(project_names) if project_names else base
        fams.append({
            "fam": group["fam"], "base": base, "client": client,
            "offer_nos": offer_nos, "project_names": project_names,
            "project_label": project_label, "offer_label": offer_label,
            "sales": group["sales"], "system": group["system"], "region": group["region"],
            "name_search": " ".join([base, client] + project_names).lower(),
            "offer_search": " ".join(offer_nos).lower(),
            "sales_people": group["sales_people"],
            "presales_people": group["presales_people"],
            "project_managers": group["project_managers"],
            "n_rev": len(group["revision_counts"]),
            "n_opt": max(group["revision_counts"].values()),
            "approved": group["approved"],
            "created_date": group["created_date"],
            "updated_date": group["updated_date"],
        })
    fams.sort(key=lambda f: (f["updated_date"], f["created_date"]), reverse=True)
    return projects, fams


@st.cache_data(show_spinner=False, max_entries=4)
def _cached_project_index(db_stamp):
    people = repo.project_people()
    if people.empty:
        people = pd.DataFrame([{
            "SalesPerson": None, "PresalesEngineer": None, "ProjectManager": None
        }])
    return people, []


@st.cache_data(show_spinner=False, max_entries=100)
def _cached_project_search(name: str, offer: str, sales: tuple,
                           presales: tuple, project_managers: tuple, db_stamp):
    return _build_project_index(
        repo.search_projects(name, offer, sales, presales, project_managers)
    )


@st.cache_data(show_spinner=False, max_entries=48)
def _cached_project_bundle(project_id: int, family_project_ids: tuple, db_stamp):
    return repo.load_project_bundle(project_id, family_project_ids)


def _render_report_builder(company):
    ds_name = st.selectbox("Dataset", list(reports.DATASETS), key="rep_ds")
    meta = reports.DATASETS[ds_name]
    df = _cached_report_dataset(ds_name, False, _db_cache_stamp())
    if df.empty:
        st.info("No data available for this dataset yet.")
        return

    with st.expander("Filters", expanded=True):
        selections, fcols = {}, [f for f in meta["filters"] if f in df.columns]
        cols = st.columns(3)
        for i, fcol in enumerate(fcols):
            opts = _report_filter_options(df, fcol)
            label = _report_filter_label(fcol)
            help_text = None
            if fcol == "Sales Person":
                help_text = ("Filters by the person assigned in the project's Sales Person field. "
                             "The list includes Sales, Pre-Sales, Project Manager, and Top "
                             "Management users who can act as Sales.")
            selections[fcol] = cols[i % 3].multiselect(
                label, opts, key=f"rf_{ds_name}_{fcol}", help=help_text)
        dcol = meta.get("date")
        date_from = date_to = None
        if dcol and dcol in df.columns and pd.to_datetime(df[dcol], errors="coerce").notna().any():
            dc1, dc2 = st.columns(2)
            date_from = dc1.date_input("From date", value=None, key=f"rf_from_{ds_name}")
            date_to = dc2.date_input("To date", value=None, key=f"rf_to_{ds_name}")

    filtered = reports.apply_filters(df, selections, meta.get("date"), date_from, date_to)

    with st.expander("Summarize (optional — group & total)"):
        group_by = st.multiselect("Group by", [c for c in (meta["filters"] + ["Month"]) if c in df.columns],
                                   key=f"rg_{ds_name}")
        metrics = st.multiselect("Metrics", meta["metrics"], default=meta["metrics"], key=f"rm_{ds_name}")

    if group_by:
        result = reports.aggregate(filtered, group_by, metrics)
        totals = reports.totals_row(result, metrics)
    else:
        show = [c for c in meta["show"] if c in filtered.columns]
        result = filtered[show].reset_index(drop=True)
        totals = reports.totals_row(filtered, meta["metrics"])

    st.caption(f"{len(result)} row(s)")
    # Money columns (… SAR) show accounting format; counts stay integers, % keeps 1 dp.
    rep_cfg = {}
    for c in result.columns:
        cs = str(c)
        if cs.endswith("SAR"):
            rep_cfg[c] = st.column_config.NumberColumn(cs, format="accounting")
        elif cs.endswith("%"):
            rep_cfg[c] = st.column_config.NumberColumn(cs, format="%.1f")
        elif cs in ("Qty", "Count"):
            rep_cfg[c] = st.column_config.NumberColumn(cs, format="%d")
    st.dataframe(result, width="stretch", hide_index=True, column_config=rep_cfg)
    if totals:
        def _ft(k, v):
            return f"{v:,.0f}" if str(k) in ("Qty", "Count") else f"{v:,.2f}"
        st.markdown("  ·  ".join(f"**{k}:** {_ft(k, v)}" for k, v in totals.items()))

    # ---- Exports ----
    sub = [f"Generated {dt.date.today().isoformat()} · {ds_name}"]
    active = [f"{_report_filter_label(k)}: {', '.join(v)}"
              for k, v in selections.items() if v]
    if date_from or date_to:
        active.append(f"Date: {date_from or '…'} → {date_to or '…'}")
    if group_by:
        active.append(f"Grouped by: {', '.join(group_by)}")
    if active:
        sub.append("Filters — " + "  |  ".join(active))

    e1, e2 = st.columns(2)
    e2.download_button("⬇️ Export Excel", _excel_bytes({"Report": result}),
                       file_name=f"report_{ds_name.split(' ')[0].lower()}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       width="stretch")
    if e1.button("📄 Build PDF report", width="stretch"):
        out = os.path.join(db.DATA_DIR, "_last_report.pdf")
        pdf_export.generate_report_pdf(out, f"{ds_name} report", sub, result, totals, company)
        with open(out, "rb") as f:
            st.session_state["report_pdf_bytes"] = f.read()
    if st.session_state.get("report_pdf_bytes"):
        st.download_button("⬇️ Download PDF report", st.session_state["report_pdf_bytes"],
                           file_name="report.pdf", mime="application/pdf")


def _dashboard_number(value, unit="SAR"):
    """Compact, readable dashboard number (avoids scientific notation such as 1e8)."""
    value = float(value or 0)
    if unit == "%":
        return f"{value:,.1f}%"
    if unit == "count":
        return f"{int(round(value)):,}"
    magnitude = abs(value)
    if magnitude >= 1_000_000_000:
        text = f"{value / 1_000_000_000:,.1f}B"
    elif magnitude >= 1_000_000:
        text = f"{value / 1_000_000:,.1f}M"
    elif magnitude >= 1_000:
        text = f"{value / 1_000:,.1f}K"
    else:
        text = f"{value:,.0f}"
    return f"{text} SAR" if unit == "SAR" else text


def _render_dashboard(company):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    stamp = _db_cache_stamp()
    offers = _cached_offers_df(False, stamp)
    finance = _cached_finance_df(False, stamp)
    if offers.empty:
        st.info("No offers yet — nothing to chart.")
        return

    # Finance is the offers dataset plus collection/purchase metrics, so it can drive
    # every chart when available without maintaining two separate filter states.
    data = finance.copy() if not finance.empty else offers.copy()
    brand = company.get("color") or "#002060"
    metric_defs = {
        "Quoted value": ("Grand Total SAR", "sum", "SAR"),
        "Offer records": ("ProjectID", "count", "count"),
        "Gross profit": ("Gross Profit SAR", "sum", "SAR"),
        "Average margin": ("Margin %", "mean", "%"),
        "Cost": ("Cost SAR", "sum", "SAR"),
        "Collected": ("Collected SAR", "sum", "SAR"),
        "Remaining / due": ("Remaining SAR", "sum", "SAR"),
        "PO spend": ("PO Spend SAR", "sum", "SAR"),
        "Net profit": ("Net Profit SAR", "sum", "SAR"),
    }
    metric_defs = {label: spec for label, spec in metric_defs.items() if spec[0] in data.columns}
    group_defs = {
        "Month": "Month",
        "Client": "Client",
        "Assigned as Sales": "Sales Person",
        "Pre-sales Engineer": "Pre-sales",
        "Project Manager": "Project Mgr",
        "System": "System",
        "Status": "Status",
    }
    group_defs = {label: col for label, col in group_defs.items() if col in data.columns}
    filter_cols = [
        col for col in ("Status", "Client", "Sales Person", "Pre-sales", "Project Mgr", "System")
        if col in data.columns
    ]

    st.caption(
        "Build one focused chart at a time. **Quoted value** is the total value of the "
        "filtered offer records; **Gross profit** is discounted sales value before VAT minus cost."
    )
    with st.form("dashboard_builder_form"):
        st.markdown("##### Choose what you want to see")
        c1, c2, c3, c4 = st.columns(4)
        metric_label = c1.selectbox("Metric", list(metric_defs), key="dash_metric")
        group_label = c2.selectbox("Group by", list(group_defs), key="dash_group")
        chart_type = c3.selectbox(
            "Chart type", ["Column", "Horizontal bar", "Line"], key="dash_chart_type")
        chart_size = c4.selectbox("Chart size", ["Compact", "Medium"], key="dash_chart_size")
        p1, p2, p3 = st.columns([1, 1, 2])
        top_n = p1.slider("Groups / periods", 3, 20, 10, key="dash_top_n")
        show_values = p2.checkbox("Show values", value=True, key="dash_show_values")
        custom_title = p3.text_input(
            "Optional chart title", key="dash_title", placeholder="Leave blank for an automatic title")

        with st.expander("Filters (optional)", expanded=False):
            selections = {}
            fcols = st.columns(3)
            for i, column in enumerate(filter_cols):
                selections[column] = fcols[i % 3].multiselect(
                    _report_filter_label(column),
                    _report_filter_options(data, column),
                    key=f"dash_filter::{column}",
                )
            d1, d2 = st.columns(2)
            date_from = d1.date_input("From date", value=None, key="dash_date_from")
            date_to = d2.date_input("To date", value=None, key="dash_date_to")

        generated = st.form_submit_button(
            "Generate dashboard", type="primary", icon="📊", width="stretch")

    if generated:
        st.session_state["dashboard_config"] = {
            "metric": metric_label,
            "group": group_label,
            "chart_type": chart_type,
            "chart_size": chart_size,
            "top_n": top_n,
            "show_values": show_values,
            "title": custom_title.strip(),
            "filters": selections,
            "date_from": date_from,
            "date_to": date_to,
        }
        st.session_state.pop("dashboard_pdf_bytes", None)

    config = st.session_state.get("dashboard_config")
    if not config:
        st.info("Choose a metric and grouping above, then click **Generate dashboard**.")
        return
    if config.get("metric") not in metric_defs or config.get("group") not in group_defs:
        st.session_state.pop("dashboard_config", None)
        st.info("The available data changed. Please generate the dashboard again.")
        return

    filtered = reports.apply_filters(
        data,
        config.get("filters", {}),
        "Date",
        config.get("date_from"),
        config.get("date_to"),
    )
    if filtered.empty:
        st.warning("No records match these filters. Change the filters and generate again.")
        return

    quoted = float(filtered["Grand Total SAR"].sum())
    approved = float(filtered.loc[filtered["Status"] == "Approved", "Grand Total SAR"].sum())
    gross_profit = float(filtered["Gross Profit SAR"].sum())
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Offer records", f"{len(filtered):,}", help="Includes revisions and options.")
    k2.metric("Quoted value", _dashboard_number(quoted))
    k3.metric("Approved value", _dashboard_number(approved))
    k4.metric("Gross profit", _dashboard_number(gross_profit))

    metric_col, aggregation, unit = metric_defs[config["metric"]]
    group_col = group_defs[config["group"]]
    work = filtered.copy()
    valid_group = work[group_col].notna() & work[group_col].astype(str).str.strip().ne("")
    if group_col == "Month":
        valid_group &= work[group_col].astype(str).ne("NaT")
    work = work[valid_group]
    if work.empty:
        st.warning(f"The matching records have no {config['group']} value to chart.")
        return

    if aggregation == "count":
        grouped = work.groupby(group_col).size()
    else:
        grouped = work.groupby(group_col)[metric_col].agg(aggregation)
    grouped = grouped.dropna()
    if group_col == "Month":
        grouped = grouped.sort_index().tail(int(config["top_n"]))
    else:
        grouped = grouped.sort_values(ascending=False).head(int(config["top_n"]))
    if grouped.empty:
        st.warning("There is no numeric data for this chart.")
        return

    labels = [str(value) for value in grouped.index]
    values = [float(value) for value in grouped.values]
    title = config.get("title") or f"{config['metric']} by {config['group']}"
    figsize = (7.2, 2.9) if config.get("chart_size") == "Compact" else (9.0, 3.8)
    fig, ax = plt.subplots(figsize=figsize)
    color = "#37689B" if config["group"] == "Assigned as Sales" else brand
    chart_type = config.get("chart_type", "Column")
    bars = None
    if chart_type == "Horizontal bar":
        bars = ax.barh(labels, values, color=color)
        ax.invert_yaxis()
        value_axis = ax.xaxis
        ax.grid(axis="x", alpha=0.2)
    elif chart_type == "Line":
        ax.plot(labels, values, marker="o", linewidth=2.2, color=color)
        value_axis = ax.yaxis
        ax.grid(axis="y", alpha=0.2)
        ax.tick_params(axis="x", rotation=35)
    else:
        bars = ax.bar(labels, values, color=color)
        value_axis = ax.yaxis
        ax.grid(axis="y", alpha=0.2)
        ax.tick_params(axis="x", rotation=35)

    if unit == "SAR":
        value_axis.set_major_formatter(FuncFormatter(lambda value, _pos: _dashboard_number(value, "").strip()))
    elif unit == "%":
        value_axis.set_major_formatter(FuncFormatter(lambda value, _pos: f"{value:,.1f}%"))
    else:
        value_axis.set_major_formatter(FuncFormatter(lambda value, _pos: f"{value:,.0f}"))
    if bars is not None and config.get("show_values"):
        ax.bar_label(
            bars,
            labels=[_dashboard_number(value, unit) for value in values],
            padding=3,
            fontsize=8,
        )
    elif chart_type == "Line" and config.get("show_values"):
        for x, value in enumerate(values):
            ax.annotate(_dashboard_number(value, unit), (x, value),
                        xytext=(0, 7), textcoords="offset points", ha="center", fontsize=8)
    ax.set_title(title, fontsize=12, fontweight="bold", pad=12)
    ax.set_xlabel(config["group"])
    ax.set_ylabel(config["metric"] + (" (SAR)" if unit == "SAR" else ""))
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    fig.tight_layout()

    table = pd.DataFrame({config["group"]: labels, config["metric"]: values})
    chart_col, table_col = st.columns([2.4, 1], vertical_alignment="top")
    chart_col.pyplot(fig, width="content", clear_figure=False)
    table_cfg = {}
    if unit == "SAR":
        table_cfg[config["metric"]] = st.column_config.NumberColumn(
            config["metric"], format="accounting")
    elif unit == "%":
        table_cfg[config["metric"]] = st.column_config.NumberColumn(config["metric"], format="%.1f%%")
    else:
        table_cfg[config["metric"]] = st.column_config.NumberColumn(config["metric"], format="%d")
    table_col.caption(f"{len(filtered)} filtered record(s)")
    table_col.dataframe(table, hide_index=True, width="stretch", height=300, column_config=table_cfg)

    png = io.BytesIO()
    fig.savefig(png, format="png", dpi=150, bbox_inches="tight")
    png_bytes = png.getvalue()
    e1, e2, e3 = st.columns(3)
    e1.download_button(
        "Download chart PNG", png_bytes, file_name="dashboard_chart.png", mime="image/png",
        width="stretch", on_click="ignore")
    e2.download_button(
        "Export chart data", _excel_bytes({"Dashboard Data": table}),
        file_name="dashboard_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch", on_click="ignore")
    if e3.button("Build dashboard PDF", width="stretch"):
        import tempfile
        import shutil
        tmp = tempfile.mkdtemp(prefix="pq_dashboard_")
        try:
            chart_path = os.path.join(tmp, "dashboard_chart.png")
            with open(chart_path, "wb") as chart_file:
                chart_file.write(png_bytes)
            filter_summary = [
                f"{_report_filter_label(key)}: {', '.join(value)}"
                for key, value in config.get("filters", {}).items() if value
            ]
            if config.get("date_from") or config.get("date_to"):
                filter_summary.append(
                    f"Date: {config.get('date_from') or '…'} to {config.get('date_to') or '…'}")
            subtitle = [f"Generated {dt.date.today().isoformat()}"]
            subtitle.append(
                f"Records: {len(filtered):,} | Quoted: {_dashboard_number(quoted)} | "
                f"Approved: {_dashboard_number(approved)} | "
                f"Gross profit: {_dashboard_number(gross_profit)}"
            )
            if filter_summary:
                subtitle.append("Filters — " + " | ".join(filter_summary))
            selected_total = (
                float(work[metric_col].mean())
                if aggregation == "mean"
                else float(grouped.sum())
            )
            out = os.path.join(db.DATA_DIR, "_last_dashboard.pdf")
            pdf_export.generate_report_pdf(
                out,
                title,
                subtitle,
                table_df=table,
                totals={config["metric"]: selected_total},
                company=company,
                chart_paths=[chart_path],
            )
            with open(out, "rb") as pdf_file:
                st.session_state["dashboard_pdf_bytes"] = pdf_file.read()
        finally:
            shutil.rmtree(tmp, ignore_errors=True)
    if st.session_state.get("dashboard_pdf_bytes"):
        st.download_button(
            "Download dashboard PDF", st.session_state["dashboard_pdf_bytes"],
            file_name="dashboard.pdf", mime="application/pdf", on_click="ignore")
    plt.close(fig)


def _render_tracking_tab(project_id: int, sheet_name: str | None):
    track = repo.load_tracking(project_id, sheet_name).reset_index(drop=True)
    if track.empty:
        st.info("No line items to track.")
        return

    _render_tracking_qty_prompt()

    cols_meta = [("Description", 2.75), ("Brand", 1.0), ("Model", 1.05), ("Qty", 0.5),
                 ("PO Number", 1.25), ("Paid/Available", 1.1), ("Received", 0.9), ("Rec. Qty", 0.75),
                 ("Delivery Note", 1.25), ("Delivered", 0.9), ("Deliv. Qty", 0.75)]
    widths = [w for _, w in cols_meta]
    hdr = st.columns(widths)
    for (label, _), col in zip(cols_meta, hdr):
        col.markdown(f"<div class='tracking-header'>{label}</div>", unsafe_allow_html=True)
    st.markdown("<hr style='margin:0.1rem 0 0.3rem;border:none;"
                "border-top:1px solid #e6e6e6'>", unsafe_allow_html=True)

    collected = []
    for row_idx, (_, row) in enumerate(track.iterrows()):
        lid = int(row["LineID"])
        rc = st.columns(widths, vertical_alignment="center")
        j = 0
        rc[j].write(row["Description"] or ""); j += 1
        rc[j].markdown(f"<div style='text-align:center'>{html.escape(_text(row['Brand']))}</div>", unsafe_allow_html=True); j += 1
        rc[j].markdown(f"<div style='text-align:center'>{html.escape(_text(row['Model']))}</div>", unsafe_allow_html=True); j += 1
        line_qty = max(_safe_float(row.get("Qty")), 0.0)
        qty_text = f"{int(line_qty)}" if float(line_qty).is_integer() else f"{line_qty:.2f}"
        _tracking_center_text(rc[j], qty_text if pd.notna(row["Qty"]) else ""); j += 1
        po = rc[j].text_input("po", value=str(row.get("PONumber") or ""),
                              key=f"po_{lid}", label_visibility="collapsed"); j += 1
        paid, paid_at = _tracking_status_cell(rc[j], lid, "paid", bool(row["Paid"]), row.get("PaidAt")); j += 1
        rec_current_qty = _bounded_tracking_qty(row.get("ReceivedQty"), line_qty)
        rec_current = bool(row["Received"]) or rec_current_qty > 0
        region_key = f"trk_received_region_{lid}"
        if region_key not in st.session_state:
            st.session_state[region_key] = _text(row.get("ReceivedRegion") or "")
        rec_col = rc[j]
        rec, rec_at = _tracking_status_cell(rec_col, lid, "received", rec_current,
                                            row.get("ReceivedAt"), full_qty=line_qty,
                                            current_qty=rec_current_qty,
                                            description=_text(row["Description"]),
                                            region_key=region_key); j += 1
        rec_qty = _tracking_qty_cell(rc[j], lid, "received", rec_current_qty, line_qty); j += 1
        delivery_note = rc[j].text_input("delivery note", value=str(row.get("DeliveryNote") or ""),
                                         key=f"dn_{lid}", label_visibility="collapsed"); j += 1
        deliv_current_qty = _bounded_tracking_qty(row.get("DeliveredQty"), line_qty)
        deliv_current = bool(row["Delivered"]) or deliv_current_qty > 0
        deliv, deliv_at = _tracking_status_cell(rc[j], lid, "delivered", deliv_current,
                                                row.get("DeliveredAt"), full_qty=line_qty,
                                                current_qty=deliv_current_qty,
                                                description=_text(row["Description"])); j += 1
        deliv_qty = _tracking_qty_cell(rc[j], lid, "delivered", deliv_current_qty, line_qty); j += 1
        rec = rec or rec_qty > 0
        deliv = deliv or deliv_qty > 0
        _, rec_stamp_key, _ = _tracking_keys(lid, "received")
        _, deliv_stamp_key, _ = _tracking_keys(lid, "delivered")
        rec_at = _text(st.session_state.get(rec_stamp_key))
        deliv_at = _text(st.session_state.get(deliv_stamp_key))
        # Region shown below the timestamp when item is received; editable at any time
        _all_regions = repo.regions()
        if _all_regions:
            if rec:
                rec_col.selectbox(
                    "Region", ["", *_all_regions],
                    key=region_key,
                    label_visibility="collapsed",
                    disabled=not can("tracking"),
                )
            elif st.session_state.get(region_key):
                st.session_state[region_key] = ""
        rec_region = _text(st.session_state.get(region_key))
        collected.append((lid, paid, rec, deliv, po, delivery_note, paid_at, rec_at, deliv_at,
                          rec_qty, deliv_qty, rec_region))
        if row_idx < len(track) - 1:
            st.markdown("<div class='tracking-row-separator'></div>", unsafe_allow_html=True)

    tot = len(track)
    total_qty = sum(max(_safe_float(row.get("Qty")), 0.0) for _, row in track.iterrows())
    rec_total = sum(c[9] for c in collected)
    deliv_total = sum(c[10] for c in collected)
    st.markdown("<div style='height:1.75rem'></div>", unsafe_allow_html=True)  # gap before totals
    pc1, pc2, pc3 = st.columns(3)
    pc1.metric("Paid/Available", f"{sum(1 for c in collected if c[1])}/{tot}")
    pc2.metric("Received Qty", f"{rec_total:g}/{total_qty:g}")
    pc3.metric("Delivered Qty", f"{deliv_total:g}/{total_qty:g}")
    if can("tracking"):
        sig = repr(collected)                       # auto-save on any change (no button)
        sig_key = f"trk_sig_{project_id}_{sheet_name or 'all'}"
        if sig_key not in st.session_state:
            st.session_state[sig_key] = sig
        elif st.session_state[sig_key] != sig:
            repo.update_tracking(collected)
            st.session_state[sig_key] = sig
            st.toast("Tracking saved", icon="💾")
    else:
        st.caption("🔒 Your role can view tracking but not change it.")


def _render_audit_page():
    st.subheader("Audit")
    st.caption(
        "Immutable history of application data changes from the date auditing was enabled. "
        "Earlier changes cannot be reconstructed; password values are always redacted."
    )
    options = audit_log.filter_options()
    f1, f2, f3, f4 = st.columns([1.25, 1.0, 1.35, 2.0])
    username = f1.selectbox("User", ["", *options["users"]],
                            format_func=lambda v: v or "All users", key="audit_user")
    action = f2.selectbox(
        "Action", ["", *dict.fromkeys(["INSERT", "UPDATE", "DELETE", *options["actions"]])],
        format_func=lambda v: audit_log.ACTION_LABELS.get(v, "All actions"),
        key="audit_action",
    )
    entity = f3.selectbox(
        "Area", ["", *options["entities"]],
        format_func=lambda v: audit_log.ENTITY_LABELS.get(v, "All areas"),
        key="audit_entity",
    )
    search = f4.text_input(
        "Search audit details", key="audit_search",
        placeholder="Offer number, project, product, setting, username…",
    )

    d1, d2, d3, d4 = st.columns([1.0, 1.0, 1.0, 1.4], vertical_alignment="bottom")
    use_dates = d1.checkbox("Filter by date", key="audit_use_dates")
    today = dt.date.today()
    date_from = d2.date_input(
        "From", value=today - dt.timedelta(days=30), key="audit_date_from",
        disabled=not use_dates,
    )
    date_to = d3.date_input(
        "To", value=today, key="audit_date_to", disabled=not use_dates,
    )
    limit = d4.selectbox("Maximum rows", [100, 250, 500, 1000, 2000], index=1,
                         key="audit_limit")

    events, total = audit_log.query_events(
        username=username, action=action, entity_type=entity, search=search,
        date_from=date_from if use_dates else None,
        date_to=date_to if use_dates else None,
        limit=limit,
    )
    m1, m2, m3 = st.columns(3)
    m1.metric("Matching changes", f"{total:,}")
    m2.metric("Displayed", f"{len(events):,}")
    m3.metric("Users in audit", f"{len(options['users']):,}")
    if events.empty:
        st.info("No audit changes match these filters.")
        return

    display = pd.DataFrame({
        "ID": events["AuditID"],
        "When": events["EventAt"].astype(str).str.replace("T", " ", regex=False),
        "User": events["DisplayName"].fillna("").where(
            events["DisplayName"].fillna("").str.strip().ne(""), events["Username"]),
        "Action": events["Action"].map(audit_log.ACTION_LABELS).fillna(events["Action"]),
        "Area": events["EntityType"].map(audit_log.ENTITY_LABELS).fillna(events["EntityType"]),
        "Record": events["EntityID"].fillna(""),
        "Changes": [audit_log.describe_event(row.to_dict()) for _, row in events.iterrows()],
    })
    st.dataframe(display, hide_index=True, width="stretch", height=min(540, 38 + len(display) * 35))
    st.download_button(
        "⬇️ Download filtered audit CSV", display.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"ProQuote_Audit_{today.isoformat()}.csv", mime="text/csv",
    )

    by_id = {int(row["AuditID"]): row.to_dict() for _, row in events.iterrows()}
    selected_id = st.selectbox(
        "Inspect a change", list(by_id), key="audit_selected_id",
        format_func=lambda audit_id: (
            f"#{audit_id} · {by_id[audit_id]['EventAt'].replace('T', ' ')} · "
            f"{by_id[audit_id]['DisplayName'] or by_id[audit_id]['Username']} · "
            f"{audit_log.ACTION_LABELS.get(by_id[audit_id]['Action'], by_id[audit_id]['Action'])} "
            f"{audit_log.ENTITY_LABELS.get(by_id[audit_id]['EntityType'], by_id[audit_id]['EntityType'])}"
        ),
    )
    event = by_id[int(selected_id)]
    changes = audit_log.changes_frame(event)
    st.markdown(
        f"**Record:** `{event.get('EntityID') or '-'}` &nbsp; · &nbsp; "
        f"**User:** {event.get('DisplayName') or event.get('Username')} "
        f"(`{event.get('Username')}`)"
    )
    if changes.empty:
        st.info("No field-level difference is available for this event.")
    else:
        st.dataframe(changes, hide_index=True, width="stretch")
    with st.expander("Raw before / after snapshots"):
        old_col, new_col = st.columns(2)
        old_col.markdown("**Before**")
        old_col.json(audit_log.parse_snapshot(event.get("OldValues")))
        new_col.markdown("**After**")
        new_col.json(audit_log.parse_snapshot(event.get("NewValues")))


def _render_login():
    st.markdown("<div style='height:1.25rem'></div>", unsafe_allow_html=True)
    c = st.columns([1, 2, 1])[1]
    if os.path.exists(_LOGO):
        c.image(_LOGO, width="stretch")
    c.subheader("Sign in")
    if auth.user_count() == 0:
        c.info("First run - create the **owner** account (full access).")
        with c.form("create_owner"):
            u = st.text_input("Username")
            dn = st.text_input("Display name")
            p1 = st.text_input("Password", type="password")
            p2 = st.text_input("Confirm password", type="password")
            if st.form_submit_button("Create owner & sign in", type="primary"):
                if not u.strip() or not p1:
                    st.error("Username and password are required.")
                elif p1 != p2:
                    st.error("Passwords don't match.")
                elif auth.create_user(u, p1, dn, role="owner"):
                    st.session_state.auth_user = auth.verify_login(u, p1)
                    st.rerun()
                else:
                    st.error("Could not create user (name already taken?).")
    else:
        with c.form("login"):
            u = st.text_input("Username")
            p = st.text_input("Password", type="password")
            if st.form_submit_button("Sign in", type="primary"):
                user = auth.verify_login(u, p)
                if user:
                    st.session_state.auth_user = user
                    st.rerun()
                else:
                    st.error("Invalid username or password (or account disabled).")


# ----------------------------- UI -----------------------------
if "db_init_v2" not in st.session_state:
    _init_conn = db.init_db()          # create tables + apply backend migrations
    _init_conn.close()
    auth.ensure_roles_seeded()         # seed default roles/permissions on first run
    st.session_state.db_init_v2 = True

if "auth_user" not in st.session_state:
    _render_login()
    st.stop()

USER = st.session_state.auth_user
db.set_audit_actor(USER)
ROLE = USER.get("Role", "viewer")
PERMS = auth.role_perms(ROLE)
def can(p):
    return p in PERMS

_ensure_state()
# Sidebar shows the standalone LOGO (full width); falls back to the banner if no logo.
if os.path.exists(db.logo_path()):
    st.sidebar.image(db.logo_path(), width="stretch")
elif os.path.exists(_LOGO):
    st.sidebar.image(_LOGO, width="stretch")
st.sidebar.title(_COMPANY)
st.sidebar.caption(f"👤 **{USER.get('DisplayName') or USER.get('Username')}** · _{_role_label(ROLE)}_")
if st.sidebar.button("🔒 Log out", width="stretch"):
    st.session_state.pop("auth_user", None)
    st.rerun()

PROJECT_WORKSPACE_LABEL = "Projects/Offers"
_SECTIONS = [(PROJECT_WORKSPACE_LABEL, "load"), ("Reports", "reports"), ("Audit", "audit"),
             ("Products Catalogue", "catalogue"), ("Settings", "settings"), ("Users", "users")]
_allowed = [name for name, p in _SECTIONS if can(p)]
# New-offer-only roles still enter through the combined project workspace.
if can("new_offer") and PROJECT_WORKSPACE_LABEL not in _allowed:
    _allowed.insert(0, PROJECT_WORKSPACE_LABEL)
if not _allowed:
    st.error("Your account has no accessible sections - contact the owner.")
    st.stop()
_nav_mode = st.session_state.pop("_nav_mode", None)
if _nav_mode in _allowed:
    st.session_state["workspace_mode"] = _nav_mode
if st.session_state.get("workspace_mode") not in _allowed:
    st.session_state["workspace_mode"] = _allowed[0]
mode = st.sidebar.radio("Workspace", _allowed, key="workspace_mode")
if st.session_state.get("_rendered_workspace_mode") != mode:
    _request_scroll_top()
st.session_state["_rendered_workspace_mode"] = mode
_scroll_to_top_if_requested()
admin = can("view_costs")          # on-screen internal cost metrics (client PDF never shows costs)
owner = ROLE == auth.PROTECTED_ROLE

# Refresh per-currency -> USD rates from Settings (EUR configurable; SAR pegged).
try:
    calc.CURRENCY_RATES["EUR"] = float(repo.get_setting("eur_to_usd") or 1.08)
except (TypeError, ValueError):
    pass
calc.CURRENCY_RATES["SAR"] = 1.0 / calc.SAR_PER_USD
calc.CURRENCY_RATES["AED"] = 1.0 / calc.AED_PER_USD
try:
    calc.VAT_RATE = float(repo.get_setting("vat_percent") or 15) / 100
except (TypeError, ValueError):
    pass

# Load Project is the main project workspace. New offers open as a nested view
# from its action button instead of occupying a separate sidebar page.
project_workspace_view = st.session_state.get("project_workspace_view", "load")
if not can("load") and can("new_offer"):
    project_workspace_view = "new"
elif not can("new_offer"):
    project_workspace_view = "load"
st.session_state["project_workspace_view"] = project_workspace_view

# ============================ NEW OFFER ============================
if mode == PROJECT_WORKSPACE_LABEL and project_workspace_view == "new":
    # If a new offer was just saved, show the edit panel directly instead of the form.
    if st.session_state.get("edit_mode") and st.session_state.get("_just_saved_meta") is not None:
        _edit_panel(st.session_state["_just_saved_meta"])
        st.stop()

    duplicate = st.session_state.pop("_duplicate_offer", None)
    if duplicate:
        _prime_new_offer_form(
            duplicate["header"], duplicate["grid"], duplicate["discount"],
            duplicate.get("commission", 0.0), duplicate.get("commission_percent", 0.0),
            duplicate.get("commission_mode", "Deduct from profit"))
        st.success(f"Duplicated from **{duplicate['source']}**. Review and save as a new offer.")
    elif st.session_state.pop("_no_reset_all", False):
        _prime_new_offer_form()
    elif st.session_state.pop("_no_reset_option", False):
        st.session_state["no_option"] = ""
        st.session_state.pop("editor", None)

    nt1, nt2 = st.columns([4, 1])
    nt1.subheader("New Project")
    if can("load") and nt2.button("← Load Project", width="stretch", key="back_to_load_project"):
        st.session_state["project_workspace_view"] = "load"
        _request_scroll_top()
        st.rerun()
    h = st.session_state.header

    # Live offer reference (from the System abbreviation + override below) - shown as a top bar.
    # Once the first option is saved, the offer # is "locked" so further options share it.
    _otype = (st.session_state.get("no_offer_type")
              if "no_offer_type" in st.session_state
              else repo.system_name(h.get("system"))) or ""
    _ov = (st.session_state.get("no_offer_ov") or "").strip()
    h["system"] = _otype
    h["offer_override"] = _ov
    _locked = st.session_state.get("no_offer_lock")
    h["offer"] = _locked or _ov or repo.make_offer_no(_otype)
    _saved = st.session_state.get("no_saved_options", [])
    _extra = (f"<span style='font-size:.8rem;opacity:.85'> &nbsp;·&nbsp; options saved: "
              f"{', '.join(_saved)}</span>") if _saved else ""
    st.markdown(
        f"<div style='background:#002060;color:#fff;padding:10px 16px;border-radius:8px;"
        f"font-size:1.2rem;margin:2px 0 12px'>🧾&nbsp;&nbsp;Offer #:&nbsp; <b>{h['offer']}</b>{_extra}</div>",
        unsafe_allow_html=True)
    _new_offer_actions()

    with st.expander("Project Details", expanded=True):
        c1, c2, c3 = st.columns(3)
        h["client"] = c1.text_input("Client", h["client"], key="no_client")
        h["project"] = c1.text_input("Project", h["project"], key="no_project")
        h["contact"] = c2.text_input("Contact", h["contact"], key="no_contact")
        h["phone"] = c2.text_input(
            "Phone", h["phone"], key="no_phone",
            on_change=_copy_widget_to_state_dict,
            args=("header", "phone", "no_phone"))
        h["contractor"] = c3.text_input(
            "Contractor", h.get("contractor", ""), key="no_contractor",
            on_change=_copy_widget_to_state_dict,
            args=("header", "contractor", "no_contractor"))
        h["region"] = _region_select(c3, h.get("region", ""), "no_region")
        p1, p2, p3 = st.columns(3)
        h["sales"] = _person_select(p1, "Sales Person", SALES_PERSON_ROLES,
                                    h.get("sales", ""), "no_sales")
        h["presales"] = _person_select(p2, "Pre-sales Engineer", PEOPLE_ROLES["presales"],
                                       h.get("presales", ""), "no_presales")
        h["pm"] = _person_select(p3, "Project Manager", PEOPLE_ROLES["pm"],
                                 h.get("pm", ""), "no_pm")
        # System name is stored on the project; its abbreviation drives *TYPE* in Offer #.
        o1, o2 = st.columns(2)
        h["system"] = _system_select(o1, h.get("system", ""), "no_offer_type")
        o2.text_input("Offer # (blank = auto)", key="no_offer_ov",
                      help="Leave blank to auto-number; type a value to override.")

    terms_form(st.session_state.header, "no")
    if _ps_enabled():
        project_sheet_info_form(st.session_state.header, "no_ps")

    _new_project_editor()


# ============================ LOAD EXISTING ============================
elif mode == PROJECT_WORKSPACE_LABEL:
    lt1, lt2 = st.columns([4, 1], vertical_alignment="center")
    lt1.subheader("Load Project")
    if can("new_offer") and lt2.button(
        "➕ New Project / Offer", type="primary", width="stretch", key="open_new_project"
    ):
        _prime_new_offer_form()
        st.session_state["project_workspace_view"] = "new"
        _request_scroll_top()
        st.rerun()
    if st.session_state.pop("_del_reset", False):        # clear delete confirm widgets
        st.session_state.pop("del_confirm", None)
        st.session_state.pop("del_scope", None)

    projects, fams = _cached_project_index(_db_cache_stamp())
    if projects.empty:
        st.info("No projects ingested yet. Run `python ingest.py`.")
    else:
        viewing_fam = st.session_state.get("load_fam")
        if not viewing_fam:
            restore = st.session_state.pop("_load_restore_search", None)
            if restore:
                st.session_state["load_search_name"] = restore.get("name", "")
                st.session_state["load_search_offer"] = restore.get("offer", "")
                st.session_state["load_filter_sales"] = restore.get("sales", [])
                st.session_state["load_filter_presales"] = restore.get("presales", [])
                st.session_state["load_filter_pm"] = restore.get("pm", [])

            sc1, sc2 = st.columns([2, 1])
            raw_name = sc1.text_input("Search by name", key="load_search_name")
            raw_offer = sc2.text_input("Search by offer #", key="load_search_offer")
            pc1, pc2, pc3 = st.columns(3)
            q_sales = pc1.multiselect(
                "Sales Person",
                _project_person_filter_options(
                    projects, "SalesPerson", auth.users_in_roles(SALES_PERSON_ROLES)),
                key="load_filter_sales",
            )
            q_presales = pc2.multiselect(
                "Pre-sales Engineer",
                _project_person_filter_options(
                    projects, "PresalesEngineer", auth.users_in_role(PEOPLE_ROLES["presales"])),
                key="load_filter_presales",
            )
            q_pm = pc3.multiselect(
                "Project Manager",
                _project_person_filter_options(
                    projects, "ProjectManager", auth.users_in_role(PEOPLE_ROLES["pm"])),
                key="load_filter_pm",
            )
            search_snapshot = {
                "name": _text(raw_name),
                "offer": _text(raw_offer),
                "sales": list(q_sales),
                "presales": list(q_presales),
                "pm": list(q_pm),
            }
            st.session_state["load_search_snapshot"] = search_snapshot
            q_name = search_snapshot["name"].lower()
            q_offer = search_snapshot["offer"].lower()

            query_key = repr((q_name, q_offer, tuple(q_sales), tuple(q_presales), tuple(q_pm)))
            if st.session_state.get("load_query") != query_key:
                st.session_state.load_query = query_key
                st.session_state.pop("view_pid", None)
                st.session_state.pop("pdf_bytes", None)
                st.session_state.pop("project_sheet_bytes", None)
                _clear_edit_widget_state()
                st.session_state.edit_mode = False
        else:
            search_snapshot = st.session_state.get("load_search_snapshot") or {
                "name": _text(st.session_state.get("load_search_name")),
                "offer": _text(st.session_state.get("load_search_offer")),
                "sales": list(st.session_state.get("load_filter_sales", [])),
                "presales": list(st.session_state.get("load_filter_presales", [])),
                "pm": list(st.session_state.get("load_filter_pm", [])),
            }
            st.session_state["load_search_snapshot"] = search_snapshot
            q_name = search_snapshot["name"].lower()
            q_offer = search_snapshot["offer"].lower()
            q_sales = search_snapshot["sales"]
            q_presales = search_snapshot["presales"]
            q_pm = search_snapshot["pm"]
            # Keep the hidden widget values alive across project-detail reruns so Back
            # restores the exact search even after edits, exports, or approval actions.
            st.session_state["load_search_name"] = search_snapshot["name"]
            st.session_state["load_search_offer"] = search_snapshot["offer"]
            st.session_state["load_filter_sales"] = list(q_sales)
            st.session_state["load_filter_presales"] = list(q_presales)
            st.session_state["load_filter_pm"] = list(q_pm)

        if not viewing_fam and not (q_name or q_offer or q_sales or q_presales or q_pm):
            st.session_state.pop("view_pid", None)
            st.session_state.pop("load_fam", None)
            st.info(
                "Search by project/client name, offer number, Sales Person, "
                "Pre-sales Engineer, or Project Manager."
            )
            st.stop()

        projects, fams = _cached_project_search(
            q_name, q_offer, tuple(q_sales), tuple(q_presales), tuple(q_pm),
            _db_cache_stamp(),
        )
        matches = [
            f for f in fams
            if (not q_name or q_name in f["name_search"])
            and (not q_offer or q_offer in f["offer_search"])
            and (not q_sales or bool(set(q_sales) & f["sales_people"]))
            and (not q_presales or bool(set(q_presales) & f["presales_people"]))
            and (not q_pm or bool(set(q_pm) & f["project_managers"]))
        ]
        if not matches:
            st.session_state.pop("view_pid", None)
            st.session_state.pop("load_fam", None)
            st.warning("No matching offers found.")
            st.stop()

        current_fam = st.session_state.get("load_fam")
        match_fams = {f["fam"] for f in matches}
        if current_fam not in match_fams:
            current_fam = None
            st.session_state.pop("load_fam", None)
            st.session_state.pop("view_pid", None)

        # While no offer is chosen, show the matching list. Once one is opened,
        # hide the list and show only that offer (with a Back button).
        if not current_fam:
            st.markdown("**Matching offers**")
            widths = [2.1, 1.3, 1.25, 1.7, 1.15, 0.85, 0.5, 0.75, 0.6, 0.7]
            hc = st.columns(widths)
            for col, t in zip(hc, ["Project", "System", "Client", "Offer #", "Sales Person",
                                   "Region", "Rev.", "Updated", "Approved", ""]):
                col.caption(t)
            for idx, f in enumerate(matches):
                rc = st.columns(widths, vertical_alignment="center")
                rc[0].write(_text(f["project_label"], "Offer"))
                rc[1].write(_text(f["system"], "-"))
                rc[2].write(_text(f["client"], "-"))
                rc[3].write(_text(f["offer_label"], "-"))
                rc[4].write(_text(f["sales"], "-"))
                rc[5].write(_text(f["region"], "-"))
                rc[6].write(str(f["n_rev"]))
                rc[7].write(_fmt_date(f["updated_date"]))
                rc[8].write("✅" if f["approved"] else "")
                if rc[9].button("View", key=f"match_view_{idx}_{f['fam']}",
                                width="stretch"):
                    st.session_state.load_fam = f["fam"]
                    st.session_state.pop("view_pid", None)
                    st.session_state.pop("pdf_bytes", None)
                    st.session_state.pop("project_sheet_bytes", None)
                    _clear_edit_widget_state()
                    st.session_state.edit_mode = False
                    _request_scroll_top()
                    st.rerun()
            st.info(f"{len(matches)} matching offer{'s' if len(matches) != 1 else ''} found. "
                    "Click View to open one.")
            st.stop()

        # ---- An offer is open: show a Back button + only this offer ----
        _sel_f = next((f for f in matches if f["fam"] == current_fam), None)
        bcol1, bcol2 = st.columns([1.3, 4], vertical_alignment="center")
        if bcol1.button("← Back to Results", width="stretch"):
            st.session_state["_load_restore_search"] = dict(
                st.session_state.get("load_search_snapshot") or {})
            for k in ("load_fam", "view_pid", "pdf_bytes", "project_sheet_bytes"):
                st.session_state.pop(k, None)
            _clear_edit_widget_state()
            st.session_state.edit_mode = False
            _request_scroll_top()
            st.rerun()
        if _sel_f:
            bcol2.markdown(
                f"**{_text(_sel_f['project_label'], 'Offer')}**"
                + (f" · {_text(_sel_f['client'])}" if _text(_sel_f["client"]) else "")
                + f"  —  {len(matches)} search match{'es' if len(matches) != 1 else ''}")
        selected_fam = current_fam

        grp = (projects[projects["_fam"] == selected_fam]
               .sort_values(["RevisionNo", "OptionLabel"], na_position="first"))

        # Hide archived entries by default; toggle to reveal them.
        n_arch = int(grp["Archived"].fillna(0).sum())
        show_arch = st.checkbox(f"Show archived ({n_arch})", value=False, key="show_archived",
                                disabled=not n_arch)
        shown = grp if show_arch else grp[grp["Archived"].fillna(0) == 0]
        if shown.empty:                                  # all archived -> show them anyway
            shown = grp

        # Default selection: the approved entry, else the newest active, else newest.
        rev_ids = shown["ProjectID"].tolist()
        approved_ids = shown[shown["Approved"].fillna(0) == 1]["ProjectID"].tolist()
        active_ids = shown[shown["Archived"].fillna(0) == 0]["ProjectID"].tolist()
        default_id = int(approved_ids[-1] if approved_ids
                         else (active_ids[-1] if active_ids else rev_ids[-1]))
        if st.session_state.get("view_pid") not in rev_ids:
            st.session_state.view_pid = default_id

        # Revisions grouped first, then the options inside each revision.
        shown_opts = shown.copy()
        shown_opts["_rev_sort"] = shown_opts["RevisionNo"].fillna(0).astype(int)
        pid = int(st.session_state.view_pid)
        family_project_ids = tuple(int(value) for value in shown_opts["ProjectID"].tolist())
        systems, meta, grid, totals_by_pid = _cached_project_bundle(
            pid, family_project_ids, _db_cache_stamp()
        )
        grid = grid.copy()

        st.markdown("**Revisions & options**")
        widths = [0.35, 1.3, 1.65, 0.8, 0.8, 1.2, 0.45, 0.9, 0.8]
        for rn, rev_grp in shown_opts.groupby("_rev_sort", sort=True):
            rev_label = repo.revision_token(rn) if rn > 0 else "Original"
            opt_count = len(rev_grp)
            approved_in_rev = bool(rev_grp["Approved"].fillna(0).max())
            st.markdown(
                f"<div style='font-weight:800;margin:14px 0 4px'>{html.escape(rev_label)}"
                f"<span style='font-size:.85rem;font-weight:600;color:#6b7280'>"
                f" · {opt_count} option{'s' if opt_count != 1 else ''}"
                f"{' · approved' if approved_in_rev else ''}</span></div>",
                unsafe_allow_html=True,
            )
            hc = st.columns(widths)
            for col, t in zip(hc, ["", "Option", "Offer #", "Created", "Updated",
                                   "Grand Total (SAR)", "✓", "Status", ""]):
                _ctr(col, t, header=True)
            for _, row in rev_grp.iterrows():
                rid = int(row["ProjectID"])
                sel = (rid == int(st.session_state.view_pid))
                rc = st.columns(widths, vertical_alignment="center")
                _ctr(rc[0], "▶" if sel else "")
                _ctr(rc[1], _text(row["OptionLabel"], "Main"))
                _ctr(rc[2], _text(row["OfferNo"]))
                _ctr(rc[3], _fmt_date(row["CreationDate"]))
                _ctr(rc[4], _fmt_date(row.get("UpdatedDate") or row["CreationDate"]))
                _ctr(rc[5], f"{totals_by_pid.get(rid, 0):,.2f}")
                _ctr(rc[6], "✅" if row["Approved"] else "")
                _ctr(rc[7], "📦 Archived" if row["Archived"] else "Active")
                if rc[8].button("View", key=f"view_{rid}", disabled=sel, width="stretch"):
                    st.session_state.view_pid = rid
                    st.session_state.pop("pdf_bytes", None)
                    st.session_state.pop("project_sheet_bytes", None)
                    _clear_edit_widget_state()
                    st.session_state.edit_mode = False
                    _request_scroll_top()
                    st.rerun()

        sheet = systems[0] if systems else None      # auto-pick the system sheet
        cur_key = f"{pid}::{sheet}"
        editing = (st.session_state.get("edit_mode")
                   and st.session_state.get("edit_key") == cur_key)

        if not editing:
            # -------------------- VIEW: tabbed offer view --------------------
            disp = grid.copy()
            for col in MONEY_COLS:
                if col in disp.columns:
                    disp[col] = disp[col].map(lambda v: calc.roundup(v, 0))
            if meta.get("InclusionMode") == "included":
                disp = calc.apply_inclusion(disp)
            s = calc.summarize(
                disp, meta.get("DiscountAmount") or 0, meta.get("CommissionAmount") or 0)

            _subj = repo.load_terms(meta).get("subject")
            if _subj:
                st.markdown(f"#### 📄 {_subj}")

            view_actions = st.container()
            _project_details_readonly(
                meta, repo.base_name(sheet or "").replace("BOQ", "").strip())
            active_tab = _offer_tab_selector(pid, bool(meta.get("Approved")))
            if active_tab == "BoQ":
                _summary_metrics(s)

                # Approval + archive controls (BoQ tab only).
                apc1, apc2, apc3 = st.columns([2.2, 1, 1], vertical_alignment="center")
                if meta.get("Archived"):
                    apc1.warning("📦 **Archived**" + (" · was ✅ approved" if meta.get("Approved") else ""))
                elif meta.get("Approved"):
                    at = (meta.get("ApprovedAt") or "")[:16].replace("T", " ")
                    apc1.success(f"✅ **Approved**{(' · ' + at) if at else ''}")
                else:
                    apc1.info("Active · not approved")
                if meta.get("Approved"):
                    if can("approve") and apc2.button("↩️ Unapprove", width="stretch"):
                        r = repo.unapprove_offer(pid)
                        st.cache_data.clear()
                        st.toast(f"Unapproved. {r} auto-archived entr{'y' if r == 1 else 'ies'} restored."
                                 if r else "Unapproved.", icon="↩️")
                        st.rerun()
                elif can("approve"):
                    if apc2.button("✅ Approve", type="primary", width="stretch"):
                        offer_label = (_text(meta.get("OfferNo"))
                                       or _text(meta.get("ProjectName"), f"Offer #{pid}"))
                        option_label = _text(meta.get("OptionLabel"))
                        if option_label:
                            offer_label = f"{offer_label} · {option_label}"
                        _approve_offer_dialog(pid, offer_label)
                if can("archive"):
                    if meta.get("Archived"):
                        if apc3.button("♻️ Restore", width="stretch"):
                            repo.unarchive_project(pid)
                            st.rerun()
                    elif apc3.button("📦 Archive", width="stretch"):
                        repo.archive_project(pid)
                        st.rerun()
                if admin:                   # gross-profit line (internal cost view) - BoQ tab only
                    _profit_banner(s)
                cfg = {c: st.column_config.NumberColumn(c, format="accounting") for c in MONEY_COLS}
                cfg["Qty"] = st.column_config.NumberColumn("Qty", format="%d")
                cfg["Shipping %"] = st.column_config.NumberColumn("Shipping %", format="%.2f")
                view_grid = disp[[c for c in BUILDER_COLS if c in disp.columns]].copy()
                if meta.get("InclusionMode") == "included" and "_IncludedInItems" in disp.columns:
                    view_grid.insert(0, "Include", disp["_IncludedInItems"].fillna(False).astype(bool))
                    cfg["Include"] = st.column_config.CheckboxColumn("Include", disabled=True)
                st.dataframe(view_grid, width="stretch", hide_index=True, column_config=cfg,
                             height=_editor_full_height(len(view_grid)), row_height=35)
            elif active_tab == "Tracking":
                _render_tracking_tab(pid, sheet)
            else:
                _render_finance_tab(pid, s["grand_total_sar"])

            with view_actions:
                b1, b2, b3, b4, b5, b6 = st.columns(
                    [1.3, 0.9, 1.05, 0.8, 1.0, 0.95]
                )
            if can("edit") and b1.button("✏️ Edit / new revision or option", type="primary",
                                          width="stretch"):
                _init_edit_state(pid, meta, grid, sheet)
                _request_scroll_top()
                st.rerun()
            if can("new_offer") and b2.button("📋 Duplicate", width="stretch"):
                dg = grid.copy()
                if dg.empty:
                    st.warning("This offer has no lines to duplicate.")
                else:
                    system_suffix = repo.system_name(
                        repo.base_name(sheet or "").replace("BOQ", "").strip()
                        or _default_system())
                    copied_terms = repo.load_terms(meta)
                    copied_header = {
                        **copied_terms,
                        "client": _text(meta.get("ClientName")),
                        "project": _text(meta.get("ProjectName")),
                        "contact": _text(meta.get("ContactName")),
                        "phone": _text(meta.get("ContactPhone")),
                        "contractor": _text(meta.get("Contractor")),
                        "region": _text(meta.get("Region")),
                        "sales": _text(meta.get("SalesPerson")),
                        "presales": _text(meta.get("PresalesEngineer")),
                        "pm": _text(meta.get("ProjectManager")),
                        "system": system_suffix,
                        "project_sheet": repo.load_project_sheet_info(meta),
                    }
                    st.session_state["_duplicate_offer"] = {
                        "source": _text(meta.get("OfferNo")) or _text(meta.get("ProjectName"), "existing offer"),
                        "header": copied_header,
                        "grid": calc.recompute(dg),
                        "discount": abs(float(meta.get("DiscountAmount") or 0)),
                        "commission": abs(float(meta.get("CommissionAmount") or 0)),
                        "commission_percent": abs(float(meta.get("CommissionPercent") or 0)),
                        "commission_mode": meta.get("CommissionMode") or "Deduct from profit",
                    }
                    st.session_state["_nav_mode"] = PROJECT_WORKSPACE_LABEL
                    st.session_state["project_workspace_view"] = "new"
                    _clear_edit_widget_state()
                    st.session_state.edit_mode = False
                    _request_scroll_top()
                    st.rerun()
            _pdf_rev = int(meta.get("RevisionNo") or 0)
            _rev_rows = shown_opts[shown_opts["RevisionNo"].fillna(0).astype(int) == _pdf_rev]
            _active_rev_rows = _rev_rows[_rev_rows["Archived"].fillna(0) == 0]
            _pdf_opt_count = len(_active_rev_rows if not _active_rev_rows.empty else _rev_rows)

            def _export_header():
                return {**DEFAULT_TERMS, **repo.load_terms(meta),
                        "client": meta.get("ClientName"), "project": meta.get("ProjectName"),
                        "contact": meta.get("ContactName"), "phone": meta.get("ContactPhone") or "",
                        "contractor": meta.get("Contractor") or "",
                        "region": meta.get("Region") or "",
                        "sales": meta.get("SalesPerson"), "presales": meta.get("PresalesEngineer"),
                        "pm": meta.get("ProjectManager"),
                        "system": repo.system_name(
                            repo.base_name(sheet or "").replace("BOQ", "").strip()),
                        "offer": meta.get("OfferNo"), "date": meta.get("UpdatedDate") or meta.get("CreationDate"),
                        "project_sheet": repo.load_project_sheet_info(meta)}

            if b3.button(f"📄 Generate Offer PDF{f' ({_pdf_opt_count} options)' if _pdf_opt_count > 1 else ''}",
                         width="stretch"):
                # A PDF generated from the loaded-offer view supersedes any
                # saved-edit export context left in this browser session.
                st.session_state.pop("saved_rev", None)
                st.session_state.pop("saved_export_header", None)
                st.session_state.pop("saved_export_grid", None)
                st.session_state.pop("saved_export_summary", None)
                opts = revision_options(pid)
                _make_pdf_download(_export_header(), disp, s, options=opts)
            _pdf_name = f"Quotation_{meta.get('OfferNo') or meta.get('ProjectName')}.pdf"
            if st.session_state.get("pdf_bytes"):
                b4.download_button(
                    "⬇️ Download PDF", st.session_state.pdf_bytes,
                    file_name=_pdf_name, mime="application/pdf", width="stretch")
            else:
                b4.button("⬇️ Download PDF", disabled=True, width="stretch")
            if _ps_enabled() and b5.button("📊 Generate Project Sheet", width="stretch"):
                st.session_state.pop("saved_rev", None)
                _make_project_sheet_download(_export_header(), s)
            if _ps_enabled():
                if st.session_state.get("project_sheet_bytes"):
                    b6.download_button(
                        "⬇️ Download Project Sheet",
                        st.session_state.project_sheet_bytes,
                        file_name=(
                            f"Project_Sheet_{_safe_filename(meta.get('OfferNo') or meta.get('ProjectName'))}.xlsx"
                        ),
                        mime=(
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet"
                        ),
                        width="stretch",
                    )
                else:
                    b6.button("⬇️ Download Project Sheet", disabled=True, width="stretch")

            if can("delete"):
              with st.expander("Delete..."):
                _rn = int(meta.get("RevisionNo") or 0)
                _rlbl = repo.revision_token(_rn) if _rn > 0 else "Original"
                _opt = meta.get("OptionLabel") or "-"
                scopes = {
                    f"This option only  ({_rlbl} · option {_opt})": "option",
                    f"This revision  ({_rlbl} and all its options)": "revision",
                    "This entire offer  (all revisions & options)": "offer",
                }
                pick = st.radio("What to delete", list(scopes), key="del_scope")
                ids = repo.deletion_ids(pid, scopes[pick])
                st.warning(f"Permanently deletes **{len(ids)}** entr"
                           f"{'y' if len(ids) == 1 else 'ies'} (and their line items). "
                           "This cannot be undone.")
                ok = st.checkbox("Yes, permanently delete", key="del_confirm")
                if st.button("Delete now", type="primary", disabled=not ok):
                    n = repo.delete_projects(ids)
                    st.session_state.pop("view_pid", None)
                    st.session_state["_del_reset"] = True
                    st.success(f"Deleted {n} entr{'y' if n == 1 else 'ies'}.")
                    st.rerun()
        else:
            _edit_panel(meta)


# ============================ REPORTS ============================
elif mode == "Reports":
    st.subheader("Reports & statistics")
    _rep_company = _company_dict()
    tab_builder, tab_dash = st.tabs(
        ["🧱 Report Builder", "📊 Dashboard"],
        key="reports_active_tab",
        on_change="rerun",
    )
    with tab_builder:
        _render_report_builder(_rep_company)
    with tab_dash:
        _render_dashboard(_rep_company)


# ============================ AUDIT ============================
elif mode == "Audit":
    _render_audit_page()


# ============================ CATALOGUE ============================
elif mode == "Products Catalogue":
    st.subheader("Products Catalogue")
    _cat_edit = can("catalogue_edit")

    # ---- Add a new item ----
    if _cat_edit:
      with st.expander("➕ Add new item"):
        with st.form("add_cat_item", clear_on_submit=True):
            ac1, ac2, ac3 = st.columns(3)
            a_brand = ac1.text_input("Brand")
            a_model = ac2.text_input("Model")
            a_desc = ac3.text_input("Description")
            fc, f1, f2, f3, f4, f5, f6 = st.columns([1, 1.2, 1.4, 1.2, 1.4, 1.5, 1.6])
            a_cur = fc.selectbox("Currency", calc.CURRENCIES, index=0,
                                 help="Currency of List Price & Ex Unit Cost. "
                                      "Unit Cost is stored in USD.")
            a_list = f1.number_input("List Price", min_value=0.0, value=0.0, step=1.0)
            a_ex = f2.number_input("Ex Unit Cost", min_value=0.0, value=0.0, step=1.0)
            a_ship = f3.number_input("Shipping %", min_value=0.0, value=30.0, step=5.0)
            a_unit = f4.number_input("Unit Cost (USD)", min_value=0.0, value=0.0, step=1.0,
                                     help="If Ex Unit Cost is entered, this is calculated (in USD) from Shipping %.")
            a_up = f5.number_input("Default U.Price $", min_value=0.0, value=0.0, step=1.0)
            a_ups = f6.number_input("Default U.Price SAR", min_value=0.0, value=0.0, step=10.0)
            if st.form_submit_button("➕ Add item", type="primary"):
                if not (a_model.strip() or a_desc.strip()):
                    st.warning("Enter at least a Model or a Description.")
                else:
                    iid = repo.add_catalog_item(
                        a_brand.strip(), a_model.strip(), a_desc.strip(),
                        a_list or None, a_ex or None, a_ship, a_unit or None, a_up or None,
                        a_ups or None, currency=a_cur)
                    if iid:
                        st.success(f"Added '{a_model or a_desc}' (ItemID {iid}).")
                        st.rerun()
                    else:
                        st.warning("An item with the same Brand + Model + Description already exists.")

    st.caption("Prices shown **rounded up**. Edit Brand, Model, Description, cost, or "
               "default-price cells inline, or tick **Del** to remove items. Use "
               "**Add new item** above to create another product.")
    term = st.text_input("Search", placeholder="Model / Description / Brand")
    term = term.strip()
    if not term:
        st.info("Search by model, description, or brand.")
        res = pd.DataFrame()
    else:
        _cat_limit = 1000
        res = repo.search_catalog(term, limit=_cat_limit, show_discontinued=True).reset_index(drop=True)
        st.caption(
            f"{len(res)} item(s)"
            + (f" shown - refine the search for more specific results" if len(res) >= _cat_limit else "")
        )
    _cat_select_reset = st.session_state.pop("_cat_clear_select_all", None)
    if _cat_select_reset:
        st.session_state[_cat_select_reset] = False
    _cat_save_result = st.session_state.pop("_cat_save_result", None)
    if _cat_save_result is not None:
        _cat_saved, _cat_conflicts = _cat_save_result
        st.success(f"Updated {_cat_saved} catalogue item(s).")
        if _cat_conflicts:
            st.warning(
                f"Skipped {_cat_conflicts} item(s) because the edited description would "
                "duplicate an existing Brand + Model + Description."
            )
    if not res.empty:
        rename = {"ListPriceUSD": "List Price $", "ExUnitCostUSD": "Ex Unit Cost $",
                  "Currency": "Cur",
                  "ShippingPercent": "Shipping %", "UnitCostUSD": "Unit Cost $", "DefaultUPriceUSD": "Default U.Price $",
                  "DefaultUPriceSAR": "Default U.Price SAR", "PriceUpdatedAt": "Price Updated",
                  "TimesQuoted": "Times Quoted"}
        edit_cols = list(repo.CATALOG_EDITABLE.keys())
        money_cols = [c for c in edit_cols if c != "Shipping %"]
        disp = res.rename(columns=rename).copy()
        if "Cur" not in disp.columns:
            disp["Cur"] = "USD"
        disp["Cur"] = disp["Cur"].fillna("USD")
        for c in money_cols:                       # always display rounded up
            disp[c] = disp[c].map(lambda v: calc.roundup(v, 0))
        if "Shipping %" in disp.columns:
            disp["Shipping %"] = disp["Shipping %"].map(lambda v: calc.shipping_percent(v))
        if "Price Updated" not in disp.columns:
            disp["Price Updated"] = "01-2025"
        else:
            disp["Price Updated"] = disp["Price Updated"].map(_fmt_month_year)
        base_cols = ["Brand", "Model", "Description", "Cur"] + edit_cols + ["Price Updated", "Times Quoted", "Discontinued"]
        colcfg = {c: st.column_config.NumberColumn(c, format="accounting", min_value=0.0)
                  for c in money_cols}
        colcfg["Cur"] = st.column_config.SelectboxColumn(
            "Cur", options=calc.CURRENCIES, required=False,
            help="Currency of List Price & Ex Unit Cost. Unit Cost (USD) recomputes when changed.")
        colcfg["List Price $"] = st.column_config.NumberColumn("List Price", format="accounting", min_value=0.0)
        colcfg["Ex Unit Cost $"] = st.column_config.NumberColumn("Ex Unit Cost", format="accounting", min_value=0.0)
        colcfg["Unit Cost $"] = st.column_config.NumberColumn("Unit Cost (USD)", format="accounting", min_value=0.0)
        colcfg["Shipping %"] = st.column_config.NumberColumn("Shipping %", format="%.2f", min_value=0.0, step=5.0)
        colcfg["Price Updated"] = st.column_config.TextColumn(
            "Updated On", width="small", alignment="center")
        colcfg["Times Quoted"] = st.column_config.NumberColumn("Times Quoted", format="%d")
        colcfg["Brand"] = st.column_config.TextColumn("Brand", width="medium")
        colcfg["Model"] = st.column_config.TextColumn("Model", width="medium")
        colcfg["Description"] = st.column_config.TextColumn("Description", width="large")
        colcfg["Discontinued"] = st.column_config.CheckboxColumn(
            "Discontinued", help="Discontinued items stay visible here but won't appear in offer searches")
        if "Discontinued" not in disp.columns:
            disp["Discontinued"] = False
        else:
            disp["Discontinued"] = disp["Discontinued"].fillna(0).astype(bool)
        if not _cat_edit:                              # read-only catalogue
            st.dataframe(disp[base_cols], width="stretch", hide_index=True,
                         column_config=colcfg)
        else:
            select_all_key = f"cat_select_all::{term}"
            select_all = st.checkbox(
                "Select all displayed products",
                key=select_all_key,
                help="Selects every product shown by the current search for deletion.",
            )
            disp["Del"] = bool(select_all)
            colcfg["Del"] = st.column_config.CheckboxColumn("Del", help="Tick to delete this item")
            editor_generation = int(st.session_state.get("_cat_editor_generation", 0))
            edited = st.data_editor(
                disp[["Del"] + base_cols], column_config=colcfg, num_rows="fixed", hide_index=True,
                width="stretch",
                key=f"cat_editor::{term}::{int(select_all)}::{editor_generation}",
                disabled=["Price Updated", "Times Quoted"])
            del_ids = [int(res.iloc[i]["ItemID"]) for i in range(len(edited))
                       if bool(edited.iloc[i]["Del"])]
            b1, b2 = st.columns(2)
            if b1.button("💾 Save catalogue changes", type="primary", width="stretch"):
                n = 0
                duplicate_conflicts = 0
                for i in range(len(edited)):
                    changes = {}
                    for text_field in ("Brand", "Model", "Description"):
                        new_text = "" if pd.isna(edited.iloc[i][text_field]) else str(
                            edited.iloc[i][text_field]).strip()
                        old_text = "" if pd.isna(disp.iloc[i][text_field]) else str(
                            disp.iloc[i][text_field]).strip()
                        if new_text != old_text:
                            changes[text_field] = new_text
                    for c in edit_cols:
                        new = edited.iloc[i][c]
                        if pd.isna(new):
                            continue
                        if abs(float(new) - float(disp.iloc[i][c])) > 1e-9:
                            changes[repo.CATALOG_EDITABLE[c]] = float(new)
                    new_cur = str(edited.iloc[i]["Cur"])
                    if new_cur in calc.CURRENCIES and new_cur != str(disp.iloc[i]["Cur"]):
                        changes["Currency"] = new_cur
                    new_disc = bool(edited.iloc[i].get("Discontinued", False))
                    old_disc = bool(disp.iloc[i].get("Discontinued", False))
                    if new_disc != old_disc:
                        repo.catalog_set_discontinued(int(res.iloc[i]["ItemID"]), new_disc)
                    if changes:
                        if repo.update_catalog_item(int(res.iloc[i]["ItemID"]), changes):
                            n += 1
                        else:
                            duplicate_conflicts += 1
                st.session_state["_cat_save_result"] = (n, duplicate_conflicts)
                st.rerun()
            if b2.button(f"Delete {len(del_ids)} checked item(s)", width="stretch",
                         disabled=not del_ids):
                n = repo.delete_catalog_items(del_ids)
                st.session_state["_cat_clear_select_all"] = select_all_key
                st.session_state["_cat_editor_generation"] = editor_generation + 1
                st.success(f"Deleted {n} item(s).")
                st.rerun()

# ============================ SETTINGS ============================
elif mode == "Settings":
    st.subheader("Settings")
    (
        tab_company,
        tab_offer,
        tab_images_pdf,
        tab_data,
        tab_backup,
        tab_updates,
    ) = st.tabs([
        "Company Details",
        "Offer & Pricing",
        "Images & PDF",
        "Data Tools",
        "Backup & Restore",
        "Updates",
    ], key="settings_active_tab", on_change="rerun")

    with tab_offer:
        st.markdown("##### Offer numbers")
        st.caption("Offer reference numbers are built from a **template** with variables.")

        with st.form("settings_offer_form"):
            template = st.text_input(
                "Offer # template",
                repo.get_setting("offer_template"),
                help=(
                    "Variables: *TYPE* = the selected System abbreviation, *YY* = 2-digit year, "
                    "*YYYY* = 4-digit year, and a run of x's = the auto-number "
                    "(its length is the zero-padding). e.g. LG-*TYPE*-*YY*/xxxx -> "
                    "LG-AV-26/0053. Omit *TYPE* for a fixed prefix (e.g. SWS-*YY*-xxxx)."
                ),
            )
            c1, c2 = st.columns(2)
            pad = c1.number_input(
                "Fallback digits (no x-run)",
                min_value=1,
                max_value=8,
                value=int(repo.get_setting("offer_number_pad") or 3),
                help="Padding used only when the template has no x's.",
            )
            dmargin = c2.number_input(
                "Default margin x",
                min_value=0.0,
                step=0.0001,
                format="%.5g",
                value=float(repo.get_setting("default_margin") or 1.6),
                help=(
                    "Applied to new blank rows and catalogue items with no historical price, "
                    "in New Project and Edit."
                ),
            )

            st.markdown("**Revision label**")
            rcol1, rcol2 = st.columns(2)
            rev_fmt = rcol1.text_input(
                "Revision format",
                repo.get_setting("revision_format"),
                help=(
                    "A run of x's = the revision number (length = padding). "
                    "e.g. Rev.x -> Rev.1 / Rev.10 ; Rxx -> R01 / R10."
                ),
            )
            sep_opts = {
                "Dash (...0053-Rev.1)": "-",
                "Space (...0053 Rev.1)": " ",
                "Underscore (...0053_Rev.1)": "_",
            }
            _cur_sep = repo.get_setting("revision_separator")
            _cur_lbl = next((k for k, v in sep_opts.items() if v == _cur_sep), list(sep_opts)[0])
            rev_sep_lbl = rcol2.selectbox(
                "Separator (offer # -> revision)",
                list(sep_opts.keys()),
                index=list(sep_opts.keys()).index(_cur_lbl),
            )

            st.markdown("**Tax**")
            vcol1, vcol2 = st.columns([1, 2])
            vat_pct = vcol1.number_input(
                "VAT %",
                min_value=0.0,
                max_value=100.0,
                step=0.5,
                value=float(repo.get_setting("vat_percent") or 15),
                help=(
                    "VAT rate applied across offers, quotations and the Finance tab. "
                    "KSA = 15%; change it for other countries."
                ),
            )
            vcol2.caption(
                "Applies everywhere VAT is shown - new offers, loaded offers, the client PDF "
                "and the Finance tab. Changing it re-computes VAT on all offers."
            )

            st.markdown("**Currencies / exchange rates**")
            ecol1, ecol2 = st.columns([1, 2])
            eur_rate = ecol1.number_input(
                "1 EUR = ? USD",
                min_value=0.0,
                step=0.01,
                format="%.4f",
                value=float(repo.get_setting("eur_to_usd") or 1.08),
                help="Converts EUR buy prices to USD when computing the Unit Cost.",
            )
            ecol2.caption(
                f"Pegged (fixed, not editable): 1 USD = {calc.SAR_PER_USD:g} SAR "
                f"(1 SAR ~= {1 / calc.SAR_PER_USD:.4f} USD)  |  "
                f"1 USD = {calc.AED_PER_USD:g} AED (1 AED ~= {1 / calc.AED_PER_USD:.4f} USD)."
            )

            saved_offer = st.form_submit_button("Save offer and pricing settings", type="primary")
            if saved_offer:
                repo.set_setting("offer_template", template.strip())
                repo.set_setting("offer_number_pad", int(pad))
                repo.set_setting("default_margin", float(dmargin))
                repo.set_setting("revision_format", rev_fmt.strip() or "Rev.x")
                repo.set_setting("revision_separator", sep_opts[rev_sep_lbl])
                repo.set_setting("eur_to_usd", float(eur_rate))
                repo.set_setting("vat_percent", float(vat_pct))
                st.session_state.pop("cached_default_margin", None)
                st.success("Offer and pricing settings saved.")

        st.divider()
        st.markdown("##### Offer-number preview")
        st.caption(
            "Numbering is **per series** - each rendered template (type + year) keeps its own "
            "counter, so LG-AV-26/... and LG-LC-26/... don't conflict."
        )
        ex_system = (repo.system_names() or [""])[0]
        ex_code = repo.system_abbreviation(ex_system)
        _ex = repo.make_offer_no(ex_system)
        st.write("Next offer # examples:")
        st.code(
            f"no system    :  {repo.make_offer_no('')}\n"
            f"system {ex_system} ({ex_code})  :  {_ex}\n"
            f"revision     :  {_ex}{repo.revision_separator()}{repo.revision_token(1)}"
            f"   /   {_ex}{repo.revision_separator()}{repo.revision_token(2)}"
        )

        st.divider()
        st.markdown("##### Reset / force a starting number")
        st.caption(
            "Force a series to restart from a chosen number. Higher historical/imported "
            "numbers will not override the restart; any exact numbers already used are "
            "skipped to prevent duplicate Offer references."
        )
        rc1, rc2, rc3 = st.columns([2, 1, 1])
        _reset_systems = repo.system_names()
        _legacy_reset_system = _text(st.session_state.get("reset_series"))
        if (_legacy_reset_system and _legacy_reset_system != "(none)"
                and _legacy_reset_system not in _reset_systems):
            _reset_systems.append(_legacy_reset_system)
        rsel = rc1.selectbox(
            "Series (System)", ["(none)"] + _reset_systems, key="reset_series")
        r_otype = "" if rsel == "(none)" else rsel
        r_next = repo.next_offer_number(r_otype)
        r_floor = repo.get_series_start(repo.series_key(r_otype))
        rc1.caption(
            f"Next: `{repo.make_offer_no(r_otype)}`"
            + (f" | forced start: {r_floor}" if r_floor else "")
        )
        start_at = rc2.number_input(
            "Start at",
            min_value=1,
            value=max(int(r_next), 1),
            step=1,
            key="reset_start_val",
        )
        if rc3.button("Apply", key="reset_series_apply", width="stretch"):
            repo.set_series_start(r_otype, int(start_at))
            st.success(
                f"This series will use the first available number from {int(start_at)} onward.")
            st.rerun()
        if r_floor and rc3.button("Clear", key="reset_series_clear", width="stretch"):
            repo.clear_series_start(r_otype)
            st.success("Forced start cleared - back to automatic numbering.")
            st.rerun()

    with tab_company:
        st.markdown("##### Company / branding")
        with st.form("settings_company_form"):
            gc1, gc2 = st.columns([3, 1])
            company_name = gc1.text_input(
                "Company name",
                repo.get_setting("company_name") or "",
                help="Shown in the page title, client PDF and project sheet.",
            )
            brand_color = gc2.color_picker(
                "Brand color",
                repo.get_setting("company_brand_color") or "#002060",
                help="Primary colour for PDF titles, table headers and footer.",
            )
            company_tagline = st.text_input("Tagline", repo.get_setting("company_tagline") or "")
            company_contact = st.text_input(
                "Contact line (city / country)",
                repo.get_setting("company_contact") or "",
            )
            id1, id2 = st.columns(2)
            company_vat_number = id1.text_input(
                "VAT Number",
                repo.get_setting("company_vat_number") or "",
            )
            company_cr_number = id2.text_input(
                "C.R. Number",
                repo.get_setting("company_cr_number") or "",
            )
            st.divider()
            ps_enabled = st.checkbox(
                "Enable Project Sheet (information section + export)",
                value=repo.get_setting("project_sheet_enabled") != "0",
                help="When off, the Project Sheet info form and its Excel export are hidden "
                     "everywhere (New Project, Edit and the offer view).")
            inclusion_enabled = st.checkbox(
                "Enable Installation Included pricing mode",
                value=repo.get_setting("installation_inclusion_enabled") == "1",
                help="When on, a dropdown appears above the BoQ table in Edit mode letting you "
                     "distribute installation/accessory costs into item prices.")
            saved_company = st.form_submit_button("Save company settings", type="primary")
            if saved_company:
                repo.set_setting("company_name", company_name.strip() or "Company Name")
                repo.set_setting("company_tagline", company_tagline.strip())
                repo.set_setting("company_contact", company_contact.strip())
                repo.set_setting("company_vat_number", company_vat_number.strip())
                repo.set_setting("company_cr_number", company_cr_number.strip())
                repo.set_setting("company_brand_color", brand_color)
                repo.set_setting("project_sheet_enabled", "1" if ps_enabled else "0")
                repo.set_setting("installation_inclusion_enabled", "1" if inclusion_enabled else "0")
                st.session_state.pop("cached_ps_enabled", None)
                st.session_state.pop("cached_inclusion_enabled", None)
                st.success("Company settings saved. (Page title updates on next reload.)")

        st.divider()
        st.markdown("##### Project systems")
        st.caption(
            "The full **System** name appears in Project Details and the BOQ. Its "
            "**Abbreviation** replaces `*TYPE*` in the Offer # template. Add, edit, delete, "
            "or reorder rows here."
        )
        if "*type*" not in (repo.get_setting("offer_template") or "").lower():
            st.warning(
                "Your current Offer # template does not contain `*TYPE*`, so System "
                "abbreviations will not appear yet. Add `*TYPE*` under Settings → "
                "Offer & Pricing → Offer # template (for example `LG-*TYPE*-*YY*-xxx`)."
            )
        _systems_saved = st.session_state.pop("_systems_saved", None)
        if _systems_saved is not None:
            st.success(f"Saved {_systems_saved} system(s).")
        _systems_generation = int(st.session_state.get("_systems_editor_generation", 0))
        with st.form("settings_systems_form"):
            systems_df = pd.DataFrame(repo.systems(), columns=["name", "abbreviation"]).rename(
                columns={"name": "System", "abbreviation": "Abbreviation"})
            edited_systems = st.data_editor(
                systems_df,
                num_rows="dynamic",
                hide_index=True,
                width="stretch",
                key=f"settings_systems_editor::{_systems_generation}",
                column_config={
                    "System": st.column_config.TextColumn(
                        "System", help="Full name shown in Project Details, e.g. Lighting Control."),
                    "Abbreviation": st.column_config.TextColumn(
                        "Abbreviation", help="Inserted at *TYPE* in Offer #, e.g. LC or LCS."),
                },
            )
            save_systems = st.form_submit_button("Save systems", type="primary")
            if save_systems:
                system_rows, partial = [], False
                for _, row in edited_systems.iterrows():
                    name = _text(row.get("System")).strip()
                    abbreviation = _text(row.get("Abbreviation")).strip()
                    if not name and not abbreviation:
                        continue
                    if not name or not abbreviation:
                        partial = True
                        continue
                    system_rows.append({"name": name, "abbreviation": abbreviation})
                name_keys = [row["name"].casefold() for row in system_rows]
                code_keys = [row["abbreviation"].casefold() for row in system_rows]
                if partial:
                    st.warning("Every system row needs both a System name and an Abbreviation.")
                elif len(name_keys) != len(set(name_keys)):
                    st.warning("System names must be unique.")
                elif len(code_keys) != len(set(code_keys)):
                    st.warning("System abbreviations must be unique.")
                else:
                    repo.set_systems(system_rows)
                    st.session_state["_systems_saved"] = len(system_rows)
                    st.session_state["_systems_editor_generation"] = _systems_generation + 1
                    st.rerun()

        st.divider()
        st.markdown("##### Project regions")
        st.caption(
            "These values appear in the Region dropdown for New Project and Edit Project. "
            "Enter one region per line; reorder the lines to reorder the dropdown. Removing "
            "a region does not change it on older saved projects."
        )
        _regions_saved = st.session_state.pop("_regions_saved", None)
        if _regions_saved is not None:
            st.success(f"Saved {_regions_saved} region(s).")
        with st.form("settings_regions_form"):
            region_lines = st.text_area(
                "Regions (one per line)",
                value="\n".join(repo.regions()),
                height=180,
                placeholder="Riyadh\nJeddah\nDammam",
            )
            if st.form_submit_button("Save regions", type="primary"):
                repo.set_regions(region_lines.splitlines())
                st.session_state["_regions_saved"] = len(repo.regions())
                st.rerun()

    with tab_images_pdf:
        st.markdown("##### PDF body template")
        with st.form("settings_pdf_form"):
            pdf_body_options = {
                "Template 1 - current ProQuote layout": "template1",
                "Template 2 - proposal page + BOQ table": "template2",
                "Template 3 - compact billed-to offer + integrated totals": "template3",
            }
            current_pdf_body_template = repo.get_setting("pdf_body_template") or "template1"
            current_pdf_body_label = next(
                (label for label, value in pdf_body_options.items() if value == current_pdf_body_template),
                list(pdf_body_options)[0],
            )
            pdf_body_template_label = st.selectbox(
                "Client PDF body",
                list(pdf_body_options.keys()),
                index=list(pdf_body_options.keys()).index(current_pdf_body_label),
                help=(
                    "Template 1 keeps the current PDF body. Template 2 follows the attached "
                    "proposal/BOQ body style."
                ),
            )

            st.markdown("**PDF header text**")
            st.caption(
                "Used when no full-width banner is uploaded. Optional placeholders: "
                "`{company}`, `{project}`, `{offer}`, `{page}`, `{vat_number}` / `{vat}`, "
                "`{cr_number}` / `{cr}`."
            )
            htxt1, htxt2, htxt3 = st.columns(3)
            header_left_text = htxt1.text_area(
                "Left header",
                repo.get_setting("header_left_text") or "",
                height=90,
            )
            header_middle_text = htxt2.text_area(
                "Middle header",
                repo.get_setting("header_middle_text") or "",
                height=90,
            )
            header_right_text = htxt3.text_area(
                "Right header",
                repo.get_setting("header_right_text") or "",
                height=90,
            )

            st.markdown("**PDF footer text**")
            st.caption(
                "Optional placeholders: `{company}`, `{project}`, `{offer}`, `{page}`, "
                "`{vat_number}` / `{vat}`, `{cr_number}` / `{cr}`."
            )
            ftxt1, ftxt2, ftxt3 = st.columns(3)
            footer_left_text = ftxt1.text_area(
                "Left footer",
                repo.get_setting("footer_left_text") or "",
                height=90,
            )
            footer_middle_text = ftxt2.text_area(
                "Middle footer",
                repo.get_setting("footer_middle_text") or "",
                height=90,
            )
            footer_right_text = ftxt3.text_area(
                "Right footer",
                repo.get_setting("footer_right_text") or "",
                height=90,
            )

            saved_pdf = st.form_submit_button("Save PDF text settings", type="primary")
            if saved_pdf:
                repo.set_setting("pdf_body_template", pdf_body_options[pdf_body_template_label])
                repo.set_setting("header_left_text", header_left_text.strip())
                repo.set_setting("header_middle_text", header_middle_text.strip())
                repo.set_setting("header_right_text", header_right_text.strip())
                repo.set_setting("footer_left_text", footer_left_text.strip())
                repo.set_setting("footer_middle_text", footer_middle_text.strip())
                repo.set_setting("footer_right_text", footer_right_text.strip())
                st.success("PDF text settings saved.")

    with tab_images_pdf:
        # ---- Branding images (per company; outside forms for file upload) ----
        st.markdown("##### Branding images")
        bcol, lcol = st.columns([2, 1])
        with bcol:
            st.markdown("**Banner** - full-width header (app, PDF, project sheet)")
            if os.path.exists(db.banner_path()):
                st.image(db.banner_path(), width="stretch")
            else:
                st.info("No banner yet.")
            up_b = st.file_uploader("Upload / replace banner (PNG)", type=["png"], key="banner_up")
            bb1, bb2 = st.columns(2)
            if up_b is not None and bb1.button("Save banner", key="save_banner"):
                db.save_asset(db.banner_path(), bytes(up_b.getbuffer()), "image/png")
                st.success("Banner updated. (Reload to see it in the header/sidebar.)")
                st.rerun()
            if os.path.exists(db.banner_path()) and bb2.button("Remove banner", key="remove_banner"):
                db.delete_asset(db.banner_path())
                st.success("Banner removed. PDF header sections will be used if configured.")
                st.rerun()
        with lcol:
            st.markdown("**Logo** - standalone mark")
            if os.path.exists(db.logo_path()):
                st.image(db.logo_path(), width=160)
            else:
                st.info("No logo yet.")
            up_l = st.file_uploader("Upload / replace logo (PNG)", type=["png"], key="logo_up")
            if up_l is not None and st.button("Save logo", key="save_logo"):
                db.save_asset(db.logo_path(), bytes(up_l.getbuffer()), "image/png")
                st.success("Logo updated.")
                st.rerun()
        st.caption("Banner: wide and shallow PNG, about 1400x155 px. Logo: square transparent PNG works best.")

        st.divider()
        st.markdown("##### PDF header images")
        st.caption(
            "The banner is the full-width PDF header. If no banner is uploaded, the left/middle/right "
            "header images and text are used. Section images work best as wide, shallow PNGs with "
            "transparent or white backgrounds."
        )
        if os.path.exists(db.banner_path()):
            st.info(
                "A banner is currently uploaded, so header section images and text are saved but not "
                "used in the PDF until the banner is removed."
            )
        head_specs = [
            ("Left", db.header_left_path(), "header_left_up", "save_header_left", "remove_header_left"),
            ("Middle", db.header_middle_path(), "header_middle_up", "save_header_middle", "remove_header_middle"),
            ("Right", db.header_right_path(), "header_right_up", "save_header_right", "remove_header_right"),
        ]
        head_cols = st.columns(3)
        for col, (label, path, upload_key, save_key, remove_key) in zip(head_cols, head_specs):
            with col:
                st.markdown(f"**{label} section image**")
                if os.path.exists(path):
                    st.image(path, width="stretch")
                else:
                    st.info("No image yet.")
                up = st.file_uploader(
                    f"Upload / replace {label.lower()} image (PNG)",
                    type=["png"],
                    key=upload_key,
                )
                if up is not None and st.button(f"Save {label.lower()} image", key=save_key):
                    db.save_asset(path, bytes(up.getbuffer()), "image/png")
                    st.success(f"{label} header image updated.")
                    st.rerun()
                if os.path.exists(path) and st.button(f"Remove {label.lower()} image", key=remove_key):
                    db.delete_asset(path)
                    st.success(f"{label} header image removed.")
                    st.rerun()

        st.divider()
        st.markdown("##### PDF footer images")
        st.caption(
            "A full-width footer image overrides the three-section footer. If no full footer image is "
            "uploaded, the left/middle/right images and text are used. Keep footer images wide and "
            "shallow so they do not crowd quotation content."
        )
        full_col, preview_col = st.columns([2, 1])
        with full_col:
            st.markdown("**Full footer image** - full-width footer section")
            if os.path.exists(db.footer_full_path()):
                st.image(db.footer_full_path(), width="stretch")
            else:
                st.info("No full footer image yet.")
            up_footer_full = st.file_uploader(
                "Upload / replace full footer (PNG)",
                type=["png"],
                key="footer_full_up",
            )
            ffu1, ffu2 = st.columns(2)
            if up_footer_full is not None and ffu1.button("Save full footer", key="save_footer_full"):
                db.save_asset(db.footer_full_path(), bytes(up_footer_full.getbuffer()), "image/png")
                st.success("Full footer image updated.")
                st.rerun()
            if os.path.exists(db.footer_full_path()) and ffu2.button("Remove full footer", key="remove_footer_full"):
                db.delete_asset(db.footer_full_path())
                st.success("Full footer image removed.")
                st.rerun()
        with preview_col:
            st.info("Suggested full footer ratio: wide and shallow, similar to the header banner.")

        foot_specs = [
            ("Left", db.footer_left_path(), "footer_left_up", "save_footer_left", "remove_footer_left"),
            ("Middle", db.footer_middle_path(), "footer_middle_up", "save_footer_middle", "remove_footer_middle"),
            ("Right", db.footer_right_path(), "footer_right_up", "save_footer_right", "remove_footer_right"),
        ]
        foot_cols = st.columns(3)
        for col, (label, path, upload_key, save_key, remove_key) in zip(foot_cols, foot_specs):
            with col:
                st.markdown(f"**{label} section image**")
                if os.path.exists(path):
                    st.image(path, width="stretch")
                else:
                    st.info("No image yet.")
                up = st.file_uploader(
                    f"Upload / replace {label.lower()} image (PNG)",
                    type=["png"],
                    key=upload_key,
                )
                if up is not None and st.button(f"Save {label.lower()} image", key=save_key):
                    db.save_asset(path, bytes(up.getbuffer()), "image/png")
                    st.success(f"{label} footer image updated.")
                    st.rerun()
                if os.path.exists(path) and st.button(f"Remove {label.lower()} image", key=remove_key):
                    db.delete_asset(path)
                    st.success(f"{label} footer image removed.")
                    st.rerun()

    with tab_data:
        st.markdown("##### Import Excel workbooks")
        st.caption(
            "A safety backup is created automatically before importing. Imports work "
            "on any deployment (upload through the browser)."
        )

        # ---- Primary: upload through the browser (works locally AND when deployed) ----
        uploads = st.file_uploader(
            "Upload Excel workbooks (.xlsx / .xlsm) - or a single .zip of your project folder",
            type=["xlsx", "xlsm", "zip"],
            accept_multiple_files=True,
            key="import_uploads",
            help="Select the files from your computer. A .zip keeps any sub-folder structure.",
        )
        if st.button(
            "Import uploaded files",
            type="primary",
            width="stretch",
            disabled=not uploads,
        ):
            import tempfile, zipfile, shutil

            tmpdir = tempfile.mkdtemp(prefix="pq_import_")
            try:
                saved = 0
                for f in uploads:
                    if f.name.lower().endswith(".zip"):
                        try:
                            with zipfile.ZipFile(io.BytesIO(f.getbuffer())) as z:
                                z.extractall(tmpdir)
                            saved += 1
                        except zipfile.BadZipFile:
                            st.error(f"'{f.name}' is not a valid .zip file.")
                    else:
                        with open(os.path.join(tmpdir, f.name), "wb") as out:
                            out.write(f.getbuffer())
                        saved += 1
                if saved:
                    _run_excel_import(tmpdir)
            finally:
                shutil.rmtree(tmpdir, ignore_errors=True)

        # ---- Secondary: a folder already on the server / this machine ----
        with st.expander("Or import from a folder on the server (files must be on the host)"):
            # Apply a pending Browse pick BEFORE the widget is created. You cannot set a
            # widget-keyed session_state value after that widget exists in the same run.
            if "_import_folder_pick" in st.session_state:
                st.session_state.import_folder_path = st.session_state.pop("_import_folder_pick")
            if "import_folder_path" not in st.session_state:
                st.session_state.import_folder_path = repo.get_setting("last_import_folder") or ""
            ipath_col, browse_col = st.columns([4, 1])
            ipath_col.text_input(
                "Folder path on the server",
                key="import_folder_path",
                placeholder=r"C:\Projects\Old BoQ Files  (or /data/imports on a Linux host)",
            )
            # The OS folder dialog only works when the server IS a local Windows desktop.
            if os.name == "nt":
                if browse_col.button("Browse...", width="stretch"):
                    picked, pick_error = _choose_local_folder(st.session_state.import_folder_path)
                    if picked:
                        st.session_state["_import_folder_pick"] = picked   # deferred; applied next run
                        st.rerun()
                    elif pick_error:
                        st.warning(f"Folder picker unavailable: {pick_error}")
            if st.button("Import from server folder", width="stretch"):
                import_root = (st.session_state.import_folder_path or "").strip().strip('"')
                if not import_root:
                    st.warning("Enter a folder path first.")
                elif not os.path.isdir(import_root):
                    st.error("Folder not found on the server. Use the upload option above instead.")
                else:
                    repo.set_setting("last_import_folder", import_root)
                    _run_excel_import(import_root)

        st.divider()
        st.markdown("##### Clean database (migrated data)")
        if not owner:
            st.info("Only the owner role can run database clean-up.")
        else:
            st.caption(
                "Fix-ups for offers imported from old Excel files. Each shows a preview "
                "first and creates a safety backup before applying."
            )

            # ---- Bulk exact-value replacement for migrated project headers ----
            st.markdown("**Bulk project field cleanup**")
            st.caption(
                "Merge one or several misspelled values into the correct value. Changes are "
                "exact-match only and apply to every matching offer, revision, and option."
            )
            _bulk_result = st.session_state.pop("_bulk_cleanup_result", None)
            if _bulk_result:
                st.success(
                    f"Updated {_bulk_result['updated']} project record(s) in "
                    f"{_bulk_result['field']}. A safety backup was created first."
                )

            _bulk_fields = list(repo.PROJECT_CLEANUP_FIELDS)
            _bulk_field = st.selectbox(
                "Field to clean",
                _bulk_fields,
                key="bulk_cleanup_field",
            )
            _bulk_rows = repo.project_cleanup_values(_bulk_field)
            _bulk_counts = {row["Value"]: int(row["OfferCount"]) for row in _bulk_rows}
            _bulk_stored = list(_bulk_counts)
            _bulk_sources = st.multiselect(
                "Misspelled / old values to replace",
                _bulk_stored,
                key=f"bulk_cleanup_sources::{_bulk_field}",
                format_func=lambda value: f"{value} ({_bulk_counts[value]} record(s))",
                help="Select every typo variant that should become the same correct value.",
            )

            _bulk_canonical = []
            if _bulk_field == "Sales Person":
                _bulk_canonical = auth.users_in_roles(SALES_PERSON_ROLES)
            elif _bulk_field == "Pre-sales Engineer":
                _bulk_canonical = auth.users_in_role(PEOPLE_ROLES["presales"])
            elif _bulk_field == "Project Manager":
                _bulk_canonical = auth.users_in_role(PEOPLE_ROLES["pm"])
            _bulk_targets = sorted(
                {str(value).strip() for value in [*_bulk_stored, *_bulk_canonical]
                 if str(value).strip() and str(value).strip() not in _bulk_sources},
                key=str.casefold,
            )
            _bulk_new_marker = "(type a new correct value)"
            _bulk_target_choice = st.selectbox(
                "Correct replacement value",
                [_bulk_new_marker, *_bulk_targets],
                key=f"bulk_cleanup_target::{_bulk_field}",
                help="People lists include the active users allowed for that assignment.",
            )
            _bulk_custom = ""
            if _bulk_target_choice == _bulk_new_marker:
                _bulk_custom = st.text_input(
                    "New correct value",
                    key=f"bulk_cleanup_custom::{_bulk_field}",
                    placeholder="Type the exact spelling to store",
                )
            _bulk_replacement = (
                _bulk_custom.strip()
                if _bulk_target_choice == _bulk_new_marker
                else _bulk_target_choice
            )

            _bulk_preview = None
            if _bulk_sources and _bulk_replacement:
                _bulk_preview = repo.bulk_replace_project_field(
                    _bulk_field, _bulk_sources, _bulk_replacement, apply=False)
                if _bulk_preview["to_update"]:
                    st.write(
                        f"Will update **{_bulk_preview['to_update']}** project record(s) to "
                        f"**{_bulk_preview['replacement']}**."
                    )
                    st.dataframe(
                        pd.DataFrame(
                            _bulk_preview["sample"],
                            columns=["ProjectID", "Offer #", "Project", "Date",
                                     "Current value", "New value"],
                        ),
                        hide_index=True,
                        width="stretch",
                    )
                    if _bulk_preview["to_update"] > len(_bulk_preview["sample"]):
                        st.caption(
                            f"Showing the first {len(_bulk_preview['sample'])} of "
                            f"{_bulk_preview['to_update']} affected records."
                        )
                else:
                    st.info("No records need changing for this selection.")

            _bulk_ready = bool(_bulk_preview and _bulk_preview["to_update"])
            _bulk_confirm = st.checkbox(
                "I reviewed the affected offers and want to apply this replacement.",
                key=f"bulk_cleanup_confirm::{_bulk_field}",
                disabled=not _bulk_ready,
            )
            if st.button(
                "Apply bulk replacement (creates backup)",
                type="primary",
                key="apply_bulk_project_cleanup",
                disabled=not (_bulk_ready and _bulk_confirm),
            ):
                db_backup.create_profile_backup("before-project-field-cleanup")
                _bulk_applied = repo.bulk_replace_project_field(
                    _bulk_field, _bulk_sources, _bulk_replacement, apply=True)
                st.session_state["_bulk_cleanup_result"] = {
                    "updated": _bulk_applied["to_update"],
                    "field": _bulk_field,
                }
                for _bulk_key in (
                    f"bulk_cleanup_sources::{_bulk_field}",
                    f"bulk_cleanup_target::{_bulk_field}",
                    f"bulk_cleanup_custom::{_bulk_field}",
                    f"bulk_cleanup_confirm::{_bulk_field}",
                ):
                    st.session_state.pop(_bulk_key, None)
                st.rerun()

            st.markdown("---")

            # ---- 1) Stamp year from offer number ----
            st.markdown(
                "**Stamp date from project name / offer #** - uses the full date in the "
                "name (e.g. ...-24.03.2024 -> 2024-03-24); if there's only a year in the offer "
                "ref, it fixes the year and keeps the month/day."
            )
            if st.button("Preview year stamping", key="prev_year"):
                st.session_state["clean_year"] = repo.cleanup_stamp_years(apply=False)
            yp = st.session_state.get("clean_year")
            if yp:
                st.write(
                    f"Will update **{yp['to_update']}** | already correct {yp['already_ok']} | "
                    f"no year in offer # {yp['no_year']}"
                )
                if yp["sample"]:
                    st.dataframe(
                        pd.DataFrame(yp["sample"], columns=["ProjectID", "Offer #", "Old date", "New date"]),
                        hide_index=True,
                        width="stretch",
                    )
                    st.caption(
                        f"Sample only - showing the first {len(yp['sample'])} of "
                        f"**{yp['to_update']}**. Clicking Apply updates all {yp['to_update']}."
                    )
                if yp["to_update"] and st.button(
                    "Apply year stamping (creates backup)",
                    type="primary",
                    key="apply_year",
                ):
                    db_backup.create_profile_backup("before-year-cleanup")
                    res = repo.cleanup_stamp_years(apply=True)
                    st.session_state.pop("clean_year", None)
                    st.success(f"Stamped year on {res['to_update']} offers. Backup created first.")
                    st.rerun()

            st.markdown("---")
            # ---- 2) Merge revisions (Rxx) ----
            st.markdown(
                "**Merge revisions (Rxx)** - links offers like `...-R01 / -R02` into one "
                "offer with multiple revisions (R01 -> Rev 1, no suffix -> original)."
            )
            if st.button("Preview revision merge", key="prev_rev"):
                st.session_state["clean_rev"] = repo.cleanup_merge_revisions(apply=False)
            rp = st.session_state.get("clean_rev")
            if rp:
                st.write(
                    f"Will link **{rp['to_update']}** revision offers into **{rp['families']}** "
                    "offer families."
                )
                if rp["sample"]:
                    st.dataframe(
                        pd.DataFrame(rp["sample"], columns=["ProjectID", "Offer #", "Rev #", "Base offer #"]),
                        hide_index=True,
                        width="stretch",
                    )
                    st.caption(
                        f"Sample only - showing the first {len(rp['sample'])} of "
                        f"**{rp['to_update']}**. Clicking Apply updates all {rp['to_update']}."
                    )
                if rp["to_update"] and st.button(
                    "Apply revision merge (creates backup)",
                    type="primary",
                    key="apply_rev",
                ):
                    db_backup.create_profile_backup("before-revision-cleanup")
                    res = repo.cleanup_merge_revisions(apply=True)
                    st.session_state.pop("clean_rev", None)
                    st.success(
                        f"Linked {res['to_update']} revisions into {res['families']} families. "
                        "Backup created first."
                    )
                    st.rerun()

            st.markdown("---")
            # ---- 3) Parse client from project name ----
            st.markdown(
                "**Parse client from project name** - fills the **Client** field (only where "
                "it's blank) from the name, e.g. `...Al Rashed Farm-Laptop...` -> **Al Rashed Farm**."
            )
            if st.button("Preview client parsing", key="prev_client"):
                st.session_state["clean_client"] = repo.cleanup_parse_clients(apply=False)
            cp = st.session_state.get("clean_client")
            if cp:
                st.write(f"Will set Client on **{cp['to_update']}** offers that are currently blank.")
                if cp["sample"]:
                    st.dataframe(
                        pd.DataFrame(cp["sample"], columns=["ProjectID", "Project name", "-> Client"]),
                        hide_index=True,
                        width="stretch",
                    )
                    st.caption(
                        f"Sample only - showing the first {len(cp['sample'])} of "
                        f"**{cp['to_update']}**. Review these - if the client looks wrong, "
                        "don't apply and tell me the pattern."
                    )
                if cp["to_update"] and st.button(
                    "Apply client parsing (creates backup)",
                    type="primary",
                    key="apply_client",
                ):
                    db_backup.create_profile_backup("before-client-cleanup")
                    res = repo.cleanup_parse_clients(apply=True)
                    st.session_state.pop("clean_client", None)
                    st.success(f"Set client on {res['to_update']} offers. Backup created first.")
                    st.rerun()

            st.markdown("---")
            st.markdown("##### Cleanup catalogue items")
            st.caption("Find duplicate catalogue items with the same Model + Description, "
                       "then review and delete selected copies.")
            _catalog_dedupe_tool()

            st.markdown("---")
            # ---- Clear everything for a clean re-import ----
            st.markdown("**Clear all imported data (for a fresh re-import)**")
            st.caption(
                "Deletes ALL projects (and their lines, finance) and the catalogue - keeps "
                "users, roles, settings and branding. A full backup is created first. Use this "
                "before re-importing so the import doesn't create duplicates."
            )
            clr_ok = st.checkbox(
                "Yes - wipe all projects & catalogue (a backup is created first).",
                key="clear_confirm",
            )
            if st.button("Clear & prepare for fresh import", disabled=not clr_ok, key="do_clear"):
                bk = db_backup.create_profile_backup("before-fresh-import")
                res = repo.clear_imported_data()
                st.session_state.pop("clear_confirm", None)
                st.success(
                    f"Cleared {res['projects']} projects and {res['catalogue']} catalogue items. "
                    f"Backup: {os.path.basename(bk)}. Now re-import your files (top of this page)."
                )
                st.rerun()

    with tab_backup:
        st.markdown("##### Backup & Restore")
        st.caption(
            "Backups are `.zip` files that include the database and branding images from `assets/`."
        )
        try:
            _counts = repo.db_counts()
            db1, db2 = st.columns(2)
            db1.metric("Projects in DB", _counts["project_families"])
            db1.caption(f"{_counts['project_records']} records incl. revisions & options")
            db2.metric("Catalogue items", _counts["catalogue_items"])
        except Exception as e:
            st.error(f"DB: {e}")
        st.divider()

        if not owner:
            st.info("Only the owner role can create, download, or restore backups.")
        else:
            b1, b2, b3 = st.columns([1, 1, 2])
            if b1.button("Create backup", width="stretch"):
                try:
                    backup_path = db_backup.create_profile_backup("manual")
                    st.session_state.latest_profile_backup = backup_path
                    st.success(f"Backup created: {os.path.basename(backup_path)}")
                except Exception as exc:
                    st.error(f"Backup failed: {exc}")

            latest_backup = st.session_state.get("latest_profile_backup")
            if latest_backup and os.path.exists(latest_backup):
                with open(latest_backup, "rb") as f:
                    b2.download_button(
                        "Download backup",
                        data=f.read(),
                        file_name=os.path.basename(latest_backup),
                        mime="application/zip",
                        width="stretch",
                    )
            else:
                b2.button("Download backup", disabled=True, width="stretch")

            backups = db_backup.list_profile_backups(limit=5)
            if backups:
                names = [f"{b['name']} ({b['size'] / 1024 / 1024:.1f} MB)" for b in backups]
                pick = b3.selectbox("Recent backups", names, label_visibility="collapsed")
                picked_backup = backups[names.index(pick)]
                with open(picked_backup["path"], "rb") as f:
                    b3.download_button(
                        "Download selected recent backup",
                        data=f.read(),
                        file_name=picked_backup["name"],
                        mime="application/zip",
                        width="stretch",
                    )
            else:
                b3.info("No local backups yet.")

            restore_file = st.file_uploader(
                "Restore from backup (.zip)",
                type=["zip"],
                key="restore_zip_upload",
            )
            restore_ok = st.checkbox(
                "I understand restore replaces the current database and branding images. A safety backup will be created first.",
                key="restore_zip_confirm",
            )
            if st.button(
                "Restore backup",
                type="primary",
                disabled=not (restore_file and restore_ok),
                width="stretch",
            ):
                try:
                    restored_path, safety_backup = db_backup.restore_profile_from_bytes(restore_file.getvalue())
                    restored_conn = db.init_db(restored_path)
                    restored_conn.close()
                    db.set_audit_actor(USER)
                    audit_log.record_event(
                        "RESTORE", "Database", os.path.basename(restored_path),
                        f"Restored profile backup {restore_file.name}",
                        new_values={"backup_file": restore_file.name},
                    )
                    st.session_state.db_init_v2 = True
                    st.success("Backup restored. The app will reload now.")
                    if safety_backup:
                        st.info(f"Safety backup created: {os.path.basename(safety_backup)}")
                    st.cache_data.clear()
                    st.cache_resource.clear()
                    st.rerun()
                except Exception as exc:
                    st.error(f"Restore failed: {exc}")

            st.divider()
            st.markdown("##### Catalogue only")
            st.caption("Back up or restore just the catalogue (item list), independent of "
                       "the full database backup above.")
            if st.button("Prepare catalogue backup (.zip)", width="stretch", key="prep_cat_backup"):
                _cat_all = repo.catalog_all()
                if _cat_all.empty:
                    st.warning("Catalogue is empty.")
                else:
                    st.session_state["catalogue_backup_zip"] = _catalogue_zip_bytes(_cat_all)
                    st.session_state["catalogue_backup_name"] = f"catalogue_backup_{dt.date.today().isoformat()}.zip"
                    st.success(f"Catalogue backup ready - {_cat_all.shape[0]:,} item(s).")
            if st.session_state.get("catalogue_backup_zip"):
                st.download_button(
                    "⬇️ Download prepared catalogue backup",
                    st.session_state["catalogue_backup_zip"],
                    file_name=st.session_state.get(
                        "catalogue_backup_name",
                        f"catalogue_backup_{dt.date.today().isoformat()}.zip",
                    ),
                    mime="application/zip",
                    width="stretch",
                )
            st.markdown("**Restore catalogue** — replaces the entire catalogue with an uploaded backup.")
            _cat_up = st.file_uploader("Catalogue backup (.zip)", type=["zip"], key="cat_restore_up")
            _cat_ok = st.checkbox("I understand this replaces the current catalogue "
                                  "(a full-database safety backup is made first).", key="cat_restore_ok")
            if st.button("♻️ Restore catalogue", type="primary",
                         disabled=not (_cat_up is not None and _cat_ok), width="stretch"):
                try:
                    _newdf = _catalogue_df_from_zip(_cat_up.getvalue())
                    db_backup.create_backup("before-catalogue-restore")
                    n = repo.replace_catalog(_newdf)
                    st.success(f"Catalogue restored — {n} item(s). (Safety backup created.)")
                    st.rerun()
                except Exception as e:
                    st.error(f"Restore failed: {e}")

            if db.is_postgres():
                st.divider()
                st.markdown("##### Migrate from SQLite")
                st.caption(
                    "Import a ProQuote SQLite profile backup (`.zip` containing `proquote.db`) "
                    "into this PostgreSQL database. All existing PostgreSQL data will be replaced. "
                    "A safety backup of the current PostgreSQL data is created first."
                )
                _mig_file = st.file_uploader(
                    "SQLite profile backup (.zip)",
                    type=["zip"],
                    key="migrate_sqlite_upload",
                )
                _mig_ok = st.checkbox(
                    "I understand this replaces ALL data in the PostgreSQL database with the "
                    "contents of the uploaded SQLite backup.",
                    key="migrate_sqlite_confirm",
                )
                if st.button(
                    "⬆️ Migrate SQLite → PostgreSQL",
                    type="primary",
                    disabled=not (_mig_file and _mig_ok),
                    width="stretch",
                    key="migrate_sqlite_btn",
                ):
                    import tempfile as _tempfile
                    import db_transfer as _db_transfer
                    try:
                        with _tempfile.TemporaryDirectory(
                            prefix="proquote-migrate-", dir=db.DATA_DIR
                        ) as _tmp:
                            _zip_path = os.path.join(_tmp, "upload.zip")
                            with open(_zip_path, "wb") as _f:
                                _f.write(_mig_file.getvalue())
                            _vok, _vmsg = db_backup.validate_profile_backup(_zip_path)
                            if not _vok:
                                st.error(f"Invalid backup: {_vmsg}")
                            else:
                                with zipfile.ZipFile(_zip_path) as _zf:
                                    if "proquote.db" not in _zf.namelist():
                                        st.error(
                                            "This ZIP does not contain `proquote.db`. "
                                            "Upload a SQLite profile backup, not a portable PostgreSQL backup."
                                        )
                                    else:
                                        _zf.extract("proquote.db", _tmp)
                                        _sqlite_path = os.path.join(_tmp, "proquote.db")
                                        with st.spinner("Creating safety backup of current PostgreSQL data…"):
                                            _safety = db_backup.create_profile_backup("before-sqlite-migration")
                                        st.info(f"Safety backup: {os.path.basename(_safety)}")
                                        _mig_log = []
                                        _mig_placeholder = st.empty()
                                        def _on_mig_progress(table, count):
                                            _mig_log.append(f"- **{table}**: {count:,} rows")
                                            _mig_placeholder.markdown("\n".join(_mig_log))
                                        with st.spinner("Migrating…"):
                                            _result = _db_transfer.migrate_sqlite_to_postgres(
                                                _sqlite_path, db.database_url(),
                                                replace=True, progress=_on_mig_progress,
                                            )
                                        st.cache_data.clear()
                                        _copied = _result.get("copied", {})
                                        _rows = [(t, c) for t, c in _copied.items() if c > 0]
                                        if _rows:
                                            st.dataframe(
                                                pd.DataFrame(_rows, columns=["Table", "Rows copied"]),
                                                hide_index=True, use_container_width=True,
                                            )
                                        if _result.get("skipped_orphans"):
                                            st.warning(f"Skipped orphan rows: {_result['skipped_orphans']}")
                                        if _result.get("repaired_values"):
                                            st.info(f"Repaired values: {_result['repaired_values']}")
                                        st.success("Migration complete! Reloading…")
                                        st.rerun()
                    except Exception as _exc:
                        st.error(f"Migration failed: {_exc}")

    with tab_updates:
        st.markdown("##### Software updates")
        gh_owner = repo.get_setting("github_owner") or "Hollako"
        gh_repo = repo.get_setting("github_repo") or "_ProQuote"
        st.metric("Installed version", APP_VERSION)
        git_update_ok, git_update_message = runtime_env.git_update_available(db.APP_DIR)
        if git_update_message:
            st.caption(git_update_message)

        if st.button("Check for updates", width="stretch"):
            try:
                rel = updater.latest_release(gh_owner, gh_repo)
                st.session_state.latest_release = rel
                st.session_state.update_available = updater.is_newer(rel.tag, APP_VERSION)
            except Exception as exc:
                st.session_state.latest_release = None
                st.session_state.update_available = False
                st.error(str(exc))

        rel = st.session_state.get("latest_release")
        if rel:
            if st.session_state.get("update_available"):
                st.warning(f"New {rel.source} available: **{rel.tag}** ({rel.name})")
                if rel.url:
                    st.link_button("Open GitHub page", rel.url, width="stretch")
                if git_update_ok:
                    if st.button("Update this instance", type="primary", width="stretch"):
                        ok, output = updater.run_git_update(db.APP_DIR)
                        if ok:
                            st.success("Update downloaded. Restart this Streamlit instance to load the new code.")
                        else:
                            st.error("Update failed. Details below.")
                        st.code(output or "No output", language="text")
                else:
                    st.info(
                        git_update_message
                        or "Use your deployment platform or a newer installer to update this instance."
                    )
            else:
                st.success(f"You are up to date. Latest {rel.source}: {rel.tag or rel.name}")

# ============================ USERS (owner) ============================
elif mode == "Users":
    st.subheader("Users & access")
    tab_users, tab_roles = st.tabs(["👤 Users", "🛡️ Roles & permissions"])

    # ---------------------------- USERS TAB ----------------------------
    with tab_users:
        roles = auth.list_roles()
        _default_role = "viewer" if "viewer" in roles else roles[-1]
        with st.expander("➕ Add user", expanded=False):
            with st.form("add_user", clear_on_submit=True):
                uc1, uc2 = st.columns(2)
                nu = uc1.text_input("Username")
                ndn = uc2.text_input("Display name")
                uc3, uc4, uc5 = st.columns(3)
                np1 = uc3.text_input("Password", type="password")
                np2 = uc4.text_input("Confirm", type="password")
                nrole = uc5.selectbox("Role", roles, index=roles.index(_default_role),
                                      format_func=_role_label)
                if st.form_submit_button("➕ Create user", type="primary"):
                    if not nu.strip() or not np1:
                        st.warning("Username and password are required.")
                    elif np1 != np2:
                        st.warning("Passwords don't match.")
                    elif auth.create_user(nu, np1, ndn, nrole):
                        st.success(f"Created user '{nu}' ({_role_label(nrole)}).")
                        st.rerun()
                    else:
                        st.warning("Username already exists.")

        st.markdown("##### Existing users")
        users = auth.list_users()
        owners = sum(1 for x in users if x["Role"] == auth.PROTECTED_ROLE and x["Active"])
        for u in users:
            uid = u["UserID"]
            is_self = (uid == USER.get("UserID"))
            last_owner = (u["Role"] == auth.PROTECTED_ROLE and owners <= 1)
            with st.container(border=True):
                cc = st.columns([2, 2, 1.5, 1, 1.2], vertical_alignment="center")
                cc[0].markdown(f"**{u['Username']}**" + (" · _(you)_" if is_self else ""))
                cc[1].write(u.get("DisplayName") or "")
                r_idx = roles.index(u["Role"]) if u["Role"] in roles else 0
                new_role = cc[2].selectbox("Role", roles, key=f"role_{uid}",
                                           index=r_idx, label_visibility="collapsed",
                                           format_func=_role_label)
                active = cc[3].toggle("On", value=bool(u["Active"]), key=f"act_{uid}",
                                      label_visibility="collapsed")
                if (new_role != u["Role"] or active != bool(u["Active"])):
                    if cc[4].button("💾 Save", key=f"saveu_{uid}", width="stretch"):
                        if last_owner and (new_role != auth.PROTECTED_ROLE or not active):
                            st.warning("Can't remove the last active owner.")
                        else:
                            auth.update_user(uid, role=new_role, active=active)
                            st.toast(f"Updated {u['Username']}.")
                            st.rerun()
                pc = st.columns([3, 1.3, 1.3])
                newpw = pc[0].text_input("pw", type="password", key=f"pw_{uid}",
                                         label_visibility="collapsed",
                                         placeholder="New password (blank = keep)")
                if pc[1].button("🔑 Set password", key=f"setpw_{uid}", width="stretch",
                                disabled=not newpw):
                    auth.set_password(uid, newpw)
                    st.toast(f"Password updated for {u['Username']}.")
                    st.rerun()
                if pc[2].button("Delete", key=f"delu_{uid}", width="stretch",
                                disabled=is_self or last_owner):
                    auth.delete_user(uid)
                    st.toast(f"Deleted {u['Username']}.")
                    st.rerun()

    # ---------------------- ROLES & PERMISSIONS TAB --------------------
    with tab_roles:
        st.caption("Tick what each role is allowed to do, then **Save**. "
                   "The **Owner** always has full access and can't be changed.")
        roles = auth.list_roles()
        rows = []
        for role in roles:
            perms = auth.role_perms(role)
            row = {"Role": _role_label(role)}
            for k, label in auth.PERMISSIONS:
                row[label] = (k in perms)
            rows.append(row)
        mat = pd.DataFrame(rows, columns=["Role"] + [lbl for _, lbl in auth.PERMISSIONS])

        colcfg = {"Role": st.column_config.TextColumn("Role", disabled=True, width="small")}
        for _, label in auth.PERMISSIONS:
            colcfg[label] = st.column_config.CheckboxColumn(label)
        edited = st.data_editor(mat, column_config=colcfg, hide_index=True,
                                num_rows="fixed", width="stretch",
                                disabled=["Role"], key="role_matrix")

        if st.button("💾 Save role permissions", type="primary"):
            for i, (_, r) in enumerate(edited.iterrows()):
                role = roles[i]                       # actual key (the matrix shows a capitalized label)
                if role == auth.PROTECTED_ROLE:
                    continue
                granted = {auth.LABEL_TO_PERM[lbl] for _, lbl in auth.PERMISSIONS if bool(r[lbl])}
                auth.set_role_perms(role, granted)
            st.success("Role permissions saved.")
            st.rerun()

        st.divider()
        rc1, rc2 = st.columns(2)
        with rc1:
            st.markdown("**Add a role**")
            arc1, arc2 = st.columns([3, 1.4], vertical_alignment="bottom")
            new_role_name = arc1.text_input("New role name", key="new_role_name",
                                            label_visibility="collapsed",
                                            placeholder="e.g. Estimator")
            if arc2.button("➕ Add", width="stretch"):
                if auth.add_role(new_role_name):
                    st.toast(f"Added role '{new_role_name.strip()}'.")
                    st.rerun()
                else:
                    st.warning("Empty or duplicate role name.")
        with rc2:
            st.markdown("**Delete a role**")
            deletable = [r for r in roles if r != auth.PROTECTED_ROLE]
            drc1, drc2 = st.columns([3, 1.4], vertical_alignment="bottom")
            drole = drc1.selectbox("Role to delete", deletable, key="del_role",
                                   label_visibility="collapsed",
                                   format_func=_role_label) if deletable else None
            if drc2.button("Delete", width="stretch", disabled=not deletable):
                in_use = auth.role_user_count(drole)
                if in_use:
                    st.warning(f"{in_use} user(s) still have the '{drole}' role - "
                               "reassign them first.")
                elif auth.delete_role(drole):
                    st.toast(f"Deleted role '{drole}'.")
                    st.rerun()
