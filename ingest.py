"""
Folder-wide BoQ ingestion engine.

Scans the assigned project directory, finds every `BOQ <System>` sheet (and its
matching `Quotation <System>` sheet for client metadata), parses the canonical
grid, and upserts into the SQLite master database.

Idempotent: re-running re-ingests each file cleanly (old rows for that file are
removed first), so it is safe to run repeatedly as the folder grows.

Usage:
    python ingest.py                 # ingest the default ROOT folder
    python ingest.py "C:\\some\\path" # ingest a different folder
"""
import os
import re
import sys
import glob
import json
import datetime as dt
import warnings

warnings.simplefilter("ignore")
import openpyxl

import calc
import db as dbmod
import repo

DEFAULT_ROOT = r"J:\My Drive\1-Projects"


def _parse_date_text(text):
    """ISO date from a string: a DD.MM.YYYY anywhere, or a worded date like
    'May 13,2026'. Returns 'YYYY-MM-DD' or None."""
    s = str(text or "").strip()
    if not s:
        return None
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(20\d{2})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= d <= 31 and 1 <= mo <= 12:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    for fmt in ("%B %d,%Y", "%B %d, %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None

# Canonical BOQ columns -> normalized keys. Detection matches these labels
# (lowercased, punctuation-stripped) against the header row.
HEADER_MAP = {
    "area": "area",
    "system": "system",
    "description": "description",
    "brand": "brand",
    "model": "model",
    "qty": "qty",
    "quantity": "qty",
    "list price": "list_price",
    "list price $": "list_price",
    "ex unit cost": "ex_unit_cost",
    "ex unit cost $": "ex_unit_cost",
    "unit cost": "unit_cost",
    "unit cost $": "unit_cost",
    "total cost": "total_cost",
    "total cost $": "total_cost",
    "u price": "u_price",
    "u price $": "u_price",
    "t price": "t_price",
    "t price $": "t_price",
    "u price sar": "u_price_sar",
    "t price sar": "t_price_sar",
}

REQUIRED_KEYS = {"description", "qty"}  # minimum to trust a header row


def norm(v):
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = s.replace("$", "").replace(".", " ").replace("/", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("$", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def find_header(ws, scan_rows=12):
    """Return (header_row_index, {key: col_index}) or (None, None)."""
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        mapping = {}
        for c in range(1, ws.max_column + 1):
            key = HEADER_MAP.get(norm(ws.cell(row=r, column=c).value))
            if key and key not in mapping:
                mapping[key] = c
        if REQUIRED_KEYS.issubset(mapping.keys()) and "model" in mapping:
            return r, mapping
    return None, None


def cell(ws, r, colmap, key):
    c = colmap.get(key)
    return ws.cell(row=r, column=c).value if c else None


def parse_boq_sheet(ws):
    """Parse one BOQ sheet -> (lines, sheet_meta)."""
    hdr, colmap = find_header(ws)
    if hdr is None:
        return None, None

    last_data_col = max(colmap.values())
    lines = []
    factors = []
    discount = 0.0
    order = 0
    in_summary = False   # True once the totals/factor block starts -> rows below = spare
    empty_run = 0        # consecutive blank rows (the 3-4 row gap before spares)
    summary_subtotal = None  # the BOQ's own offer subtotal (first totals row, SAR)

    for r in range(hdr + 1, ws.max_row + 1):
        desc = cell(ws, r, colmap, "description")
        desc_s = norm(desc)
        qty = to_num(cell(ws, r, colmap, "qty"))
        model = cell(ws, r, colmap, "model")
        brand = cell(ws, r, colmap, "brand")

        # Discount row (still part of the offer)
        if desc_s == "discount":
            dval = to_num(cell(ws, r, colmap, "t_price_sar")) \
                or to_num(cell(ws, r, colmap, "total_cost")) \
                or to_num(cell(ws, r, colmap, "t_price"))
            if dval:
                dval = abs(dval)
                discount = dval
            order += 1
            lines.append({"_type": "discount", "_order": order, "description": "Discount",
                          "qty": None, "t_price_sar": dval})
            empty_run = 0
            continue

        has_label = bool((desc and str(desc).strip()) or (model and str(model).strip()))

        # No item label and no qty -> blank / totals / conversion-factor row.
        # The first such non-blank row (or a 3+ row gap) marks the end of the
        # offer; everything below it is a parked "spare" device.
        if not has_label and qty is None:
            row_vals = [to_num(ws.cell(row=r, column=c).value)
                        for c in range(1, last_data_col + 2)]
            nums = [v for v in row_vals if v is not None]
            if not nums:                                # truly empty row
                empty_run += 1
                if empty_run >= 3 and order > 0:
                    in_summary = True
                continue
            empty_run = 0
            if len(nums) == 1 and 1.0 <= nums[0] <= 2.5:
                factors.append(nums[0])                 # 1.69 / 1.49 / 1.42
            elif summary_subtotal is None:              # first totals row = offer subtotal
                summary_subtotal = to_num(cell(ws, r, colmap, "t_price_sar"))
            in_summary = True                           # totals/factor row ends the offer
            continue

        # Item row.
        if has_label and qty is not None:
            empty_run = 0
            if in_summary:
                ltype = "spare"                         # parked below the block
            elif (qty == 1 and not to_num(cell(ws, r, colmap, "ex_unit_cost"))
                  and norm(model) in ("sws", "")):
                ltype = "service"
            else:
                ltype = "item"
            order += 1
            ex_unit_cost = to_num(cell(ws, r, colmap, "ex_unit_cost"))
            unit_cost = to_num(cell(ws, r, colmap, "unit_cost"))
            shipping_percent = calc.infer_shipping_percent(ex_unit_cost, unit_cost)
            editor_unit_cost = (
                calc.roundup(unit_cost, 0) if unit_cost is not None
                else calc.unit_cost_from_ex(ex_unit_cost, shipping_percent)
            )
            total_cost = to_num(cell(ws, r, colmap, "total_cost"))
            u_price = to_num(cell(ws, r, colmap, "u_price"))
            t_price = to_num(cell(ws, r, colmap, "t_price"))
            lines.append({
                "_type": ltype, "_order": order,
                "area": cell(ws, r, colmap, "area"),
                "system": cell(ws, r, colmap, "system"),
                "description": str(desc).strip() if desc else None,
                "brand": str(brand).strip() if brand else None,
                "model": str(model).strip() if model else None,
                "qty": qty,
                "list_price": to_num(cell(ws, r, colmap, "list_price")),
                "ex_unit_cost": ex_unit_cost,
                "shipping_percent": shipping_percent,
                "unit_cost": unit_cost,
                "total_cost": total_cost,
                "u_price": u_price,
                "t_price": t_price,
                "u_price_sar": to_num(cell(ws, r, colmap, "u_price_sar")),
                "t_price_sar": to_num(cell(ws, r, colmap, "t_price_sar")),
                "margin": _pricing_margin(total_cost, t_price, editor_unit_cost, u_price),
            })
            continue

        # has_label but no qty (e.g. a section sub-header) -> ignore.
        empty_run = 0

    meta = {"discount": discount, "factors": factors[:3], "subtotal_sar": summary_subtotal}
    return lines, meta


def _pricing_margin(total_cost, total_price, unit_cost, unit_price):
    """Derive the imported pricing multiplier from selling price / cost.

    Imported workbooks often contain unrelated formulas immediately after the
    named BOQ columns, so that trailing cell must never be treated as Margin x.
    Prefer unit values because the editor recalculates ``U. Price = Unit Cost x
    Margin``; this preserves the imported selling price even when workbook totals
    contain fractional-cost rounding. Line totals are the fallback.
    """
    for cost, price in ((unit_cost, unit_price), (total_cost, total_price)):
        cost_num, price_num = to_num(cost), to_num(price)
        if cost_num is not None and cost_num > 0 and price_num is not None:
            return round(max(price_num / cost_num, 0.0), 4)
    return 0.0


def _find_quotation_sheet(wb, suffix):
    name = f"Quotation {suffix}".strip()
    if name in wb.sheetnames:
        return wb[name]
    for s in wb.sheetnames:
        if s.lower().startswith("quotation") and suffix.lower() in s.lower():
            return wb[s]
    for s in wb.sheetnames:                      # fall back to any quotation sheet
        if s.lower().startswith("quotation"):
            return wb[s]
    return None


def parse_quotation_meta(wb, suffix):
    """Pull client/project/offer metadata from the matching Quotation sheet."""
    ws = _find_quotation_sheet(wb, suffix)
    meta = {"client": None, "project": None, "contact": None, "phone": None,
            "offer": None, "date": None, "sales": None}
    if ws is None:
        return meta
    labels = {
        "client name": "client", "project name": "project",
        "contact": "contact", "phone": "phone", "offer #": "offer", "offer no": "offer",
        "date": "date", "date:": "date", "billed to": None,
    }
    for r in range(1, min(ws.max_row, 20) + 1):
        for c in range(1, min(ws.max_column, 9) + 1):
            key = norm(ws.cell(row=r, column=c).value)
            if key in labels and labels[key] is not None:
                # value is the next non-empty cell to the right, but stop if we
                # hit another label cell (blank value field -> leave as None).
                for cc in range(c + 1, min(ws.max_column, 9) + 1):
                    val = ws.cell(row=r, column=cc).value
                    if val in (None, ""):
                        continue
                    if norm(val) in labels:          # adjacent label, not a value
                        break
                    meta[labels[key]] = val
                    break

    # Single-cell labelled fields ("M/S MJS", "Project:Rowleys...", "Reference: ...",
    # "From: ...", "Date:..."). Only fill what the label|value pass left blank.
    prefixes = [("m/s", "client"), ("messrs", "client"),
                ("project:", "project"), ("project :", "project"),
                ("reference:", "offer"), ("offer #:", "offer"), ("offer no:", "offer"),
                ("from:", "sales"), ("date:", "date"), ("date :", "date"),
                ("phone:", "phone"), ("phone :", "phone"),
                ("attn:", "contact"), ("attn :", "contact"), ("attention:", "contact")]
    for r in range(1, min(ws.max_row, 20) + 1):
        for c in range(1, min(ws.max_column, 9) + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            vs = v.strip()
            low = vs.lower()
            for pref, field in prefixes:
                if low.startswith(pref) and not meta.get(field):
                    val = vs[len(pref):].strip(" :\t ")
                    if val:
                        meta[field] = val
                    break
    return meta


# Quotation note labels -> offer-terms keys (longest first for greedy matching).
QTERM_LABELS = sorted([
    ("special notes and instructions", "notes"), ("special notes", "notes"),
    ("pre-requirements", "prerequisites"), ("pre requirements", "prerequisites"),
    ("prerequisites", "prerequisites"),
    ("payment terms", "payment"), ("payment", "payment"),
    ("scope of work", "scope"), ("scope", "scope"),
    ("exclusions", "exclusions"), ("exclusion", "exclusions"),
    ("delivery time", "delivery"), ("delivery", "delivery"),
    ("offer validity", "validity"), ("validity", "validity"),
    ("warranty", "notes"), ("remarks", "notes"),
    ("system", "system_note"),
], key=lambda x: len(x[0]), reverse=True)


def parse_quotation_terms(wb, suffix):
    """Extract the offer Subject (title), greeting and note blocks (scope, payment,
    exclusions, delivery, validity, system, ...) from the Quotation sheet."""
    ws = _find_quotation_sheet(wb, suffix)
    terms = {}
    if ws is None:
        return terms
    mc = min(ws.max_column, 10)

    hdr = None
    for r in range(1, min(ws.max_row, 30) + 1):
        rowtext = " | ".join(norm(ws.cell(row=r, column=c).value)
                             for c in range(1, mc + 1) if ws.cell(row=r, column=c).value)
        if "description" in rowtext and ("unit price" in rowtext or "qty" in rowtext):
            hdr = r
            break

    # Subject (title cell ending in 'Offer') + greeting, above the items table.
    for r in range(1, (hdr or 12)):
        for c in range(1, mc + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            vn = norm(v)
            if not terms.get("subject") and vn.endswith("offer") and 3 < len(vn) < 55:
                terms["subject"] = str(v).strip()
            if not terms.get("greeting") and vn.startswith("dear"):
                terms["greeting"] = str(v).strip()

    # Note blocks below the items table, left columns only.
    for r in range((hdr or 14) + 1, ws.max_row + 1):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str) or not v.strip():
                continue
            vn = norm(v)
            for lab, key in QTERM_LABELS:
                if vn.startswith(lab) and not terms.get(key):
                    pat = re.compile(r"^\s*" + r"\s+".join(re.escape(w) for w in lab.split())
                                     + r"\s*[:\-]?\s*", re.I)
                    val = pat.sub("", str(v).strip(), count=1).strip()
                    if val:
                        terms[key] = val
                    break
    return terms


def system_suffix(sheet_name):
    return re.sub(r"^\s*boq\s*", "", sheet_name, flags=re.I).strip()


# Sheets that are never the priced BoQ grid (matched by name prefix).
_NON_GRID_SHEETS = ("pivot", "quotation", "mark up", "markup", "sheet", "balance",
                    "summary", "cost", "profit")


def _detect_boq_sheets(wb):
    """BoQ grid sheets. Prefer the canonical 'BOQ*'-named sheet(s) (original
    behaviour). ONLY when a file has no 'BOQ'-named sheet do we detect the grid by
    its header columns (the grid is then named by system/area, e.g. 'Sound System',
    'LANDSCAPE AREA'). This avoids picking up draft/duplicate sheets like
    'Lighting Control BOQ' alongside the real 'BOQ'."""
    named = [s for s in wb.sheetnames if s.lower().strip().startswith("boq")]
    if named:
        return named
    out = []
    for s in wb.sheetnames:
        sl = s.lower().strip()
        if any(sl.startswith(x) for x in _NON_GRID_SHEETS):
            continue
        try:
            hdr, _ = find_header(wb[s])
        except Exception:
            hdr = None
        if hdr is not None:
            out.append(s)
    return out


def upsert_item(conn, ln, src_file, now):
    """Upsert into Items_Catalog by (Brand, Model, Description); refresh defaults."""
    if not (ln.get("description") or ln.get("model")):
        return None
    cur = conn.execute(
        """SELECT ItemID,ListPriceUSD,ExUnitCostUSD,ShippingPercent,UnitCostUSD,
                  DefaultUPriceUSD,DefaultUPriceSAR
             FROM Items_Catalog
            WHERE IFNULL(Brand,'')=? AND IFNULL(Model,'')=? AND IFNULL(Description,'')=?""",
        (ln.get("brand") or "", ln.get("model") or "", ln.get("description") or ""),
    )
    row = cur.fetchone()
    if row:
        iid = row["ItemID"]
        price_cols = {
            "list_price": "ListPriceUSD",
            "ex_unit_cost": "ExUnitCostUSD",
            "shipping_percent": "ShippingPercent",
            "unit_cost": "UnitCostUSD",
            "u_price": "DefaultUPriceUSD",
            "u_price_sar": "DefaultUPriceSAR",
        }
        price_changed = any(
            ln.get(src) is not None and abs(calc._num(ln.get(src)) - calc._num(row[dst])) > 1e-9
            for src, dst in price_cols.items()
        )
        conn.execute(
            """UPDATE Items_Catalog SET
                 ListPriceUSD=COALESCE(?,ListPriceUSD),
                 ExUnitCostUSD=COALESCE(?,ExUnitCostUSD),
                 ShippingPercent=COALESCE(?,ShippingPercent),
                 UnitCostUSD=COALESCE(?,UnitCostUSD),
                 DefaultUPriceUSD=COALESCE(?,DefaultUPriceUSD),
                 DefaultUPriceSAR=COALESCE(?,DefaultUPriceSAR),
                 PriceUpdatedAt=COALESCE(?,PriceUpdatedAt),
                 TimesQuoted=TimesQuoted+1, LastSeenFile=?, LastSeenAt=?
               WHERE ItemID=?""",
            (ln.get("list_price"), ln.get("ex_unit_cost"), ln.get("shipping_percent"), ln.get("unit_cost"),
             ln.get("u_price"), ln.get("u_price_sar"),
             dt.date.today().isoformat() if price_changed else None, src_file, now, iid),
        )
        return iid
    cur = conn.execute(
        """INSERT INTO Items_Catalog
              (Description,Brand,Model,ListPriceUSD,ExUnitCostUSD,ShippingPercent,UnitCostUSD,
               DefaultUPriceUSD,DefaultUPriceSAR,PriceUpdatedAt,TimesQuoted,LastSeenFile,LastSeenAt)
            VALUES (?,?,?,?,?,?,?,?,?,?,1,?,?) RETURNING ItemID""",
        (ln.get("description"), ln.get("brand"), ln.get("model"), ln.get("list_price"),
         ln.get("ex_unit_cost"), ln.get("shipping_percent"), ln.get("unit_cost"), ln.get("u_price"),
         ln.get("u_price_sar"), repo.CATALOG_INITIAL_PRICE_DATE, src_file, now),
    )
    return cur.fetchone()["ItemID"]


def ingest_file(conn, path, stats):
    now = dt.datetime.now().isoformat(timespec="seconds")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        stats["errors"].append((os.path.basename(path), repr(e)[:90]))
        return

    boq_sheets = _detect_boq_sheets(wb)
    if not boq_sheets:
        wb.close()
        stats["skipped_no_boq"] += 1
        return

    # Idempotent by full path: re-ingesting the same file replaces its prior copy,
    # but the original creation date remains fixed.
    existing_project = conn.execute(
        "SELECT CreationDate FROM Projects_Master WHERE SourceFile=?", (path,)
    ).fetchone()
    preserved_creation_date = (
        _s(existing_project["CreationDate"]) if existing_project else None
    )
    conn.execute("DELETE FROM Projects_Master WHERE SourceFile=?", (path,))

    # project-level metadata from the first system's Quotation sheet
    first_suffix = system_suffix(boq_sheets[0])
    qmeta = parse_quotation_meta(wb, first_suffix)
    qterms = parse_quotation_terms(wb, first_suffix)
    terms_json = json.dumps(qterms) if qterms else None
    fname = os.path.basename(path)
    # CreationDate: Quotation date cell -> a date in the filename -> file mtime.
    qd = qmeta.get("date")
    if isinstance(qd, (dt.date, dt.datetime)):
        cdate = qd.date().isoformat() if isinstance(qd, dt.datetime) else qd.isoformat()
    else:
        cdate = _parse_date_text(qd) or _parse_date_text(fname)
    if not cdate:
        cdate = dt.datetime.fromtimestamp(os.path.getmtime(path)).date().isoformat()
    cdate = preserved_creation_date or cdate
    updated_date = dt.date.today().isoformat()
    project_name = str(qmeta.get("project") or os.path.splitext(fname)[0]).strip()

    # Offer number + revision: base ref from the Quotation 'Reference' (else the
    # filename), revision from a trailing 'R0x'; OfferNo written in the configured
    # revision format so an offer + its revisions group as one family.
    ref_raw = _s(qmeta.get("offer")) or (repo.parse_offer_ref(fname)[0] or "")
    base_ref = repo.base_name(ref_raw) if ref_raw else ""
    revno = repo.parse_revision_no(fname) or repo.parse_revision_no(ref_raw)
    if base_ref and revno:
        offer_no = f"{base_ref}{repo.revision_separator()}{repo.revision_token(revno)}"
    else:
        offer_no = base_ref or _s(qmeta.get("offer"))
    base_name_val = base_ref or None

    # Client: the Quotation 'M/S' line when present. Older 2023-template offers
    # have no M/S (only 'Attn:'), so fall back to the first 1-2 words of the
    # clean project name (user-chosen heuristic for migrated/legacy files).
    client = _s(qmeta.get("client"))
    if not client:
        proj_src = _s(qmeta.get("project")) or (repo.parse_client_from_name(fname) or "")
        words = proj_src.split()
        client = " ".join(words[:2]) if words else None

    cur = conn.execute(
        """INSERT INTO Projects_Master
             (ProjectName,ClientName,ContactName,ContactPhone,SalesPerson,OfferNo,
               CreationDate,UpdatedDate,RevisionNo,BaseName,SourceFile,IngestedAt,OfferTerms)
             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?) RETURNING ProjectID""",
        (project_name, client, _s(qmeta.get("contact")), _s(qmeta.get("phone")),
          _s(qmeta.get("sales")), offer_no, cdate, updated_date, revno or 0, base_name_val,
         path, now, terms_json),
    )
    pid = cur.fetchone()["ProjectID"]
    primary_discount = None
    primary_factor = None

    for sheet in boq_sheets:
        suffix = system_suffix(sheet)
        lines, meta = parse_boq_sheet(wb[sheet])
        if lines is None:
            stats["sheets_unparsed"] += 1
            continue
        f = (meta["factors"] + [None, None, None])[:3]
        scur = conn.execute(
            """INSERT INTO Project_Sheets
                 (ProjectID,SheetName,SystemSuffix,DiscountAmount,Factor1,Factor2,Factor3,SubtotalSAR)
               VALUES (?,?,?,?,?,?,?,?) RETURNING SheetID""",
            (pid, sheet, suffix, meta["discount"], f[0], f[1], f[2], meta.get("subtotal_sar")),
        )
        sid = scur.fetchone()["SheetID"]
        if primary_discount is None:
            primary_discount = meta["discount"]
            primary_factor = f[0]

        for ln in lines:
            if ln["_type"] == "discount":
                conn.execute(
                    """INSERT INTO Project_BoQ_Lines
                         (ProjectID,SheetID,RowOrder,Description,LineType,TPriceSAR)
                       VALUES (?,?,?,?,?,?)""",
                    (pid, sid, ln["_order"], "Discount", "discount", ln.get("t_price_sar")),
                )
                continue
            iid = upsert_item(conn, ln, path, now)
            conn.execute(
                """INSERT INTO Project_BoQ_Lines
                     (ProjectID,SheetID,ItemID,RowOrder,Area,System,Description,Brand,Model,
                      Qty,ListPriceUSD,ExUnitCostUSD,ShippingPercent,FinalUnitCostUSD,TotalCostUSD,
                      FinalUPriceUSD,TPriceUSD,FinalUPriceSAR,TPriceSAR,MarginExtra,LineType)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (pid, sid, iid, ln["_order"], _s(ln.get("area")), _s(ln.get("system")),
                 ln.get("description"), ln.get("brand"), ln.get("model"), ln.get("qty"),
                 ln.get("list_price"), ln.get("ex_unit_cost"), ln.get("shipping_percent"), ln.get("unit_cost"),
                 ln.get("total_cost"), ln.get("u_price"), ln.get("t_price"),
                 ln.get("u_price_sar"), ln.get("t_price_sar"), ln.get("margin"),
                 ln["_type"]),
            )
            if ln["_type"] == "spare":
                stats["spares"] += 1
            else:
                stats["lines"] += 1
        stats["sheets"] += 1

    conn.execute(
        "UPDATE Projects_Master SET DiscountAmount=?, ConversionFactor=? WHERE ProjectID=?",
        (primary_discount or 0, primary_factor, pid),
    )
    wb.close()
    stats["files"] += 1


def _s(v):
    if v is None:
        return None
    if isinstance(v, (dt.date, dt.datetime)):
        return v.isoformat()
    return str(v).strip()


def scan_workbooks(root):
    files = []
    for ext in ("*.xlsx", "*.xlsm"):
        files += glob.glob(os.path.join(root, "**", ext), recursive=True)
    return sorted(f for f in files if "~$" not in f)


def ingest_folder(root, progress=None):
    conn = dbmod.init_db()
    files = scan_workbooks(root)
    stats = {"workbooks_found": len(files), "files": 0, "sheets": 0, "lines": 0,
             "spares": 0, "skipped_no_boq": 0, "sheets_unparsed": 0, "errors": []}
    for i, f in enumerate(files, 1):
        ingest_file(conn, f, stats)
        conn.commit()
        if progress:
            progress(i, len(files), f)
    conn.commit()
    stats["catalogue_items"] = conn.execute("SELECT COUNT(*) FROM Items_Catalog").fetchone()[0]
    conn.close()
    return stats


def print_summary(root, stats):
    print(f"Scanning {stats['workbooks_found']} workbook(s) under:\n  {root}\n")
    print("\n=== INGEST SUMMARY ===")
    print(f"  Files ingested (had BOQ sheets) : {stats['files']}")
    print(f"  BOQ system-sheets parsed        : {stats['sheets']}")
    print(f"  Offer line items stored         : {stats['lines']}")
    print(f"  Spare (parked) items tagged     : {stats['spares']}")
    print(f"  Files skipped (no BOQ sheet)     : {stats['skipped_no_boq']}")
    print(f"  Sheets with no canonical header  : {stats['sheets_unparsed']}")
    print(f"  Distinct catalogue items         : {stats['catalogue_items']}")
    if stats["errors"]:
        print(f"  Errors ({len(stats['errors'])}):")
        for n, e in stats["errors"][:15]:
            print(f"    - {n}: {e}")


def main(root):
    stats = ingest_folder(root, progress=lambda i, total, _f: print(f"  ...{i}/{total}") if i % 10 == 0 else None)
    print_summary(root, stats)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_ROOT)
